import os
import matplotlib
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import json
import calendar
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import *
from datetime import timedelta, datetime, time, date
import holidays as hd
import numpy as np
import re
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas


# For Clickable Icons
class ClickableLabel(QtWidgets.QLabel):
    clicked = pyqtSignal()

    def mousePressEvent(self, event):
        self.clicked.emit()


class Ui_LoginWindow(object):
    def setupUi(self, LoginWindow):
        LoginWindow.setObjectName("LoginWindow")
        LoginWindow.resize(800, 600)
        LoginWindow.setWindowIcon(QtGui.QIcon("logo-logo.png"))
        self.login_window = QtWidgets.QWidget(LoginWindow)
        self.login_window.setStyleSheet("""background-color : qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:1, 
        stop:0 rgba(176, 0, 0, 255), stop:0.738636 rgba(255, 113, 250, 255))""")
        self.login_window.setObjectName("MainWindow")
        self.username = QtWidgets.QLineEdit(self.login_window)
        self.username.setGeometry(QtCore.QRect(310, 140, 171, 31))
        self.username.setAutoFillBackground(False)
        self.username.setStyleSheet("background-color: rgb(255, 255, 255); border-radius: 5px")
        self.username.setObjectName("username")
        self.password = QtWidgets.QLineEdit(self.login_window)
        self.password.setGeometry(QtCore.QRect(310, 230, 171, 31))
        self.password.setAutoFillBackground(False)
        self.password.setStyleSheet("background-color: rgb(255, 255, 255); border-radius: 5px")
        self.password.setObjectName("password")
        self.password.setEchoMode(QtWidgets.QLineEdit.Password)  # Make the input hidden
        self.login_btn = QtWidgets.QPushButton(self.login_window)
        self.login_btn.setGeometry(QtCore.QRect(360, 340, 75, 23))
        self.login_btn.setStyleSheet("\n"
                                     "color: rgb(0, 0, 0);\n"
                                     "background-color: rgb(255, 255, 255); \n"
                                     "border-radius: 10px;")
        self.login_btn.setObjectName("Login")
        self.login_btn.clicked.connect(self.login)

        LoginWindow.setCentralWidget(self.login_window)
        self.retranslateUi(LoginWindow)

        QtCore.QMetaObject.connectSlotsByName(LoginWindow)

    def retranslateUi(self, LoginWindow):
        _translate = QtCore.QCoreApplication.translate
        LoginWindow.setWindowTitle(_translate("LoginWindow", "MBPI"))
        self.login_btn.setText(_translate("LoginWindow", "Login"))
        self.login_btn.setShortcut('Return')

    def login(self):
        import psycopg2

        username = self.username.text()
        pass1 = self.password.text()

        # Connect to the Database
        try:
            self.conn = psycopg2.connect(
                host="192.168.1.13",
                port=5432,
                dbname='postgres',
                user=f'postgres',
                password=f'mbpi'
            )
            self.cursor = self.conn.cursor()

            self.cursor.execute(f"""
            SELECT * FROM users 
            WHERE username = '{username}' AND password = '{pass1}'

            """)

            account = self.cursor.fetchone()

            if account:

                global extruder_access
                extruder_access = account[5]
                global qc_access
                qc_access = account[6]
                global warehouse_access
                warehouse_access = account[7]

                self.username.deleteLater()
                self.password.deleteLater()
                self.login_btn.deleteLater()
                pass
            else:
                QMessageBox.information(self.login_window, "INVALID CREDENTIALS", "Check your Username and Password")
                return


        except psycopg2.Error:
            QMessageBox.information(self.login_window, "INVALID CREDENTIALS", "Check your Username and Password")
            return

        # Get log info
        self.cursor.execute(f"""
        INSERT INTO logs
        VALUES('{username}', '{os.environ['COMPUTERNAME']}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}')

        """)

        self.conn.commit()

        self.launch_main()

    # This is the main window after login screen
    def launch_main(self):

        def show_production_tabs():

            def mixer_clicked():
                self.mixer_btn_clicked = True
                self.production()

            try:
                if self.qc_tab_showed == True:
                    # Hide the Buttons
                    self.qc_data_btn.hide()
                    self.qc_entry_btn.hide()
                    self.dashboard_btn.hide()
                    self.returns_btn.hide()
                    self.dropdown_qc.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))
                    self.qc_tab_showed = False
                elif self.warehouse_tabs_showed == True:
                    self.fgIncoming_btn.hide()
                    self.fgOutgoing_btn.hide()
                    self.dropdown_wh.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))
                    self.warehouse_tabs_showed = False
                else:
                    pass
            except:
                pass

            if self.production_tabs_showed == False:
                self.production_tabs_showed = True
                self.qualityControl_btn.move(20, 240)
                self.warehouse_btn.move(20, 300)
                self.dropdown_production.setPixmap(QtGui.QIcon('up.png').pixmap(20, 20))

                # Create Buttons for Tabs Sections
                self.extruder_btn = QPushButton(self.login_window)
                self.extruder_btn.setGeometry(60, 180, 150, 30)
                self.extruder_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.extruder_btn.setText('Extruder')
                self.extruder_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.extruder_btn.setCursor(Qt.PointingHandCursor)
                self.extruder_btn.clicked.connect(self.production)
                self.extruder_btn.show()

                self.mixer_btn = QPushButton(self.login_window)
                self.mixer_btn.setGeometry(60, 210, 150, 30)
                self.mixer_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.mixer_btn.setText('Mixer')
                self.mixer_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.mixer_btn.setCursor(Qt.PointingHandCursor)
                self.mixer_btn_clicked = False
                self.mixer_btn.clicked.connect(mixer_clicked)
                self.mixer_btn.show()

            else:
                self.production_tabs_showed = False
                self.qualityControl_btn.move(20, 200)
                self.warehouse_btn.move(20, 260)
                self.extruder_btn.hide()
                self.mixer_btn.hide()
                self.dropdown_production.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))

            self.qualityControl_btn.show()
            self.warehouse_btn.show()
            self.dropdown_production.show()

        def show_qc_tabs():
            # Hide other tabs opened
            def qc_clicked():
                self.quality_control()

            def qc_entry_clicked():
                self.qc_entry_clicked_status = True
                self.quality_control()

            def dashboard_clicked():
                self.dashboard_clicked_status = True
                self.quality_control()

            def returns_clicked():
                self.returns_clicked_status = True
                self.quality_control()

            try:
                if self.production_tabs_showed == True:
                    self.mixer_btn.hide()
                    self.extruder_btn.hide()
                    self.qualityControl_btn.move(20, 200)
                    self.warehouse_btn.move(20, 260)
                    self.dropdown_production.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))
                    self.dropdown_production.show()
                    self.production_tabs_showed = False
                elif self.warehouse_tabs_showed == True:
                    self.fgIncoming_btn.hide()
                    self.fgOutgoing_btn.hide()
                    self.dropdown_wh.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))
                    self.warehouse_tabs_showed = False
            except:
                pass

            if self.qc_tab_showed == False:
                self.qc_tab_showed = True
                self.warehouse_btn.move(20, 360)
                self.dropdown_qc.setPixmap(QtGui.QIcon('up.png').pixmap(20, 20))

                # Create Buttons for Tabs Sections
                self.qc_data_btn = QPushButton(self.login_window)
                self.qc_data_btn.setGeometry(60, 240, 150, 30)
                self.qc_data_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.qc_data_btn.setText('QC Data')
                self.qc_data_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.qc_data_btn.setCursor(Qt.PointingHandCursor)
                self.qc_data_btn.clicked.connect(qc_clicked)
                self.qc_data_btn.show()

                self.qc_entry_btn = QPushButton(self.login_window)
                self.qc_entry_btn.setGeometry(60, 270, 150, 30)
                self.qc_entry_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.qc_entry_btn.setText('QC Entry')
                self.qc_entry_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.qc_entry_btn.setCursor(Qt.PointingHandCursor)
                self.qc_entry_clicked_status = False
                self.qc_entry_btn.clicked.connect(qc_entry_clicked)
                self.qc_entry_btn.show()

                self.dashboard_btn = QPushButton(self.login_window)
                self.dashboard_btn.setGeometry(60, 300, 150, 30)
                self.dashboard_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.dashboard_btn.setText('Analytics')
                self.dashboard_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.dashboard_btn.setCursor(Qt.PointingHandCursor)
                self.dashboard_clicked_status = False
                self.dashboard_btn.clicked.connect(dashboard_clicked)
                self.dashboard_btn.show()

                self.returns_btn = QPushButton(self.login_window)
                self.returns_btn.setGeometry(60, 330, 150, 30)
                self.returns_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.returns_btn.setText('Returns')
                self.returns_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.returns_btn.setCursor(Qt.PointingHandCursor)
                self.returns_clicked_status = False
                self.returns_btn.clicked.connect(returns_clicked)
                self.returns_btn.show()

            else:
                self.qc_tab_showed = False
                self.warehouse_btn.move(20, 260)
                self.dropdown_qc.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))

                # Hide the Buttons
                self.qc_data_btn.hide()
                self.qc_entry_btn.hide()
                self.dashboard_btn.hide()
                self.returns_btn.hide()

            self.warehouse_btn.show()
            self.dropdown_production.show()

        def show_warehouse_tabs():

            def fg_outgoing_clicked():
                self.fgOutgoing_btn_clicked_status = True
                self.warehouse()

            try:
                if self.production_tabs_showed == True:
                    self.mixer_btn.hide()
                    self.extruder_btn.hide()
                    self.qualityControl_btn.move(20, 200)
                    self.warehouse_btn.move(20, 260)
                    self.dropdown_production.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))
                    self.dropdown_production.show()
                    self.production_tabs_showed = False
                elif self.qc_tab_showed == True:
                    # Hide the Buttons
                    self.qc_data_btn.hide()
                    self.qc_entry_btn.hide()
                    self.dashboard_btn.hide()
                    self.returns_btn.hide()
                    self.dropdown_qc.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))
                    self.qc_tab_showed = False
                else:
                    pass

            except:
                pass

            if self.warehouse_tabs_showed == False:
                self.warehouse_tabs_showed = True
                self.dropdown_wh.setPixmap(QtGui.QIcon('up.png').pixmap(20, 20))
                self.warehouse_btn.move(20, 260)

                # Create Dropdown Buttons
                self.fgIncoming_btn = QPushButton(self.login_window)
                self.fgIncoming_btn.setGeometry(60, 300, 150, 30)
                self.fgIncoming_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.fgIncoming_btn.setText('FG Incoming')
                self.fgIncoming_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.fgIncoming_btn.setCursor(Qt.PointingHandCursor)
                self.fgIncoming_btn.clicked.connect(self.warehouse)
                self.fgIncoming_btn.show()

                self.fgOutgoing_btn = QPushButton(self.login_window)
                self.fgOutgoing_btn.setGeometry(60, 330, 150, 30)
                self.fgOutgoing_btn.setStyleSheet(
                    'QPushButton{ border: none; text-align: left; padding-left: 40px; background-color: rgb(217, 218, 221)}'
                    'QPushButton:hover{background-color: yellow}')
                self.fgOutgoing_btn.setText('FG Outgoing')
                self.fgOutgoing_btn.setFont(QtGui.QFont("Segoe UI", 12))
                self.fgOutgoing_btn.setCursor(Qt.PointingHandCursor)
                self.fgOutgoing_btn_clicked_status = False
                self.fgOutgoing_btn.clicked.connect(fg_outgoing_clicked)
                self.fgOutgoing_btn.show()

            else:
                self.warehouse_tabs_showed = False
                self.warehouse_btn.move(20, 260)
                self.dropdown_wh.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))

                self.fgOutgoing_btn.hide()
                self.fgIncoming_btn.hide()

            self.warehouse_btn.show()

        LoginWindow.move(75, 0)
        LoginWindow.setFixedSize(1200, 700)
        self.login_window.setStyleSheet("background-color: rgb(239, 243, 254);")
        self.main_widget = QtWidgets.QWidget(self.login_window)
        self.main_widget.setStyleSheet("""
        background-color: rgb(240,240,240);     

        """)
        self.main_widget.setGeometry(210, 0, 991, 700)
        self.main_widget.show()

        self.production_btn = QtWidgets.QPushButton(self.login_window)
        self.production_btn.setGeometry(20, 140, 190, 40)
        self.production_btn.setCursor(Qt.PointingHandCursor)
        self.production_btn.setStyleSheet("""
        background-color: white; 
        border:none; 
        border-top-left-radius: 20px;
        border-bottom-left-radius: 20px;

        """)
        self.production_btn.clicked.connect(self.production)
        self.production_btn.show()

        self.production_lbl = QtWidgets.QLabel(self.production_btn)
        self.production_lbl.setText("Production")
        self.production_lbl.setGeometry(70, 5, 100, 30)
        self.production_lbl.setFont(QtGui.QFont("Segoe UI", 12))
        self.production_lbl.setStyleSheet("color: gray;")
        self.production_lbl.setCursor(Qt.PointingHandCursor)
        self.production_lbl.show()

        self.production_icon = ClickableLabel(self.production_btn)
        self.production_icon.setGeometry(0, 0, 40, 40)
        self.production_icon.setPixmap(QtGui.QIcon('production_icon.png').pixmap(40, 40))  # Set icon
        self.production_icon.setScaledContents(True)  # Scale icon to fit the label
        self.production_icon.setCursor(Qt.PointingHandCursor)
        self.production_icon.clicked.connect(self.production)
        self.production_icon.show()

        self.production_tabs_showed = False
        self.dropdown_production = ClickableLabel(self.production_btn)
        self.dropdown_production.setGeometry(165, 0, 40, 40)
        self.dropdown_production.setPixmap(
            QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))  # Set icon# Scale icon to fit the label
        self.dropdown_production.setCursor(Qt.PointingHandCursor)
        self.dropdown_production.clicked.connect(show_production_tabs)

        self.dropdown_production.show()

        self.qualityControl_btn = QtWidgets.QPushButton(self.login_window)
        self.qualityControl_btn.setGeometry(20, 200, 190, 40)
        self.qualityControl_btn.setCursor(Qt.PointingHandCursor)
        self.qualityControl_btn.setStyleSheet("""
                border-top-left-radius: 20px; 
                border-bottom-left-radius: 20px; 
                background-color: white;
                """)
        self.qualityControl_btn.clicked.connect(self.quality_control)
        self.qualityControl_btn.show()

        self.qc_icon = ClickableLabel(self.qualityControl_btn)
        self.qc_icon.setGeometry(0, 0, 40, 40)
        self.qc_icon.setPixmap(QtGui.QIcon('qc-icon.png').pixmap(40, 40))  # Set icon
        self.qc_icon.setScaledContents(True)  # Scale icon to fit the label
        self.qc_icon.setCursor(Qt.PointingHandCursor)
        self.qc_icon.clicked.connect(self.quality_control)
        self.qc_icon.show()

        self.qualityControl_lbl = QtWidgets.QLabel(self.qualityControl_btn)
        self.qualityControl_lbl.setText("Quality Control")
        self.qualityControl_lbl.setGeometry(50, 5, 120, 30)
        self.qualityControl_lbl.setFont(QtGui.QFont("Segoe UI", 12))
        self.qualityControl_lbl.setStyleSheet("color: gray;")
        self.qualityControl_lbl.setCursor(Qt.PointingHandCursor)
        self.qualityControl_lbl.show()

        self.dropdown_qc = ClickableLabel(self.qualityControl_btn)
        self.dropdown_qc.setGeometry(165, 0, 40, 40)
        self.dropdown_qc.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))  # Set icon
        self.dropdown_qc.setCursor(Qt.PointingHandCursor)
        self.dropdown_qc.clicked.connect(show_qc_tabs)
        self.qc_tab_showed = False
        self.dropdown_qc.show()

        self.warehouse_btn = QtWidgets.QPushButton(self.login_window)
        self.warehouse_btn.setGeometry(20, 260, 190, 40)
        self.warehouse_btn.setCursor(Qt.PointingHandCursor)
        self.warehouse_btn.setStyleSheet("""
                        border-top-left-radius: 20px; 
                        border-bottom-left-radius: 20px; 
                        background-color: white;
                        """)
        self.warehouse_btn.clicked.connect(self.warehouse)

        self.warehouse_btn.show()

        self.warehouse_icon = ClickableLabel(self.warehouse_btn)
        self.warehouse_icon.setGeometry(0, 0, 40, 40)
        self.warehouse_icon.setPixmap(QtGui.QIcon('warehouse2.png').pixmap(40, 40))
        self.warehouse_icon.setScaledContents(True)  # Scale icon to fit the label
        self.warehouse_icon.setCursor(Qt.PointingHandCursor)
        self.warehouse_icon.show()

        self.warehouse_lbl = QLabel(self.warehouse_btn)
        self.warehouse_lbl.setGeometry(50, 5, 120, 30)
        self.warehouse_lbl.setText("Warehouse")
        self.warehouse_lbl.setFont(QtGui.QFont("Segoe UI", 12))
        self.warehouse_lbl.setStyleSheet("color: gray;")
        self.warehouse_lbl.show()

        self.dropdown_wh = ClickableLabel(self.warehouse_btn)
        self.dropdown_wh.setGeometry(165, 0, 40, 40)
        self.dropdown_wh.setPixmap(QtGui.QIcon('icons8-dropdown-48.png').pixmap(20, 20))  # Set icon
        self.dropdown_wh.setCursor(Qt.PointingHandCursor)
        self.warehouse_tabs_showed = False
        self.dropdown_wh.clicked.connect(show_warehouse_tabs)
        self.dropdown_wh.show()

        self.logo = QLabel(self.login_window)
        self.logo.setGeometry(23, 10, 170, 121)
        pixmap = QtGui.QPixmap('MBPI-ADJUST.png')
        pixmap = pixmap.scaled(170, 121)
        self.logo.setPixmap(pixmap)
        self.logo.show()

        self.production()

    def production(self):
        # Delete If there are existing Widgets
        try:
            self.info_widget.deleteLater()
            self.temp_table.deleteLater()
            self.time_table.deleteLater()
            self.material_table.deleteLater()
            self.group_box.deleteLater()
            self.export_btn.deleteLater()
            print("Widgets Cleared")

        except Exception as e:
            print(e)

        def show_form():

            try:
                selected = self.extruder_table.selectedItems()
                selected = [i.text() for i in selected]

                # Query the whole columns
                self.cursor.execute(f"SELECT * FROM extruder WHERE process_id = {selected[0]}")
                selected = self.cursor.fetchall()

                # Clear all the widget first
                self.extruder_table.deleteLater()
                self.view_btn.deleteLater()
                self.add_btn.deleteLater()
                self.update_btn.deleteLater()
                self.print_btn.deleteLater()
                main_time_table.deleteLater()
                material_table.deleteLater()
                lotNumber_table.deleteLater()

            except:
                QMessageBox.information(self.production_widget, "ERROR", "No Selected Item")
                return

            selected = selected[0]

            # Unpack all the items for convenience
            machine, ordered_qty, product_output, customer = selected[1:5]
            formula_id, product_code, order_id, total_time, time_start = selected[5:10]
            time_end, output_percent, loss, loss_percent, purging = selected[10:15]
            resin, remarks, screw_config, feed_rate, rpm = selected[15:20]
            screen_size, operator, supervisor, materials, temperature = selected[20:25]
            purge_duration, outputs, output_per_hour = selected[25:28]
            production_ID = selected[28]
            total_input = selected[29]

            def exportToExcel():

                process_id = selected[0]

                self.cursor.execute(f"""
                           SELECT * FROM extruder 
                           WHERE process_id = '{process_id}';

                           """)
                items = self.cursor.fetchall()[0]
                # Unpack the Items
                machine_number = items[1]
                quantity_order = items[2]
                customer = items[4]
                code = items[6]
                total_input = items[-2]
                total_time = items[8]
                time_start = items[9]
                time_end = items[10]
                outputPerHour = items[27]
                total_output = items[3]
                outputPercent = items[11]
                loss = items[12]
                lossPercent = items[13]
                purging = items[14]
                resin = items[15]
                remarks = items[16]
                operator = items[21]
                supervisor = items[22]
                outputs = items[-7]
                materials = items[-10]
                lot_number = items[30]
                purge_duration = time(hour=items[-8])

                # Loading the Extruder Template
                wb = load_workbook(
                    r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist\Extruder Template.xlsx")
                worksheet = wb.active

                font = Font(size=8, bold=True, name='Arial')
                center_Alignment = Alignment(horizontal='center', vertical='center')

                worksheet["F5"] = "Extruder Machine No. " + machine_number[-1]
                worksheet["A8"] = machine_number[-1]
                worksheet["B8"] = quantity_order  # quantity order
                worksheet["C8"].font = font
                worksheet["C8"].alignment = center_Alignment
                worksheet["C8"] = customer  # customer

                worksheet["F8"] = code  # product code
                worksheet["G9"] = total_input  # total input
                worksheet["H9"] = total_time  # total time used
                worksheet["I9"] = outputPerHour  # output Per Hour
                worksheet["K9"] = total_output  # total Output
                worksheet["L9"] = outputPercent  # Total Output Percentage
                worksheet["M9"] = loss
                worksheet["N9"] = lossPercent

                total_sec = timedelta()
                for row in range(len(time_start)):
                    worksheet["A" + str(12 + row)] = time_start[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["D" + str(12 + row)] = time_end[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["F" + str(12 + row)] = time_end[row] - time_start[row]
                    worksheet["G" + str(12 + row)] = outputs[row]
                    total_sec = total_sec + (time_end[row] - time_start[row])

                try:
                    hour = str(int(total_sec.total_seconds() // 3600))
                    minute = str((int(total_sec.total_seconds() % 3600) // 60))

                    total_time_used = time(int(hour), int(minute))

                    worksheet["F25"] = total_time_used
                except ValueError:
                    worksheet["F25"] = hour + ":" + minute

                for key in list(materials.keys()):
                    worksheet["I" + str(12 + list(materials.keys()).index(key))] = key
                    worksheet["K" + str(12 + list(materials.keys()).index(key))] = materials[key]

                for ln in range(len(lot_number)):
                    worksheet["M" + str(12 + ln)] = lot_number[ln]

                worksheet["B27"] = purging
                worksheet["E28"] = purge_duration
                worksheet["B29"] = resin
                worksheet["G26"] = remarks
                worksheet["M28"] = operator
                worksheet["M29"] = supervisor

                # Save the Workbook
                wb.save(r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist\text.xlsx")
                print("load successful")

            # Convert string of json to JSON
            materials = str(materials).replace("'", '"')
            materials = json.loads(materials)

            # Main Widget
            self.info_widget = QtWidgets.QWidget(self.production_widget)
            self.info_widget.setGeometry(20, 0, 951, 450)
            self.info_widget.setStyleSheet("background-color: rgb(0,109,189);")
            self.info_widget.show()

            # Set Font Style and Size
            label_stylesheet = "color: rgb(195, 164, 86)"
            font = QtGui.QFont("Arial", 14)  # Set Font for Labels

            # Create 3 widgets for division
            info_widget1 = QtWidgets.QWidget(self.info_widget)
            info_widget1.setGeometry(0, 80, 158, 380)

            info_widget2 = QtWidgets.QWidget(self.info_widget)
            info_widget2.setGeometry(158, 80, 158, 380)

            info_widget3 = QtWidgets.QWidget(self.info_widget)
            info_widget3.setGeometry(316, 80, 158, 380)

            info_widget4 = QtWidgets.QWidget(self.info_widget)
            info_widget4.setGeometry(474, 80, 158, 380)

            info_widget5 = QtWidgets.QWidget(self.info_widget)
            info_widget5.setGeometry(632, 80, 158, 380)

            info_widget6 = QtWidgets.QWidget(self.info_widget)
            info_widget6.setGeometry(790, 80, 158, 380)

            # create Vertical Layouts
            left_vertical_layout = QVBoxLayout()

            self.ordered_company = QtWidgets.QLabel(self.info_widget)
            self.ordered_company.setFont(QtGui.QFont("Arial", 30))
            self.ordered_company.setGeometry(0, 30, 950, 50)
            self.ordered_company.setText(customer)
            self.ordered_company.setAlignment(Qt.AlignCenter)
            self.ordered_company.show()

            # Extruder Label
            self.machine_label = QtWidgets.QLabel(self.info_widget)
            self.machine_label.setText("Extruder:")
            self.machine_label.setStyleSheet(label_stylesheet)
            self.machine_label.setFont(font)

            # Show Extruder Value
            self.extruder_val = QtWidgets.QLabel(self.info_widget)
            self.extruder_val.setText(machine)
            self.extruder_val.setFont(font)

            # Product Code Label
            self.code_label = QtWidgets.QLabel(self.info_widget)
            self.code_label.setText("Product Code:")
            self.code_label.setFont(font)
            self.code_label.setStyleSheet(label_stylesheet)

            # Show Product Code Value
            self.product_code_val = QtWidgets.QLabel(self.info_widget)
            self.product_code_val.setText(product_code)
            self.product_code_val.setFont(font)

            # Quantity Order Label
            self.order_label = QtWidgets.QLabel(self.info_widget)
            self.order_label.setText("Quantity Order:")
            self.order_label.setFont(font)
            self.order_label.setStyleSheet(label_stylesheet)

            # Show Order Value
            self.order_val = QtWidgets.QLabel(self.info_widget)
            self.order_val.setText(str(ordered_qty))
            self.order_val.setFont(font)

            # Show Output Label
            self.output_label = QtWidgets.QLabel(self.info_widget)
            self.output_label.setText("Output:")
            self.output_label.setFont(font)
            self.output_label.setStyleSheet(label_stylesheet)

            # Show Output Value
            self.output_val = QtWidgets.QLabel(self.info_widget)
            self.output_val.setText(str(product_output))
            self.output_val.setFont(font)

            # Show formula Label
            self.formula_label = QtWidgets.QLabel(self.info_widget)
            self.formula_label.setText("Formula ID:")
            self.formula_label.setFont(font)
            self.formula_label.setStyleSheet(label_stylesheet)

            # Show Formula ID Value
            self.formulaID_val = QtWidgets.QLabel(self.info_widget)
            self.formulaID_val.setText(str(formula_id))
            self.formulaID_val.setFont(font)

            self.resin_label = QtWidgets.QLabel(self.info_widget)
            self.resin_label.setText("Resin:")
            self.resin_label.setFont(font)
            self.resin_label.setStyleSheet(label_stylesheet)

            # Show Resin Value
            self.resin_val = QtWidgets.QLabel(self.info_widget)
            self.resin_val.setText(str(resin))
            self.resin_val.setFont(font)

            # Lot Label
            self.lot_label = QtWidgets.QLabel(self.info_widget)
            self.lot_label.setText("LOT Number:")
            self.lot_label.setFont(font)
            self.lot_label.setStyleSheet(label_stylesheet)

            # Show Lot Number Value
            self.lotNum_val = QtWidgets.QLabel(self.info_widget)
            self.lotNum_val.setText(str(resin))
            self.lotNum_val.setFont(font)

            self.feedrate_label = QtWidgets.QLabel(self.info_widget)
            self.feedrate_label.setText("Feed Rate:")
            self.feedrate_label.setFont(font)
            self.feedrate_label.setStyleSheet(label_stylesheet)

            # Show Feed Rate Value
            self.feedrate_val = QtWidgets.QLabel(self.info_widget)
            self.feedrate_val.setText(str(feed_rate))
            self.feedrate_val.setFont(font)

            # RPM label
            self.rpm_label = QtWidgets.QLabel(self.info_widget)
            self.rpm_label.setText("RPM:")
            self.rpm_label.setFont(font)
            self.rpm_label.setStyleSheet(label_stylesheet)

            # Show RPM Value
            self.rpm_val = QtWidgets.QLabel(self.info_widget)
            self.rpm_val.setText(str(rpm))
            self.rpm_val.setFont(font)

            # Screen Size Label
            self.screen_size_label = QtWidgets.QLabel(self.info_widget)
            self.screen_size_label.setFont(font)
            self.screen_size_label.setText("Screen Size:")
            self.screen_size_label.setStyleSheet(label_stylesheet)

            # Show Screen Size Value
            self.screenSize_val = QtWidgets.QLabel(self.info_widget)
            self.screenSize_val.setText(str(screen_size))
            self.screenSize_val.setFont(font)

            self.screwconfig_label = QtWidgets.QLabel(self.info_widget)
            self.screwconfig_label.setFont(font)
            self.screwconfig_label.setText("Screw Config:")
            self.screwconfig_label.setStyleSheet(label_stylesheet)

            # Show Screw Config Value
            self.screwConf_val = QtWidgets.QLabel(self.info_widget)
            self.screwConf_val.setText(str(screw_config))
            self.screwConf_val.setFont(font)

            # Output % Label
            self.output_percentage_lbl = QtWidgets.QLabel(self.info_widget)
            self.output_percentage_lbl.setText("Output %:")
            self.output_percentage_lbl.setFont(font)
            self.output_percentage_lbl.setStyleSheet(label_stylesheet)

            # Show Output Percentage Value
            self.outputPercent_val = QtWidgets.QLabel(self.info_widget)
            self.outputPercent_val.setText(str(output_percent))
            self.outputPercent_val.setFont(font)

            # Loss Label
            self.loss_label = QtWidgets.QLabel(self.info_widget)
            self.loss_label.setText("Loss:")
            self.loss_label.setFont(QtGui.QFont(font))
            self.loss_label.setStyleSheet(label_stylesheet)

            # Show Loss Value
            self.loss_val = QtWidgets.QLabel(self.info_widget)
            self.loss_val.setText(str(loss))
            self.loss_val.setFont(font)

            # loss percentage Label
            self.loss_percent_label = QtWidgets.QLabel(self.info_widget)
            self.loss_percent_label.setText("Loss %:")
            self.loss_percent_label.setFont(font)
            self.loss_percent_label.setStyleSheet(label_stylesheet)

            # Show Loss Percentage Value
            self.lossPercent_val = QtWidgets.QLabel(self.info_widget)
            self.lossPercent_val.setText(str(loss_percent))
            self.lossPercent_val.setFont(font)

            # Purge Duration Label
            self.purge_duration_label = QtWidgets.QLabel(self.info_widget)
            self.purge_duration_label.setText("Purge Duration:")
            self.purge_duration_label.setFont(font)
            self.purge_duration_label.setStyleSheet(label_stylesheet)

            # Show Resin Value
            self.purgeDuration_val = QtWidgets.QLabel(self.info_widget)
            self.purgeDuration_val.setText(str(purge_duration))
            self.purgeDuration_val.setFont(font)

            # Production ID
            self.productionID_value = QtWidgets.QLabel(self.info_widget)
            self.productionID_value.setText(str(production_ID))
            self.productionID_value.setFont(font)

            self.productionID_label = QtWidgets.QLabel(self.info_widget)
            self.productionID_label.setText("Production ID:")
            self.productionID_label.setFont(font)
            self.productionID_label.setStyleSheet(label_stylesheet)

            # Order ID
            self.orderID_label = QtWidgets.QLabel(self.info_widget)
            self.orderID_label.setText("Order ID:")
            self.orderID_label.setFont(font)
            self.orderID_label.setStyleSheet(label_stylesheet)

            self.orderID_value = QtWidgets.QLabel(self.info_widget)
            self.orderID_value.setText(str(order_id))
            self.orderID_value.setFont(font)

            # Total Time
            self.total_time = QtWidgets.QLabel(self.info_widget)
            self.total_time.setText("Total Hours:")
            self.total_time.setFont(font)
            self.total_time.setStyleSheet(label_stylesheet)

            self.total_time_value = QtWidgets.QLabel(self.info_widget)
            self.total_time_value.setText(str(total_time))
            self.total_time_value.setFont(font)

            # Output Per Hour
            self.outputPerHour_label = QtWidgets.QLabel(self.info_widget)
            self.outputPerHour_label.setText("Output / Hour:")
            self.outputPerHour_label.setFont(font)
            self.outputPerHour_label.setStyleSheet(label_stylesheet)

            self.outputPerHour_val = QtWidgets.QLabel(self.info_widget)
            self.outputPerHour_val.setText(str(output_per_hour))
            self.outputPerHour_val.setFont(font)

            # Total Input
            self.total_input = QtWidgets.QLabel(self.info_widget)
            self.total_input.setFont(font)
            self.total_input.setText(str(total_input))

            self.totalInput_label = QtWidgets.QLabel(self.info_widget)
            self.totalInput_label.setFont(font)
            self.totalInput_label.setText("Total Input:")
            self.totalInput_label.setStyleSheet(label_stylesheet)

            # Purging Labels
            self.purging_label = QtWidgets.QLabel(self.info_widget)
            self.purging_label.setText("Purging To:")
            self.purging_label.setFont(font)
            self.purging_label.setStyleSheet(label_stylesheet)

            self.purging_val = QtWidgets.QLabel(self.info_widget)
            self.purging_val.setFont(font)
            self.purging_val.setText(purging)

            # Operator Labels
            self.operator_label = QtWidgets.QLabel(self.info_widget)
            self.operator_label.setFont(font)
            self.operator_label.setText("Operator:")
            self.operator_label.setStyleSheet(label_stylesheet)

            self.operator_value = QtWidgets.QLabel(self.info_widget)
            self.operator_value.setFont(font)
            self.operator_value.setText(operator)

            # Supervisor Labels
            self.supervisor_label = QtWidgets.QLabel(self.info_widget)
            self.supervisor_label.setText("Supervisor")
            self.supervisor_label.setFont(font)
            self.supervisor_label.setStyleSheet(label_stylesheet)

            self.supervisor_value = QtWidgets.QLabel(self.info_widget)
            self.supervisor_value.setText(supervisor)
            self.supervisor_value.setFont(font)

            info_vbox1 = QVBoxLayout(info_widget1)
            info_vbox2 = QVBoxLayout(info_widget2)
            info_vbox3 = QVBoxLayout(info_widget3)
            info_vbox4 = QVBoxLayout(info_widget4)
            info_vbox5 = QVBoxLayout(info_widget5)
            info_vbox6 = QVBoxLayout(info_widget6)

            # First VBOX for Labels
            info_vbox1.addWidget(self.machine_label)
            info_vbox1.addWidget(self.productionID_label)
            info_vbox1.addWidget(self.code_label)
            info_vbox1.addWidget(self.orderID_label)
            info_vbox1.addWidget(self.formula_label)
            info_vbox1.addWidget(self.order_label)
            info_vbox1.addWidget(self.total_time)

            # Second VBOX for Inputs
            info_vbox2.addWidget(self.extruder_val)
            info_vbox2.addWidget(self.productionID_value)
            info_vbox2.addWidget(self.product_code_val)
            info_vbox2.addWidget(self.orderID_value)
            info_vbox2.addWidget(self.formulaID_val)
            info_vbox2.addWidget(self.order_val)
            info_vbox2.addWidget(self.total_time_value)

            # third VBOX for Second Label
            info_vbox3.addWidget(self.totalInput_label)
            info_vbox3.addWidget(self.output_label)
            info_vbox3.addWidget(self.output_percentage_lbl)
            info_vbox3.addWidget(self.outputPerHour_label)
            info_vbox3.addWidget(self.loss_label)
            info_vbox3.addWidget(self.loss_percent_label)
            info_vbox3.addWidget(self.feedrate_label)
            info_vbox3.addWidget(self.rpm_label)

            # Fourth Vbox for the Second Row Inputs Value
            info_vbox4.addWidget(self.total_input)
            info_vbox4.addWidget(self.output_val)
            info_vbox4.addWidget(self.outputPercent_val)
            info_vbox4.addWidget(self.outputPerHour_val)
            info_vbox4.addWidget(self.loss_val)
            info_vbox4.addWidget(self.lossPercent_val)
            info_vbox4.addWidget(self.feedrate_val)
            info_vbox4.addWidget(self.rpm_val)

            # Fifth VBOX for Labels
            info_vbox5.addWidget(self.purging_label)
            info_vbox5.addWidget(self.resin_label)
            info_vbox5.addWidget(self.screwconfig_label)
            info_vbox5.addWidget(self.screen_size_label)
            info_vbox5.addWidget(self.purge_duration_label)
            info_vbox5.addWidget(self.operator_label)
            info_vbox5.addWidget(self.supervisor_label)

            info_vbox6.addWidget(self.purging_val)
            info_vbox6.addWidget(self.resin_val)
            info_vbox6.addWidget(self.screwConf_val)
            info_vbox6.addWidget(self.screenSize_val)
            info_vbox6.addWidget(self.purgeDuration_val)
            info_vbox6.addWidget(self.operator_value)
            info_vbox6.addWidget(self.supervisor_value)

            info_widget1.show()
            info_widget2.show()
            info_widget3.show()
            info_widget4.show()
            info_widget5.show()
            info_widget6.show()

            # Create 3 tables for Time, Materials and Temperature
            try:
                self.time_table = QtWidgets.QTableWidget(self.production_widget)
                self.time_table.setGeometry(20, 450, 330, 300)
                self.time_table.setColumnCount(3)
                self.time_table.setRowCount(len(time_start) + 2)
                self.time_table.setRowCount(9)
                self.time_rows = len(time_end)
                self.time_cols = 3
                self.time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Time Diff"])
                total_timediff = timedelta()

                for i in range(self.time_rows):
                    for j in range(self.time_cols):
                        if j == 0:
                            item = QtWidgets.QTableWidgetItem(
                                time_start[i].strftime("%b-%d-%Y %H:%M"))  # Convert to string
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                            self.time_table.setItem(i, j, item)
                        elif j == 1:
                            item = QtWidgets.QTableWidgetItem(
                                time_end[i].strftime("%b-%d-%Y %H:%M"))  # Convert to string
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                            self.time_table.setItem(i, j, item)
                        else:
                            datetime1 = time_start[i]
                            datetime2 = time_end[i]
                            t_diff = datetime2 - datetime1
                            # Convert timedelta to a string representation (e.g., "X days, HH:MM:SS")
                            total_timediff = total_timediff + t_diff
                            t_diff_str = str(t_diff)
                            item = QtWidgets.QTableWidgetItem(t_diff_str)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            self.time_table.setItem(i, j, item)

                # Convert it to Hours and Minutes only
                total_hours = total_timediff.days * 24 + total_timediff.seconds // 3600
                total_minutes = (total_timediff.seconds % 3600) // 60

                # Populate the Time Table
                self.time_table.setItem(self.time_rows + 1, 1, QtWidgets.QTableWidgetItem("Total"))
                self.time_table.setItem(self.time_rows + 1, 2,
                                        QtWidgets.QTableWidgetItem(str(total_hours) + ":" + str(total_minutes)))
                self.time_table.show()

                # Material Table
                self.material_table = QtWidgets.QTableWidget(self.production_widget)
                self.material_table.setRowCount(len(materials))
                self.material_table.setColumnCount(2)
                self.material_table.setRowCount(9)
                self.material_table.setGeometry(350, 450, 225, 300)
                self.material_table.setHorizontalHeaderLabels(["Materials", "Quantity(Kg)"])

                # Populate the Materials Table
                for key in list(materials.keys()):
                    key_item = QtWidgets.QTableWidgetItem(str(key))
                    value_item = QtWidgets.QTableWidgetItem(str(materials[key]))
                    key_item.setFlags(key_item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    value_item.setFlags(value_item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.material_table.setItem(list(materials.keys()).index(key), 0, key_item)
                    self.material_table.setItem(list(materials.keys()).index(key), 1, value_item)

                self.material_table.show()

                # Temperature Table
                self.temp_table = QtWidgets.QTableWidget(self.production_widget)
                self.temp_table.setGeometry(575, 450, 140, 300)
                self.temp_table.setRowCount(12)
                self.temp_table.setColumnCount(1)
                self.temp_table.setVerticalHeaderLabels(["Z" + str(i + 1) for i in range(12)])
                self.temp_table.setHorizontalHeaderLabels(["Temperature"])

                # Populate the Table
                for i in range(len(temperature)):
                    item = QtWidgets.QTableWidgetItem(str(temperature[i]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.temp_table.setItem(i, 0, item)

                self.temp_table.show()
            except Exception as e:
                print(e)

            self.group_box = QtWidgets.QGroupBox(self.production_widget)
            self.group_box.setGeometry(715, 450, 276, 301)
            self.group_box.setTitle("Remarks")
            self.group_box.setFont(QtGui.QFont("Arial", 10))

            self.show_remarks = QtWidgets.QTextEdit(self.group_box)
            self.show_remarks.setGeometry(0, 20, 230, 130)
            self.show_remarks.setAutoFillBackground(True)
            self.show_remarks.setText(remarks)
            self.show_remarks.setEnabled(False)
            self.show_remarks.show()

            self.group_box.show()

            self.export_btn = QtWidgets.QPushButton(self.production_widget)
            self.export_btn.setGeometry(730, 610, 100, 30)
            self.export_btn.setText("Export")
            self.export_btn.clicked.connect(exportToExcel)
            self.export_btn.show()

        def add_entry():

            if extruder_access:
                pass
            else:
                QMessageBox.critical(self.production_widget, 'Restricted Access',
                                     "You Dont Have Permission. \n Contact the Admin.")
                return

            self.entry_widget = QtWidgets.QWidget()
            self.entry_widget.setGeometry(300, 100, 750, 670)
            self.entry_widget.setWindowIcon(QtGui.QIcon("setting.png"))
            self.entry_widget.setWindowTitle("ADD EXTRUDER DATA")
            self.entry_widget.setStyleSheet("background-color : rgb(240,240,240);")
            self.entry_widget.setWindowModality(Qt.ApplicationModal)
            self.entry_widget.show()

            def get_entries():
                # get the data from the tables

                try:

                    # Time Table
                    temp_row = time_table.rowCount()
                    time_start = []
                    time_end = []
                    outputs = []

                    # Getting the data from the Time Table
                    for i in range(self.time_entry):
                        time_start.append(time_table.item(i, 0))  # time start
                        time_end.append(time_table.item(i, 1))  # time end
                        outputs.append(time_table.item(i, 2))

                    # Removing Null Values
                    time_start = [i for i in time_start if i is not None]
                    time_start = [i.text() for i in time_start]
                    time_end = [i for i in time_end if i is not None]
                    time_end = [i.text() for i in time_end]
                    outputs = [i for i in outputs if i is not None]
                    outputs = [i.text() for i in outputs]

                    total_time = timedelta()
                    try:
                        for i in range(len(time_start)):
                            t_start = datetime.strptime(time_start[i], "%m-%d-%Y %H:%M")
                            t_end = datetime.strptime(time_end[i], "%m-%d-%Y %H:%M")
                            total_time = total_time + (t_start - t_end)
                    except Exception as e:
                        print(e)

                    hours = str(int(total_time.total_seconds() // 3600))
                    minutes = str((int(total_time.total_seconds() % 3600) // 60))
                    seconds = str(int(total_time.total_seconds() % 60))

                    total_hours = round(abs(total_time.total_seconds() / 3600), 2)

                    time_start = ', '.join(["'{}'".format(time) for time in time_start])
                    time_end = ', '.join(["'{}'".format(time) for time in time_end])

                    # Getting the Data for temperature
                    temperature = []
                    for i in range(temperature_table.rowCount()):
                        temperature.append(temperature_table.item(i, 0))

                    temperature = [i for i in temperature if i is not None]
                    temperature = [i.text() for i in temperature]

                    # Declare additional variables need here like loss percentage
                    output_percent = round((float(product_output_input.text()) / float(product_input.text())) * 100,
                                           4)  # Round to the 4th decimal
                    loss_percent = round((float(loss_input.text()) / float(product_input.text())) * 100,
                                         4)  # Round to the 4th decimal
                    purge_duration = timedelta()
                    outputPerHour = round(float(product_output_input.text()) / total_hours, 4)

                    try:
                        purge_start = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                            "%Y-%m-%d %H:%M")

                        if purge_end < purge_start:
                            purge_end += timedelta(days=1)  # Adjust for next day

                        purge_duration = purge_end - purge_start

                    except:

                        purge_start = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                        purge_end = datetime.strptime(
                            datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                            "%Y-%m-%d %H:%M")
                        purge_duration = (purge_end - purge_start).total_seconds()

                    purge_duration = purge_duration.total_seconds() / 60
                    # SQL command here to insert Items
                    self.cursor.execute(
                        f"SELECT materials FROM production_merge WHERE t_prodid = '{productionID_input.text()}'")

                    self.total_mats = json.dumps(self.total_mats)
                    self.total_mats = self.total_mats.replace('\\', "")

                    # Convert the list to string
                    temperature = str(temperature).replace("[", "").replace("]", "")
                    outputs = str(outputs).replace("[", "").replace("]", "")
                    self.lot_numberList = str(self.lot_numberList).replace("[", "").replace("]", "")

                    try:

                        self.cursor.execute(f"""
                            INSERT INTO extruder( machine, qty_order, total_output, customer,
                            formula_id, product_code, order_id, total_time, time_start, time_end, output_percent,
                            loss, loss_percent, materials, purging, resin, purge_duration, screw_config, feed_rate, 
                            rpm, screen_size, operator, supervisor, temperature, outputs, output_per_hour, production_id, total_input,
                            remarks, lot_number, resin_quantity, encoded_on, machine_start, machine_off) 
                            VALUES('{machine_input.currentText()}', '{orderedQuantity_input.text()}', '{product_output_input.text()}',
                            '{customer_input.text().replace("'", "''")}', '{self.formulaID_input.text()}', '{productCode_input.text()}',
                            '{order_number_input.text()}', '{total_hours}', ARRAY[{time_start}]::timestamp[], ARRAY[{time_end}]::timestamp[], 
                            '{str(output_percent)}', '{loss_input.text()}', '{loss_percent}', '{self.total_mats}', '{purging_input.text()}',
                             '{resin_input.currentText()}', {purge_duration}, '{screwConf_input.currentText()}', '{feedRate_input.text()}',
                             '{rpm_input.text()}','{screenSize_input.text()}', '{operator_input.currentText()}', '{supervisor_input.currentText()}',
                             ARRAY[{temperature}]::INTEGER[], ARRAY[{outputs}]::FLOAT[], {outputPerHour}, {productionID_input.text()},
                             {product_input.text()},'{self.remarks_textBox.toPlainText()}', 
                             ARRAY[{self.lot_numberList}]::VARCHAR[], {resin_quantity.text()}, '{date.today()}',
                             '{machineStart_input.text()}', '{machineOff_input.text()}')

                                                """)
                        print("query successful")
                        self.conn.commit()
                        clear_inputs()

                        # Refresh the table
                        self.cursor.execute("""SELECT 
                                                process_id, TO_CHAR(DATE(time_start[1]), 'MM/DD/YYYY') as date,machine, customer, qty_order, total_output, output_per_hour, formula_id, product_code, total_time
                                                FROM extruder
                                                ORDER BY date DESC;
                                                """)
                        result = self.cursor.fetchall()
                        self.extruder_table.setRowCount(len(result))

                        # Populate table with data
                        for i in range(len(result)):
                            for j in range(len(column_names)):
                                item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                                if j == 3:
                                    pass
                                else:
                                    item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                                self.extruder_table.setItem(i, j, item)


                    except Exception as e:
                        print(e)
                        QMessageBox.critical(self.entry_widget, "ERROR", "INVALID ENTRY")

                        self.conn.rollback()
                except Exception as e:
                    print(e)
                    QMessageBox.critical(self.entry_widget, "ERROR", "INVALID ENTRY")
                    return

            def select_production():

                try:
                    self.table.deleteLater()
                    self.table2.deleteLater()
                    self.table3.deleteLater()

                except:
                    pass

                def show_table():
                    # Showing the Selected Item
                    self.table2.clearContents()
                    item = self.table.selectedItems()
                    item = [i.text() for i in item]

                    self.cursor.execute(f"""
                    SELECT t_prodid, t_lotnum, t_qtyreq, materials, t_prodcode
                    FROM production_merge
                    WHERE t_prodid = '{item[0]}'
                    """)

                    result = self.cursor.fetchall()
                    result = result[0]

                    material = result[3]

                    try:

                        for keys in list(material.keys()):
                            key = QTableWidgetItem(str(keys))
                            value = QTableWidgetItem(str(material[keys]))
                            self.table2.setItem(list(material.keys()).index(keys), 0, key)
                            key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                            self.table2.setItem(list(material.keys()).index(keys), 1, value)
                            value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                            self.table2.show()

                        label2.setText(str(result[2]))
                        prod_id = QTableWidgetItem(str(result[0]))
                        prod_id.setTextAlignment(Qt.AlignCenter)  # Align Center
                        lot_num = QTableWidgetItem(str(result[1]))
                        lot_num.setTextAlignment(Qt.AlignCenter)

                        prod_code = QTableWidgetItem(str(result[4]))
                        prod_code.setTextAlignment(Qt.AlignCenter)

                        self.table3.setItem(0, 0, prod_id)
                        prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                        self.table3.setItem(0, 1, lot_num)
                        lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)
                        self.table3.setItem(0, 2, prod_code)
                        prod_code.setFlags(prod_code.flags() & ~Qt.ItemIsEditable)

                    except Exception as e:
                        print(e)

                self.added_entry = 0  # for counting the lot numbers added
                self.total_mats = {}
                self.total_materialQty = 0
                self.total_quantity_order = 0
                self.total_output = 0
                self.lot_numberList = []
                self.total_outputPercent = 0

                def push_data():
                    item = self.table.selectedItems()
                    item = [i.text() for i in item]

                    try:
                        self.cursor.execute(f"""
                            SELECT * FROM production_merge
                            WHERE t_prodid = '{item[0]}' 
                            """)

                    except IndexError:
                        QMessageBox.critical(self.selectProd_widget, 'ERROR', "No Lot Number Selected")
                        return

                    result = self.cursor.fetchall()
                    result = result[0]

                    # Unpack the result
                    prod_id = result[25]
                    customer = result[1]
                    formula_id = result[26]
                    product_code = result[3]
                    product_color = result[4]
                    lot_number = result[7]
                    order_number = result[8]
                    machine_name = result[12]
                    quantity_order = result[13]
                    output_quantity = result[15]
                    remarks = result[16]
                    materials = result[-1]

                    # Get the Range of Lot Number
                    num1 = int(re.findall(r'\d+', lot_number)[0])
                    try:
                        num2 = int(re.findall(r'\d+', lot_number)[1])
                    except IndexError:
                        num2 = None
                        if int(cut_input.text()) > 0:
                            QMessageBox.critical(self.selectProd_widget, 'ERROR', 'Unable to Cut a Single Lot!')
                            return

                    if num2 == None:
                        lot_count = 1
                    else:
                        lot_count = (num2 - num1) + 1

                    if int(cut_input.text()) > 0:
                        for key in materials.keys():
                            materials[key] = round(materials[key] * (int(cut_input.text()) / lot_count), 3)
                    else:
                        pass

                    self.added_entry += 1
                    self.lot_numberList.append(lot_number)
                    self.total_output += output_quantity
                    self.total_quantity_order += quantity_order

                    # Getting the materials
                    for key in materials.keys():
                        if key in list(self.total_mats.keys()):
                            self.total_mats[key] = self.total_mats[key] + materials[key]
                            self.total_materialQty += materials[key]
                        else:
                            self.total_mats[key] = materials[key]
                            self.total_materialQty += materials[key]

                    # Set the Text to the Extruder Entry Form
                    productionID_input.setText(str(prod_id))
                    customer_input.setText(customer.strip())
                    productCode_input.setText(product_code.strip())
                    lot_number_input.setText('/'.join(self.lot_numberList))
                    self.formulaID_input.setText(str(formula_id).strip())
                    order_number_input.setText(str(order_number).strip())
                    product_input.setText(str(round(self.total_materialQty, 2)))
                    label2.setText(str(self.total_materialQty))

                    self.cursor.execute(f"""
                    SELECT t_prodid, t_lotnum
                    FROM production_merge
                    WHERE t_prodcode = '{product_code}' AND 
                    t_machine = '{machine_name}' AND t_fid = '{formula_id}';
                    """)
                    query_result = self.cursor.fetchall()
                    self.table.itemSelectionChanged.disconnect(show_table)
                    self.table.clearSelection()
                    self.table.clear()
                    self.table.setRowCount(len(query_result))
                    self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])

                    for row in range(len(query_result)):
                        prod = QTableWidgetItem(str(query_result[row][0]))
                        prod.setFlags(prod.flags() & ~Qt.ItemIsEditable)
                        lot = QTableWidgetItem(query_result[row][1])
                        lot.setFlags(lot.flags() & ~Qt.ItemIsEditable)
                        self.table.setItem(row, 0, prod)
                        self.table.setItem(row, 1, lot)
                    self.table.itemSelectionChanged.connect(show_table)
                    self.table.show()

                    self.table2.clearSelection()
                    self.table2.clear()
                    self.table2.setHorizontalHeaderLabels(["Materials", "Quantity"])
                    for keys in list(self.total_mats.keys()):
                        key = QTableWidgetItem(str(keys))
                        value = QTableWidgetItem(str(self.total_mats[keys]))
                        self.table2.setItem(list(self.total_mats.keys()).index(keys), 0, key)
                        key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                        self.table2.setItem(list(self.total_mats.keys()).index(keys), 1, value)
                        value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                    self.table2.show()

                    # Set It back to 0
                    cut_input.setText('0')

                def search():
                    try:
                        self.table.itemSelectionChanged.disconnect(show_table)
                        self.table.clearContents()
                        self.cursor.execute(f"""
                                            SELECT t_prodid, t_lotnum
                                            FROM production_merge
                                            WHERE t_lotnum ILIKE '%{search_bar.text()}%'
                                            """)
                        search_result = self.cursor.fetchall()

                        for i in range(len(search_result)):
                            item_pair = search_result[i]

                            item1 = QTableWidgetItem(str(item_pair[0]))
                            item2 = QTableWidgetItem(item_pair[1])
                            item2.setFlags(item2.flags() & ~Qt.ItemIsEditable)
                            self.table.setItem(i, 0, item1)
                            self.table.setItem(i, 1, item2)
                        self.table.itemSelectionChanged.connect(show_table)

                    except Exception as e:
                        print(e)

                def clear():

                    try:
                        self.cursor.execute("""
                            SELECT t_prodid, t_lotnum
                            FROM production_merge
                                            """)
                        result = self.cursor.fetchall()
                        self.table.setRowCount(len(result))
                        self.table.setColumnCount(2)
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])

                        self.table.itemSelectionChanged.disconnect(show_table)
                        self.table.clearSelection()
                        self.table.clear()
                        search_bar.clear()

                        self.table2.clearContents()
                        self.table3.clearContents()
                        self.table.itemSelectionChanged.connect(show_table)

                    except Exception as e:
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                        self.table.itemSelectionChanged.connect(show_table)

                    try:
                        self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                        for i in range(len(result)):
                            prod_id = QTableWidgetItem(str(result[i][0]))
                            lot_num = QTableWidgetItem(str(result[i][1]))
                            self.table.setItem(i, 0, prod_id)
                            prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                            self.table.setItem(i, 1, lot_num)
                            lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)
                            self.table.show()

                    except Exception as e:
                        print(e)

                def close_selection():
                    self.selectProd_widget.close()

                self.cursor.execute("""
                SELECT t_prodid, t_lotnum
                FROM production_merge
                ORDER BY t_prodid::Integer DESC

                """)
                result = self.cursor.fetchall()

                self.selectProd_widget = QtWidgets.QWidget()
                self.selectProd_widget.setGeometry(400, 200, 800, 630)
                self.selectProd_widget.setFixedSize(800, 600)

                search_bar = QtWidgets.QLineEdit(self.selectProd_widget)
                search_bar.setGeometry(40, 25, 170, 25)
                search_bar.setFont(QtGui.QFont("Arial", 10))
                search_bar.setPlaceholderText("Search Lot Number")
                search_bar.show()

                search_btn = QtWidgets.QPushButton(self.selectProd_widget)
                search_btn.setGeometry(210, 25, 70, 27)
                search_btn.setText("Search")
                search_btn.clicked.connect(search)
                search_btn.show()

                clear_btn = QtWidgets.QPushButton(self.selectProd_widget)
                clear_btn.setGeometry(280, 25, 70, 27)
                clear_btn.setText("Clear")
                clear_btn.clicked.connect(clear)
                clear_btn.show()

                # ProductionId and Lot Number Table widget
                self.table = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table.setGeometry(0, 70, 350, 500)
                self.table.setColumnCount(2)
                self.table.setRowCount(len(result))
                self.table.setHorizontalHeaderLabels(["Production ID", "Lot Number"])
                self.table.setColumnWidth(1, 180)
                self.table.setColumnWidth(0, 150)
                self.table.setFont(QtGui.QFont("Arial", 12))
                self.table.setStyleSheet('color: rgb(0,109,189);')
                self.table.verticalHeader().setVisible(False)
                try:
                    for i in range(len(result)):
                        prod_id = QTableWidgetItem(str(result[i][0]))
                        lot_num = QTableWidgetItem(str(result[i][1]))
                        self.table.setItem(i, 0, prod_id)
                        prod_id.setFlags(prod_id.flags() & ~Qt.ItemIsEditable)
                        self.table.setItem(i, 1, lot_num)
                        lot_num.setFlags(lot_num.flags() & ~Qt.ItemIsEditable)
                except Exception as e:
                    print(e)

                self.table.show()
                self.table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
                self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
                self.table.itemSelectionChanged.connect(show_table)

                # Materials Table
                self.table2 = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table2.setGeometry(350, 70, 450, 500)
                self.table2.setColumnCount(2)
                self.table2.setRowCount(16)
                self.table2.setColumnWidth(0, 205)
                self.table2.setColumnWidth(1, 225)
                self.table2.verticalHeader().setVisible(False)
                self.table2.setHorizontalHeaderLabels(["Materials", "Quantity"])
                self.table2.show()

                # Table 3 For showing Selected ProdID and Lot Number
                self.table3 = QtWidgets.QTableWidget(self.selectProd_widget)
                self.table3.setGeometry(380, 0, 402, 60)
                self.table3.setColumnCount(3)
                self.table3.setRowCount(1)
                self.table3.setHorizontalHeaderLabels(["Production ID", "Lot Number", "Product Code"])
                self.table3.setColumnWidth(0, 133)
                self.table3.setColumnWidth(1, 133)
                self.table3.setColumnWidth(2, 133)
                self.table3.setRowHeight(0, 35)
                font = QtGui.QFont("Arial", 12)
                font.setBold(True)
                self.table3.setFont(font)
                self.table3.setStyleSheet("color: rgb(0,109,189) ")
                self.table3.horizontalHeader().setStyleSheet("""
                    QHeaderView::section{
                        font-weight: bold;
                        background-color: rgb(0,109,189);
                        color: white;
                    }
                """)

                for i in range(2):
                    item = QTableWidgetItem("")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.table3.setItem(0, i, item)

                self.table3.verticalHeader().setVisible(False)
                self.table3.show()

                label1 = QtWidgets.QLabel(self.selectProd_widget)
                label1.setGeometry(80, 570, 120, 30)
                label1.setText("Total QTY (Kg)")
                label1.setFont(QtGui.QFont("Arial", 11))
                label1.setAlignment(Qt.AlignCenter)
                label1.setStyleSheet('border: 1px solid black;')
                label1.show()

                label2 = QtWidgets.QLabel(self.selectProd_widget)
                label2.setGeometry(200, 570, 150, 30)
                label2.setFont(QtGui.QFont("Arial", 13))
                label2.setAlignment(Qt.AlignCenter)
                label2.setStyleSheet("background-color: white; color: blue;")
                label2.show()

                cut_label = QLabel(self.selectProd_widget)
                cut_label.setGeometry(400, 570, 70, 30)
                cut_label.setText('Cut')
                cut_label.setFont(QtGui.QFont("Arial", 13))
                cut_label.show()

                cut_input = QLineEdit(self.selectProd_widget)
                cut_input.setGeometry(470, 570, 100, 30)
                cut_input.setFont(QtGui.QFont("Arial", 13))
                cut_input.setStyleSheet('color: rgb(0, 109, 189)')
                cut_input.setText('0')
                cut_input.show()

                # Save Button
                save_prod = QtWidgets.QPushButton(self.selectProd_widget)
                save_prod.setGeometry(630, 570, 70, 30)
                save_prod.setStyleSheet(
                    "background-color: rgb(194, 232, 255); border: 1px solid rgb(92, 154, 255); border-radius: 5px")
                save_prod.setText("Push Data")
                save_prod.clicked.connect(push_data)
                save_prod.show()

                close = QtWidgets.QPushButton(self.selectProd_widget)
                close.setGeometry(700, 570, 70, 30)
                close.setStyleSheet(
                    "background-color: rgb(194, 232, 255); border: 1px solid rgb(92, 154, 255); border-radius: 5px;")
                close.setText("Close")
                close.clicked.connect(close_selection)
                close.show()

                self.selectProd_widget.setWindowModality(Qt.ApplicationModal)
                self.selectProd_widget.show()

            self.time_entry = 0

            def output_changed(row, column):
                total_sum = 0
                if column == 2:
                    for i in range(time_table.rowCount()):
                        output = time_table.item(i, 2)
                        if output == None:
                            pass
                        else:
                            try:
                                total_sum += float(output.text())
                            except ValueError:
                                pass
                    product_output_input.setText(str(total_sum))

            def add_time():
                time_table.setRowCount(time_table.rowCount() + 1)
                try:
                    product_output_input.setText(
                        str(float(product_output_input.text()) + float(output_lineEdit.text())))

                except ValueError:
                    QMessageBox.critical(self.entry_widget, 'Value Error', "Invalid Input")
                    return

                item1 = QTableWidgetItem(time_start_input.text())
                item2 = QTableWidgetItem(time_end_input.text())
                item3 = QTableWidgetItem(output_lineEdit.text())

                item3.setTextAlignment(Qt.AlignCenter)

                time_table.setItem(self.time_entry, 0, item1)
                time_table.setItem(self.time_entry, 1, item2)
                time_table.setItem(self.time_entry, 2, item3)

                self.time_entry += 1

                output_lineEdit.clear()
                # Set the Focus back to date 1 entry
                date1.setFocus()

            def reset_table():
                time_table.clearContents()
                product_output_input.setText("0.0")
                self.time_entry = 0

            def loss_auto():
                if product_output_input.text() != "":
                    try:
                        loss_input.setText(
                            str(round(float(product_input.text()) - float(product_output_input.text()), 4)))
                    except:
                        loss_input.setText("INVALID")

            def autofill_temperature():
                if machine_input.currentText() == '1':

                    temperature_table.clear()
                    temperature_table.setHorizontalHeaderLabels(['Temperature'])
                    if resin_input.currentText() == 'LDPE(HMI)':
                        temp = [120, 140, 150, 150, 150, 150, 150, 150, 140, 140]
                    elif resin_input.currentText() == 'LDPE(LMI)':
                        temp = [200, 220, 220, 220, 220, 220, 220, 220, 220, 220]
                    elif resin_input.currentText() == 'HDPE(HMI)':
                        temp = [160, 180, 180, 180, 180, 180, 180, 180, 160, 160]
                    elif resin_input.currentText() == 'PP':
                        temp = [180, 200, 180, 180, 180, 180, 180, 180, 180, 180]
                    elif resin_input.currentText() == 'PET':
                        temp = [200, 220, 220, 220, 220, 220, 220, 220, 220, 220]
                    elif resin_input.currentText() == 'GPPS':
                        temp = [200, 220, 220, 220, 220, 220, 220, 220, 220, 220]
                    else:
                        return

                    # Set the setting to the Temperature Table
                    for i in range(len(temp)):
                        item = QTableWidgetItem(str(temp[i]))
                        temperature_table.setItem(i, 0, item)

                elif machine_input.currentText() == '2':

                    temperature_table.clear()
                    temperature_table.setHorizontalHeaderLabels(['Temperature'])

                    if resin_input.currentText() == 'LDPE(HMI)':
                        temp = [150, 160, 160, 160, 160]
                    elif resin_input.currentText() == 'HDPE(HMI)':
                        temp = [160, 160, 180, 180, 180]
                    elif resin_input.currentText() == 'PP':
                        temp = [180, 200, 200, 200, 200]
                    else:
                        return

                    for i in range(len(temp)):
                        item = QTableWidgetItem(str(temp[i]))
                        temperature_table.setItem(i, 0, item)

                elif machine_input.currentText() == '3':

                    temperature_table.clear()
                    temperature_table.setHorizontalHeaderLabels(['Temperature'])

                    if resin_input.currentText() == 'LDPE(HMI)':
                        temp = [150, 150, 150, 150, 150, 130, 130]
                    elif resin_input.currentText() == 'HDPE(HMI)':
                        temp = [160, 180, 180, 180, 180, 180, 160]
                    elif resin_input.currentText() == 'PP':
                        temp = [180, 200, 180, 180, 180, 180, 180]
                    else:
                        return

                    for i in range(len(temp)):
                        item = QTableWidgetItem(str(temp[i]))
                        temperature_table.setItem(i, 0, item)

                elif machine_input.currentText() == '5':

                    temperature_table.clear()
                    temperature_table.setHorizontalHeaderLabels(['Temperature'])

                    if resin_input.currentText() == 'LDPE(HMI)':
                        temp = [150, 150, 150, 150, 150, 130, 130]
                    elif resin_input.currentText() == 'HDPE(HMI)':
                        temp = [160, 180, 180, 180, 180, 180, 160]
                    elif resin_input.currentText() == 'PP':
                        temp = [180, 200, 180, 180, 180, 180, 180]
                    else:
                        return

                    for i in range(len(temp)):
                        item = QTableWidgetItem(str(temp[i]))
                        temperature_table.setItem(i, 0, item)

                elif machine_input.currentText() == '8':

                    temperature_table.clear()
                    temperature_table.setHorizontalHeaderLabels(['Temperature'])

                    if resin_input.currentText() == 'LDPE(HMI)':
                        temp = [170, 170, 170, 170, 170, 170, 160, 160, 160, 160, 150, 150]
                    elif resin_input.currentText() == 'HDPE(HMI)':
                        temp = [180, 180, 180, 180, 180, 180, 180, 180, 180, 160, 140, 140]
                    elif resin_input.currentText() == 'PP':
                        temp = [180, 200, 180, 180, 180, 180, 180, 180, 180, 180, 180, 180]
                    elif resin_input.currentText() == 'PET':
                        temp = [210, 220, 200, 220, 220, 220, 220, 220, 220, 220, 230, 230]
                    else:
                        return

                    for i in range(len(temp)):
                        item = QTableWidgetItem(str(temp[i]))
                        temperature_table.setItem(i, 0, item)

            def clear_inputs():
                try:

                    productionID_input.clear()
                    customer_input.clear()
                    orderedQuantity_input.clear()
                    productCode_input.clear()
                    product_output_input.setText("0.0")
                    self.formulaID_input.clear()
                    lot_number_input.clear()
                    feedRate_input.clear()
                    rpm_input.clear()
                    screenSize_input.clear()
                    loss_input.clear()
                    purgeStart_input.setText("00:00")
                    purgeEnd_input.setText("00:00")
                    order_number_input.clear()
                    product_input.clear()
                    purging_input.clear()
                    resin_quantity.clear()
                    time_table.clearContents()
                    temperature_table.clearContents()
                    self.remarks_textBox.clear()
                    self.total_mats = {}

                    # Set the Row to 1 again
                    self.time_entry = 0

                except:
                    pass

            # Create two new widget for the VBOX Layout
            self.leftInput_side = QtWidgets.QWidget(self.entry_widget)
            self.leftInput_side.setGeometry(0, 0, 375, 400)
            self.leftInput_side.show()

            self.right_side = QtWidgets.QWidget(self.entry_widget)
            self.right_side.setGeometry(375, 0, 375, 400)
            self.right_side.show()

            # Create Vertical Box Layout
            self.left_vbox = QtWidgets.QFormLayout(self.leftInput_side)
            self.left_vbox.setSpacing(15)
            self.right_vbox = QtWidgets.QFormLayout(self.right_side)
            self.right_vbox.setSpacing(15)

            font = QtGui.QFont("Berlin Sans FB", 13)

            productionID_label = QtWidgets.QLabel()
            productionID_label.setText("Production ID")
            productionID_label.setFont(font)

            customer_label = QtWidgets.QLabel()
            customer_label.setText("Customer")
            customer_label.setFont(font)

            machine_label = QtWidgets.QLabel()
            machine_label.setText("Machine No.")
            machine_label.setFont(font)

            productCode_label = QtWidgets.QLabel()
            productCode_label.setText("Product Code")
            productCode_label.setFont(font)

            productOutput_label = QtWidgets.QLabel()
            productOutput_label.setText("Output (kg)")
            productOutput_label.setFont(font)

            formulaID_label = QtWidgets.QLabel()
            formulaID_label.setText("Formula ID")
            formulaID_label.setFont(font)

            lotnumber_label = QtWidgets.QLabel()
            lotnumber_label.setText("Lot Number")
            lotnumber_label.setFont(font)

            orderedQuantity_label = QtWidgets.QLabel()
            orderedQuantity_label.setText("Ordered Qty")
            orderedQuantity_label.setFont(font)

            feedrate_label = QtWidgets.QLabel()
            feedrate_label.setText("Feed Rate")
            feedrate_label.setFont(font)

            rpm_label = QtWidgets.QLabel()
            rpm_label.setText("RPM")
            rpm_label.setFont(font)

            screenSize_label = QtWidgets.QLabel()
            screenSize_label.setText("Screen Size")
            screenSize_label.setFont(font)

            screwConf_label = QtWidgets.QLabel()
            screwConf_label.setText("Screw Config")
            screwConf_label.setFont(font)

            loss_label = QtWidgets.QLabel()
            loss_label.setText("Loss")
            loss_label.setFont(font)

            purgeStart_label = QtWidgets.QLabel()
            purgeStart_label.setText("Purge Start")
            purgeStart_label.setFont(font)

            purgeEnd_label = QtWidgets.QLabel()
            purgeEnd_label.setText("Purge End")
            purgeEnd_label.setFont(font)

            remarks_label = QtWidgets.QLabel()

            operator_label = QtWidgets.QLabel()
            operator_label.setText("Operator")
            operator_label.setFont(font)

            supervisor_label = QtWidgets.QLabel()
            supervisor_label.setText("Supervisor")
            supervisor_label.setFont(font)

            order_number_lbl = QtWidgets.QLabel()
            order_number_lbl.setText("Order Number")
            order_number_lbl.setFont(font)

            resin_label = QtWidgets.QLabel()
            resin_label.setText("Resin (Prod)")
            resin_label.setFont(font)

            purging_label = QtWidgets.QLabel()
            purging_label.setText("Purging To")
            purging_label.setFont(font)

            product_input_label = QtWidgets.QLabel()
            product_input_label.setText("Input")
            product_input_label.setFont(font)

            machineStart_label = QtWidgets.QLabel()
            machineStart_label.setText("Machine Start")
            machineStart_label.setFont(font)

            machineOff_label = QtWidgets.QLabel()
            machineOff_label.setText("Machine Off")
            machineOff_label.setFont(font)

            # QLineEdit Boxes
            productionID_input = QtWidgets.QLineEdit()
            productionID_input.setFixedHeight(25)
            productionID_input.setEnabled(False)
            productionID_input.setAlignment(Qt.AlignCenter)
            productionID_input.setStyleSheet("background-color: white; border: 1px solid black")

            machine_input = QtWidgets.QComboBox()
            machine_input.setFixedHeight(25)
            machine_input.addItem("1")
            machine_input.addItem("2")
            machine_input.addItem("3")
            machine_input.addItem("5")
            machine_input.addItem("6")
            machine_input.addItem("8")
            machine_input.addItem("9")
            machine_input.activated.connect(autofill_temperature)
            machine_input.setStyleSheet("background-color: white; border: 1px solid black")

            customer_input = QtWidgets.QLineEdit()
            customer_input.setFixedHeight(25)
            customer_input.setEnabled(False)
            customer_input.setAlignment(Qt.AlignCenter)
            customer_input.setStyleSheet("background-color: white; border: 1px solid black")

            orderedQuantity_input = QtWidgets.QLineEdit()
            orderedQuantity_input.setAlignment(Qt.AlignCenter)
            orderedQuantity_input.setFixedHeight(25)
            orderedQuantity_input.setStyleSheet("background-color: white; border: 1px solid black")

            productCode_input = QtWidgets.QLineEdit()
            productCode_input.setFixedHeight(25)
            productCode_input.setEnabled(False)
            productCode_input.setAlignment(Qt.AlignCenter)
            productCode_input.setStyleSheet("background-color: white; border: 1px solid black")

            product_output_input = QtWidgets.QLineEdit()
            product_output_input.setFixedHeight(25)
            product_output_input.setEnabled(False)
            product_output_input.setText("0.0")
            product_output_input.setAlignment(Qt.AlignCenter)
            product_output_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_output_input.textChanged.connect(loss_auto)

            self.formulaID_input = QtWidgets.QLineEdit()
            self.formulaID_input.setAlignment(Qt.AlignCenter)
            self.formulaID_input.setFixedHeight(25)
            self.formulaID_input.setEnabled(False)
            self.formulaID_input.setStyleSheet("background-color: white; border: 1px solid black")

            lot_number_input = QtWidgets.QLineEdit()
            lot_number_input.setAlignment(Qt.AlignCenter)
            lot_number_input.setFixedHeight(25)
            lot_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            feedRate_input = QtWidgets.QLineEdit()
            feedRate_input.setFixedHeight(25)
            feedRate_input.setStyleSheet("background-color: white; border: 1px solid black")

            rpm_input = QtWidgets.QLineEdit()
            rpm_input.setFixedHeight(25)
            rpm_input.setStyleSheet("background-color: white; border: 1px solid black")

            screenSize_input = QtWidgets.QLineEdit()
            screenSize_input.setFixedHeight(25)
            screenSize_input.setStyleSheet("background-color: white; border: 1px solid black")

            screwConf_input = QtWidgets.QComboBox()
            screwConf_input.setFixedHeight(25)
            screwConf_input.setStyleSheet("background-color: white; border: 1px solid black ")
            screwConf_input.addItem('STANDARD')
            screwConf_input.addItem('COIL')

            loss_input = QtWidgets.QLineEdit()
            loss_input.setFixedHeight(25)
            loss_input.setEnabled(False)
            loss_input.setAlignment(Qt.AlignCenter)
            loss_input.setStyleSheet("background-color: white; border: 1px solid black")

            purgeStart_input = QtWidgets.QLineEdit()
            purgeStart_input.setFixedHeight(25)
            purgeStart_input.setAlignment(Qt.AlignCenter)
            purgeStart_input.setStyleSheet("background-color: white; border: 1px solid black")
            purgeStart_input.setInputMask("99:99;_")

            purgeEnd_input = QtWidgets.QLineEdit()
            purgeEnd_input.setFixedHeight(25)
            purgeEnd_input.setAlignment(Qt.AlignCenter)
            purgeEnd_input.setStyleSheet("background-color: white; border: 1px solid black")
            purgeEnd_input.setInputMask("99:99;_")

            operator_input = QtWidgets.QComboBox()
            operator_input.setFixedHeight(25)
            operator_input.setStyleSheet("background-color: white; border: 1px solid black")
            operator_input.addItem('Arnel Rosario')
            operator_input.addItem('Edsel Mangarin')
            operator_input.addItem('Arjie Galgo')
            operator_input.addItem('Aldrin Villacrusis')
            operator_input.addItem('Joel Miras')
            operator_input.addItem('Jowie Vasquez')
            operator_input.addItem('Roselie Mapandi')
            operator_input.setEditable(True)

            supervisor_input = QtWidgets.QComboBox()
            supervisor_input.setFixedHeight(25)
            supervisor_input.setStyleSheet("background-color: white; border: 1px solid black")
            supervisor_input.addItem('Julius Fundano')
            supervisor_input.setEditable(True)

            order_number_input = QtWidgets.QLineEdit()
            order_number_input.setFixedHeight(25)
            order_number_input.setEnabled(False)
            order_number_input.setAlignment(Qt.AlignCenter)
            order_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            resin_input = QtWidgets.QComboBox()
            resin_input.setFixedHeight(25)
            resin_input.addItem('LDPE(HMI)')
            resin_input.addItem('LDPE(LMI)')
            resin_input.addItem('HDPE(HMI)')
            resin_input.addItem('PP')
            resin_input.addItem('PET')
            resin_input.addItem('GPPS')
            resin_input.addItem('RUBBER')
            resin_input.activated.connect(autofill_temperature)
            resin_input.setStyleSheet("background-color: white; border: 1px solid black")

            resin_quantity = QLineEdit()
            resin_quantity.setFixedHeight(25)
            resin_quantity.setAlignment(Qt.AlignCenter)
            resin_quantity.setStyleSheet("background-color: white; border: 1px solid black")

            resin_quantity_label = QLabel()
            resin_quantity_label.setFont(font)
            resin_quantity_label.setText('Purging Resin Qty')

            purging_input = QtWidgets.QLineEdit()
            purging_input.setFixedHeight(25)
            purging_input.setAlignment(Qt.AlignCenter)
            purging_input.setStyleSheet("background-color: white; border: 1px solid black")

            product_input = QtWidgets.QLineEdit()
            product_input.setFixedHeight(25)
            product_input.setAlignment(Qt.AlignCenter)
            product_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_input.setInputMask('')

            machineStart_input = QDateTimeEdit()
            machineStart_input.setFixedHeight(25)
            machineStart_input.setStyleSheet("background-color: white; border: 1px solid black")
            machineStart_input.setDisplayFormat("yyyy-MM-dd HH:mm")

            machineOff_input = QDateTimeEdit()
            machineOff_input.setFixedHeight(25)
            machineOff_input.setStyleSheet("background-color: white; border: 1px solid black")
            machineOff_input.setDisplayFormat("yyyy-MM-dd HH:mm")

            validator = QtGui.QDoubleValidator()
            validator.setNotation(QtGui.QDoubleValidator.StandardNotation)

            product_input.setValidator(validator)

            product_input.textChanged.connect(loss_auto)

            self.groupBoxRemarks = QtWidgets.QGroupBox(self.entry_widget)
            self.groupBoxRemarks.setGeometry(550, 450, 190, 150)
            self.groupBoxRemarks.setTitle("Remarks")
            self.groupBoxRemarks.show()

            self.remarks_textBox = QtWidgets.QTextEdit(self.groupBoxRemarks)
            self.remarks_textBox.setGeometry(0, 20, 190, 130)
            self.remarks_textBox.show()

            # Left Side of Vertical Box
            self.left_vbox.addRow(productionID_label, productionID_input)
            self.left_vbox.addRow(productCode_label, productCode_input)
            self.left_vbox.addRow(customer_label, customer_input)
            self.left_vbox.addRow(orderedQuantity_label, orderedQuantity_input)
            self.left_vbox.addRow(lotnumber_label, lot_number_input)
            self.left_vbox.addRow(product_input_label, product_input)
            self.left_vbox.addRow(productOutput_label, product_output_input)
            self.left_vbox.addRow(loss_label, loss_input)
            self.left_vbox.addRow(machine_label, machine_input)
            self.left_vbox.addRow(formulaID_label, self.formulaID_input)
            self.left_vbox.addRow(order_number_lbl, order_number_input)
            self.left_vbox.addRow(feedrate_label, feedRate_input)
            # Add widgets to the right Form Box

            self.right_vbox.addRow(rpm_label, rpm_input)
            self.right_vbox.addRow(screenSize_label, screenSize_input)
            self.right_vbox.addRow(screwConf_label, screwConf_input)
            self.right_vbox.addRow(purging_label, purging_input)
            self.right_vbox.addRow(resin_label, resin_input)
            self.right_vbox.addRow(resin_quantity_label, resin_quantity)
            self.right_vbox.addRow(purgeStart_label, purgeStart_input)
            self.right_vbox.addRow(purgeEnd_label, purgeEnd_input)
            self.right_vbox.addRow(machineStart_label, machineStart_input)
            self.right_vbox.addRow(machineOff_label, machineOff_input)
            self.right_vbox.addRow(operator_label, operator_input)
            self.right_vbox.addRow(supervisor_label, supervisor_input)

            # Time Table Entry
            time_table = QtWidgets.QTableWidget(self.entry_widget)
            time_table.setGeometry(0, 450, 400, 180)
            time_table.setColumnCount(3)
            time_table.setRowCount(0)
            time_table.setColumnWidth(0, 145)
            time_table.setColumnWidth(1, 145)
            time_table.setStyleSheet("background-color: white;")
            time_table.setFont(QtGui.QFont("Arial", 10))

            time_table.cellChanged.connect(output_changed)

            time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])
            time_table.show()

            # Temperature Table Entry
            temperature_table = QtWidgets.QTableWidget(self.entry_widget)
            temperature_table.setGeometry(400, 450, 150, 180)
            temperature_table.setColumnCount(1)
            temperature_table.setRowCount(12)
            temperature_table.setStyleSheet("background-color: white;")
            temperature_table.setHorizontalHeaderLabels(["Temperature"])
            temperature_index = ["Z" + str(i + 1) for i in range(12)]  # set the index
            temperature_table.setVerticalHeaderLabels(temperature_index)
            temperature_table.show()

            # Select Production Data Button
            select_prod = QtWidgets.QPushButton(self.entry_widget)
            select_prod.setGeometry(555, 635, 60, 25)
            select_prod.setStyleSheet(
                'background-color: rgb(194, 232, 255); border-radius: 5px; border: 1px solid rgb(92, 154, 255)')
            select_prod.setText("Select")
            select_prod.clicked.connect(select_production)
            select_prod.setCursor(Qt.PointingHandCursor)
            select_prod.show()

            save_btn = QtWidgets.QPushButton(self.entry_widget)
            save_btn.setGeometry(490, 635, 60, 25)
            save_btn.setStyleSheet(
                'background-color: rgb(194, 232, 255); border-radius: 5px; border: 1px solid rgb(92, 154, 255)')
            save_btn.clicked.connect(get_entries)
            save_btn.setText("Save")
            save_btn.setCursor(Qt.PointingHandCursor)
            save_btn.show()

            clear_btn = QtWidgets.QPushButton(self.entry_widget)
            clear_btn.setGeometry(620, 635, 60, 25)
            clear_btn.setStyleSheet(
                'background-color: rgb(194, 232, 255); border-radius: 5px; border: 1px solid rgb(92, 154, 255)')
            clear_btn.clicked.connect(clear_inputs)
            clear_btn.setText("Clear")
            clear_btn.show()

            default_date = QtCore.QDateTime(2024, 1, 1, 0, 0)

            time_start_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_start_input.setGeometry(30, 425, 120, 25)
            time_start_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_start_input.setDateTime(default_date)
            time_start_input.show()

            time_end_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_end_input.setGeometry(180, 425, 120, 25)
            time_end_input.setDisplayFormat("MM-dd-yyyy HH:mm")
            time_end_input.setDateTime(default_date)
            time_end_input.show()

            output_lineEdit = QtWidgets.QLineEdit(self.entry_widget)
            output_lineEdit.setGeometry(310, 425, 80, 25)
            output_lineEdit.setAlignment(Qt.AlignCenter)
            output_lineEdit.setStyleSheet("background-color: white; border: 1px solid black")
            output_lineEdit.show()

            self.plus_icon = ClickableLabel(self.entry_widget)
            self.plus_icon.setGeometry(390, 425, 25, 25)
            self.plus_icon.setPixmap(QtGui.QIcon('plus.png').pixmap(25, 25))
            self.plus_icon.setCursor(Qt.PointingHandCursor)
            self.plus_icon.clicked.connect(add_time)
            self.plus_icon.show()

            self.reset_icon = ClickableLabel(self.entry_widget)
            self.reset_icon.setGeometry(425, 425, 25, 25)
            self.reset_icon.setPixmap(QtGui.QIcon('reset.png').pixmap(20, 20))
            self.reset_icon.setCursor(Qt.PointingHandCursor)
            self.reset_icon.clicked.connect(reset_table)
            self.reset_icon.show()

        def update_entry():

            if extruder_access:
                pass
            else:
                QMessageBox.critical(self.production_widget, 'Restricted Access',
                                     "You Dont Have Permission. \n Contact the Admin.")
                return

            try:
                selected = self.extruder_table.selectedItems()
                selected = [i.text() for i in selected]
                self.cursor.execute(f"SELECT * FROM extruder WHERE process_id = {selected[0]}")
                result = self.cursor.fetchall()
                result = result[0]

            except:
                QMessageBox.critical(self.production_widget, "ERROR", "No Data Selected")
                return

            self.entry_widget = QtWidgets.QWidget()
            self.entry_widget.setGeometry(300, 100, 800, 750)
            self.entry_widget.setStyleSheet("background-color : rgb(240,240,240);")
            self.entry_widget.setWindowModality(Qt.ApplicationModal)
            self.entry_widget.show()

            def update():
                # get the data from the tables

                # Time Table
                temp_row = time_table.rowCount()
                time_start = []
                time_end = []
                outputs = []

                # Getting the data from the Time Table
                for i in range(8):
                    if time_table.item(i, 0) == None:
                        break
                    else:
                        time_start.append(time_table.item(i, 0))  # time start
                        time_end.append(time_table.item(i, 1))  # time end
                        outputs.append(time_table.item(i, 2))

                # Removing Null Values
                time_start = [i for i in time_start if i is not None]
                time_start = [i.text() for i in time_start]
                time_end = [i for i in time_end if i is not None]
                time_end = [i.text() for i in time_end]
                outputs = [i for i in outputs if i is not None]
                outputs = [i.text() for i in outputs]

                print(time_start)

                total_time = timedelta()

                try:
                    for i in range(len(time_start)):
                        t_start = datetime.strptime(time_start[i], "%Y-%m-%d %H:%M")
                        t_end = datetime.strptime(time_end[i], "%Y-%m-%d %H:%M")
                        total_time = total_time + (t_end - t_start)
                        print(total_time)
                    print(total_time)
                except Exception as e:
                    print(e)

                hours = str(int(total_time.total_seconds() // 3600)).replace('-', '')
                minutes = str((int(total_time.total_seconds() % 3600) // 60))
                seconds = str(int(total_time.total_seconds() % 60))

                total_hours = round(abs(total_time.total_seconds() / 3600), 2)

                time_start = ', '.join(["'{}'".format(time) for time in time_start])
                time_end = ', '.join(["'{}'".format(time) for time in time_end])

                # Getting the Data for temperature
                temperature = []
                for i in range(temperature_table.rowCount()):
                    temperature.append(temperature_table.item(i, 0))

                temperature = [i for i in temperature if i is not None]
                temperature = [i.text() for i in temperature]

                # Declare additional variables need here like loss percentage
                output_percent = round((float(product_output_input.text()) / float(product_input.text())) * 100,
                                       4)  # Round to the 4th decimal
                loss_percent = round((float(loss_input.text()) / float(product_input.text())) * 100,
                                     4)  # Round to the 4th decimal
                outputPerHour = round(float(product_output_input.text()) / total_hours, 4)

                try:
                    purge_start = datetime.strptime(
                        datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                    purge_end = datetime.strptime(
                        datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                        "%Y-%m-%d %H:%M")

                    print(purge_start, purge_end)
                    if purge_end < purge_start:
                        purge_end += timedelta(days=1)  # Adjust for next day
                    purge_duration = purge_end - purge_start

                except:

                    purge_start = datetime.strptime(
                        datetime.today().strftime("%Y-%m-%d") + " " + purgeStart_input.text(), "%Y-%m-%d %H:%M")
                    purge_end = datetime.strptime(
                        datetime.today().strftime("%Y-%m-%d") + " " + purgeEnd_input.text(),
                        "%Y-%m-%d %H:%M")
                    purge_duration = (purge_end - purge_start).total_seconds()

                purge_duration = purge_duration.total_seconds() / 60

                # SQL command here to insert Items
                self.cursor.execute(
                    f"SELECT materials FROM production_merge WHERE t_prodid = '{productionID_input.text()}'")
                material = self.cursor.fetchall()
                material = material[0][0]
                material = json.dumps(material)

                # Convert the list to string
                temperature = str(temperature).replace("[", "").replace("]", "")
                outputs = str(outputs).replace("[", "").replace("]", "")

                try:
                    self.cursor.execute(f"""
                        UPDATE extruder
                        SET  total_time = {total_hours}, machine = '{machine_input.currentText()}', total_input = {product_input.text()}, outputs = ARRAY[{outputs}]::FLOAT[],
                        temperature = ARRAY[{temperature}]::INTEGER[], remarks = '{self.remarks_textBox.toPlainText()}',
                        feed_rate = '{feedRate_input.text()}', rpm = '{rpm_input.text()}', screen_size = '{screenSize_input.text()}',
                        screw_config = '{screwConf_input.currentText()}', purging = '{purging_input.text()}', resin = '{resin_input.currentText()}',
                        purge_duration = {purge_duration}, operator = '{operator_input.text()}', supervisor = '{supervisor_input.text()}',
                        time_start = ARRAY[{time_start}]::timestamp[], time_end =  ARRAY[{time_end}]::timestamp[],
                        output_percent = '{str(output_percent)}', loss = '{loss_input.text()}', loss_percent = '{loss_percent}',
                        output_per_hour = '{outputPerHour}', total_output = {product_output_input.text()}, resin_quantity = {resin_quantity.text()},
                        qty_order = {orderedQuantity_input.text()}, machine_start = '{machineStart_input.text()}', machine_off = '{machineOff_input.text()}'
                        WHERE process_id = {selected[0]};

                        """)

                    QMessageBox.information(self.entry_widget, "UPDATE SUCCESSFUL",
                                            f"Successfully Updated \n Form No. {selected[0]}")

                    self.conn.commit()
                    self.production()  # Refreshes the Table
                    self.entry_widget.close()
                except Exception as e:
                    print("Insert Failed")
                    print(e)
                    self.conn.rollback()

            self.time_entry = len(result[9])

            def add_time():

                item1 = QTableWidgetItem(time_start_input.text())
                item2 = QTableWidgetItem(time_end_input.text())
                item3 = QTableWidgetItem(output_lineEdit.text())

                item3.setTextAlignment(Qt.AlignCenter)

                time_table.setItem(self.time_entry, 0, item1)
                time_table.setItem(self.time_entry, 1, item2)
                time_table.setItem(self.time_entry, 2, item3)

                self.time_entry += 1

                # Auto increment Output
                product_output_input.setText(str(float(product_output_input.text()) + float(output_lineEdit.text())))

                output_lineEdit.clear()

            def reset_table():
                time_table.clearContents()
                product_output_input.setText("0.0")
                self.time_entry = 0

            def loss_auto():
                if product_output_input.text() != "":
                    try:
                        loss_input.setText(
                            str(round(float(product_input.text()) - float(product_output_input.text()), 4)))
                    except:
                        loss_input.setText("INVALID")

            # Create two new widget for the VBOX Layout
            self.leftInput_side = QtWidgets.QWidget(self.entry_widget)
            self.leftInput_side.setGeometry(0, 0, 400, 450)
            self.leftInput_side.show()

            self.right_side = QtWidgets.QWidget(self.entry_widget)
            self.right_side.setGeometry(400, 0, 400, 450)
            self.right_side.show()

            # Create Vertical Box Layout
            self.left_vbox = QtWidgets.QFormLayout(self.leftInput_side)
            self.left_vbox.setSpacing(20)
            self.right_vbox = QtWidgets.QFormLayout(self.right_side)
            self.right_vbox.setSpacing(20)

            font = QtGui.QFont("Berlin Sans FB", 14)

            productionID_label = QtWidgets.QLabel()
            productionID_label.setText("Production ID")
            productionID_label.setFont(font)

            customer_label = QtWidgets.QLabel()
            customer_label.setText("Customer")
            customer_label.setFont(font)

            machine_label = QtWidgets.QLabel()
            machine_label.setText("Machine No.")
            machine_label.setFont(font)

            productCode_label = QtWidgets.QLabel()
            productCode_label.setText("Product Code")
            productCode_label.setFont(font)

            productOutput_label = QtWidgets.QLabel()
            productOutput_label.setText("Output (kg)")
            productOutput_label.setFont(font)

            formulaID_label = QtWidgets.QLabel()
            formulaID_label.setText("Formula ID")
            formulaID_label.setFont(font)

            lotnumber_label = QtWidgets.QLabel()
            lotnumber_label.setText("Lot Number")
            lotnumber_label.setFont(font)

            orderedQuantity_label = QtWidgets.QLabel()
            orderedQuantity_label.setText("Ordered Qty")
            orderedQuantity_label.setFont(font)

            feedrate_label = QtWidgets.QLabel()
            feedrate_label.setText("Feed Rate")
            feedrate_label.setFont(font)

            rpm_label = QtWidgets.QLabel()
            rpm_label.setText("RPM")
            rpm_label.setFont(font)

            screenSize_label = QtWidgets.QLabel()
            screenSize_label.setText("Screen Size")
            screenSize_label.setFont(font)

            screwConf_label = QtWidgets.QLabel()
            screwConf_label.setText("Screw Config")
            screwConf_label.setFont(font)

            loss_label = QtWidgets.QLabel()
            loss_label.setText("Loss")
            loss_label.setFont(font)

            purgeStart_label = QtWidgets.QLabel()
            purgeStart_label.setText("Purge Start")
            purgeStart_label.setFont(font)

            purgeEnd_label = QtWidgets.QLabel()
            purgeEnd_label.setText("Purge End")
            purgeEnd_label.setFont(font)

            remarks_label = QtWidgets.QLabel()

            operator_label = QtWidgets.QLabel()
            operator_label.setText("Operator")
            operator_label.setFont(font)

            supervisor_label = QtWidgets.QLabel()
            supervisor_label.setText("Supervisor")
            supervisor_label.setFont(font)

            order_number_lbl = QtWidgets.QLabel()
            order_number_lbl.setText("Order Number")
            order_number_lbl.setFont(font)

            resin_label = QtWidgets.QLabel()
            resin_label.setText("Resin")
            resin_label.setFont(font)

            purging_label = QtWidgets.QLabel()
            purging_label.setText("Purging")
            purging_label.setFont(font)

            product_input_label = QtWidgets.QLabel()
            product_input_label.setText("Input")
            product_input_label.setFont(font)

            machineStart_label = QtWidgets.QLabel()
            machineStart_label.setText("Machine Start")
            machineStart_label.setFont(font)

            machineOff_label = QtWidgets.QLabel()
            machineOff_label.setText("Machine Off")
            machineOff_label.setFont(font)

            # QLineEdit Boxes
            productionID_input = QtWidgets.QLineEdit()
            productionID_input.setFixedHeight(25)
            productionID_input.setEnabled(False)
            productionID_input.setAlignment(Qt.AlignCenter)
            productionID_input.setStyleSheet("background-color: white; border: 1px solid black")
            productionID_input.setText(str(result[28]))
            productionID_input.setEnabled(False)

            machine_input = QtWidgets.QComboBox()
            machine_input.setFixedHeight(25)
            machine_input.addItem("1")
            machine_input.addItem("2")
            machine_input.addItem("3")
            machine_input.addItem("5")
            machine_input.addItem("6")
            machine_input.addItem("8")
            machine_input.addItem("9")
            machine_input.setStyleSheet("background-color: white; border: 1px solid black")
            machine_input.setCurrentText(result[1])

            customer_input = QtWidgets.QLineEdit()
            customer_input.setFixedHeight(25)
            customer_input.setAlignment(Qt.AlignCenter)
            customer_input.setStyleSheet("background-color: white; border: 1px solid black")
            customer_input.setText(result[4])
            customer_input.setEnabled(True)

            orderedQuantity_input = QtWidgets.QLineEdit()
            orderedQuantity_input.setAlignment(Qt.AlignCenter)
            orderedQuantity_input.setFixedHeight(25)
            orderedQuantity_input.setStyleSheet("background-color: white; border: 1px solid black")
            orderedQuantity_input.setText(str(result[2]))

            productCode_input = QtWidgets.QLineEdit()
            productCode_input.setFixedHeight(25)
            productCode_input.setEnabled(False)
            productCode_input.setAlignment(Qt.AlignCenter)
            productCode_input.setStyleSheet("background-color: white; border: 1px solid black")
            productCode_input.setText(result[6])

            product_output_input = QtWidgets.QLineEdit()
            product_output_input.setFixedHeight(25)
            product_output_input.setAlignment(Qt.AlignCenter)
            product_output_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_output_input.setText(str(result[3]))

            self.formulaID_input = QtWidgets.QLineEdit()
            self.formulaID_input.setAlignment(Qt.AlignCenter)
            self.formulaID_input.setFixedHeight(25)
            self.formulaID_input.setEnabled(False)
            self.formulaID_input.setStyleSheet("background-color: white; border: 1px solid black")
            self.formulaID_input.setText(str(result[5]))

            lot_number_input = QtWidgets.QLineEdit()
            lot_number_input.setAlignment(Qt.AlignCenter)
            lot_number_input.setFixedHeight(25)
            lot_number_input.setStyleSheet("background-color: white; border: 1px solid black")

            try:
                lot_number_input.setText('/'.join(result[30]))
            except:
                lot_number_input.setText(None)

            feedRate_input = QtWidgets.QLineEdit()
            feedRate_input.setFixedHeight(25)
            feedRate_input.setStyleSheet("background-color: white; border: 1px solid black")
            feedRate_input.setText(str(result[18]))

            rpm_input = QtWidgets.QLineEdit()
            rpm_input.setFixedHeight(25)
            rpm_input.setStyleSheet("background-color: white; border: 1px solid black")
            rpm_input.setText(str(result[19]))

            screenSize_input = QtWidgets.QLineEdit()
            screenSize_input.setFixedHeight(25)
            screenSize_input.setStyleSheet("background-color: white; border: 1px solid black")
            screenSize_input.setText(result[20])

            screwConf_input = QtWidgets.QComboBox()
            screwConf_input.setFixedHeight(25)
            screwConf_input.setStyleSheet("background-color: white; border: 1px solid black ")
            screwConf_input.addItem('STANDARD')
            screwConf_input.addItem('COIL')
            screwConf_input.setCurrentText(result[17])

            loss_input = QtWidgets.QLineEdit()
            loss_input.setFixedHeight(25)
            loss_input.setEnabled(False)
            loss_input.setAlignment(Qt.AlignCenter)
            loss_input.setStyleSheet("background-color: white; border: 1px solid black")
            loss_input.setText(str(result[12]))

            purgeStart_input = QtWidgets.QLineEdit()
            purgeStart_input.setFixedHeight(25)
            purgeStart_input.setAlignment(Qt.AlignCenter)
            purgeStart_input.setStyleSheet("background-color: white; border: 1px solid black")

            purgeEnd_input = QtWidgets.QLineEdit()
            purgeEnd_input.setFixedHeight(25)
            purgeEnd_input.setAlignment(Qt.AlignCenter)
            purgeEnd_input.setStyleSheet("background-color: white; border: 1px solid black")

            remarks = QtWidgets.QTextEdit()

            operator_input = QtWidgets.QLineEdit()
            operator_input.setFixedHeight(25)
            operator_input.setAlignment(Qt.AlignCenter)
            operator_input.setStyleSheet("background-color: white; border: 1px solid black")
            operator_input.setText(result[21])

            supervisor_input = QtWidgets.QLineEdit()
            supervisor_input.setFixedHeight(25)
            supervisor_input.setAlignment(Qt.AlignCenter)
            supervisor_input.setStyleSheet("background-color: white; border: 1px solid black")
            supervisor_input.setText(result[22])

            order_number_input = QtWidgets.QLineEdit()
            order_number_input.setFixedHeight(25)
            order_number_input.setEnabled(False)
            order_number_input.setAlignment(Qt.AlignCenter)
            order_number_input.setStyleSheet("background-color: white; border: 1px solid black")
            order_number_input.setText(str(result[7]))

            resin_input = QtWidgets.QComboBox()
            resin_input.setFixedHeight(25)
            resin_input.addItem('LDPE(HMI)')
            resin_input.addItem('LDPE(LMI)')
            resin_input.addItem('HDPE(HMI)')
            resin_input.addItem('PP')
            resin_input.addItem('PET')
            resin_input.addItem('GPPS')
            resin_input.setStyleSheet("background-color: white; border: 1px solid black")
            resin_input.setCurrentText(result[15])

            resin_quantity = QLineEdit()
            resin_quantity.setFixedHeight(25)
            resin_quantity.setAlignment(Qt.AlignCenter)
            resin_quantity.setStyleSheet("background-color: white; border: 1px solid black")
            resin_quantity.setText(str(result[33]))

            resin_quantity_label = QLabel()
            resin_quantity_label.setFont(font)
            resin_quantity_label.setText('Resin Qty(kg)')

            purging_input = QtWidgets.QLineEdit()
            purging_input.setFixedHeight(25)
            purging_input.setAlignment(Qt.AlignCenter)
            purging_input.setStyleSheet("background-color: white; border: 1px solid black")
            purging_input.setText(result[14])

            product_input = QtWidgets.QLineEdit()
            product_input.setFixedHeight(25)
            product_input.setAlignment(Qt.AlignCenter)
            product_input.setStyleSheet("background-color: white; border: 1px solid black")
            product_input.textChanged.connect(loss_auto)
            product_input.setText(str(result[29]))

            machineStart_input = QDateTimeEdit()
            machineStart_input.setFixedHeight(25)
            machineStart_input.setStyleSheet("background-color: white; border: 1px solid black")
            try:
                machineStart_input.setDateTime(result[-2])
            except TypeError:
                pass
            machineStart_input.setDisplayFormat("yyyy-MM-dd HH:mm")

            machineOff_input = QDateTimeEdit()
            machineOff_input.setFixedHeight(25)
            machineOff_input.setStyleSheet("background-color: white; border: 1px solid black")
            try:
                machineOff_input.setDateTime(result[-1])
            except TypeError:
                pass
            machineOff_input.setDisplayFormat("yyyy-MM-dd HH:mm")

            self.groupBoxRemarks = QtWidgets.QGroupBox(self.entry_widget)
            self.groupBoxRemarks.setGeometry(600, 500, 200, 150)
            self.groupBoxRemarks.setTitle("Remarks")
            self.groupBoxRemarks.show()

            self.remarks_textBox = QtWidgets.QTextEdit(self.groupBoxRemarks)
            self.remarks_textBox.setGeometry(0, 20, 200, 130)
            self.remarks_textBox.setText(result[16])
            self.remarks_textBox.show()

            # Left Side of Vertical Box
            self.left_vbox.addRow(productionID_label, productionID_input)
            self.left_vbox.addRow(productCode_label, productCode_input)
            self.left_vbox.addRow(customer_label, customer_input)
            self.left_vbox.addRow(orderedQuantity_label, orderedQuantity_input)
            self.left_vbox.addRow(lotnumber_label, lot_number_input)
            self.left_vbox.addRow(product_input_label, product_input)
            self.left_vbox.addRow(productOutput_label, product_output_input)
            self.left_vbox.addRow(loss_label, loss_input)
            self.left_vbox.addRow(machine_label, machine_input)
            self.left_vbox.addRow(formulaID_label, self.formulaID_input)
            self.left_vbox.addRow(order_number_lbl, order_number_input)
            self.left_vbox.addRow(feedrate_label, feedRate_input)

            # Add widgets to the right Form Box
            self.right_vbox.addRow(rpm_label, rpm_input)
            self.right_vbox.addRow(screenSize_label, screenSize_input)
            self.right_vbox.addRow(screwConf_label, screwConf_input)
            self.right_vbox.addRow(purging_label, purging_input)
            self.right_vbox.addRow(resin_label, resin_input)
            self.right_vbox.addRow(resin_quantity_label, resin_quantity)
            self.right_vbox.addRow(purgeStart_label, purgeStart_input)
            self.right_vbox.addRow(purgeEnd_label, purgeEnd_input)
            self.right_vbox.addRow(machineStart_label, machineStart_input)
            self.right_vbox.addRow(machineOff_label, machineOff_input)
            self.right_vbox.addRow(operator_label, operator_input)
            self.right_vbox.addRow(supervisor_label, supervisor_input)

            # Time Table Entry
            time_table = QtWidgets.QTableWidget(self.entry_widget)
            time_table.setGeometry(0, 500, 450, 200)
            time_table.setColumnCount(3)
            time_table.setRowCount(8)
            time_table.setColumnWidth(0, 150)
            time_table.setColumnWidth(1, 150)
            time_table.setStyleSheet("background-color: white;")
            time_table.setFont(QtGui.QFont("Arial", 10))
            time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])

            # Populate the Table
            for i in range(len(result[9])):
                item = QTableWidgetItem(str(result[9][i].strftime("%Y-%m-%d %H:%M")))
                item2 = QTableWidgetItem(str(result[10][i].strftime("%Y-%m-%d %H:%M")))
                item3 = QTableWidgetItem(str(result[26][i]))

                print(type(result[9][i]), str(result[10][i]))
                time_table.setItem(i, 0, item)
                time_table.setItem(i, 1, item2)
                time_table.setItem(i, 2, item3)

            time_table.show()

            # Temperature Table Entry
            temperature_table = QtWidgets.QTableWidget(self.entry_widget)
            temperature_table.setGeometry(450, 500, 150, 200)
            temperature_table.setColumnCount(1)
            temperature_table.setRowCount(12)
            temperature_table.setStyleSheet("background-color: white;")
            temperature_table.setHorizontalHeaderLabels(["Temperature"])
            temperature_index = ["Z" + str(i + 1) for i in range(12)]  # set the index
            temperature_table.setVerticalHeaderLabels(temperature_index)

            # Populate the table
            for i in range(len(result[24])):
                item = QTableWidgetItem(str(result[24][i]))
                temperature_table.setItem(i, 0, item)

            temperature_table.show()

            save_btn = QtWidgets.QPushButton(self.entry_widget)
            save_btn.setGeometry(540, 705, 60, 25)
            save_btn.clicked.connect(update)
            save_btn.setText("Save")
            save_btn.setCursor(Qt.PointingHandCursor)
            save_btn.show()

            default_date = QtCore.QDateTime(2024, 1, 1, 0, 0)

            time_start_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_start_input.setGeometry(30, 475, 120, 25)
            time_start_input.setDisplayFormat("yyyy-MM-dd HH:mm")
            time_start_input.setDateTime(default_date)
            time_start_input.show()

            time_end_input = QtWidgets.QDateTimeEdit(self.entry_widget)
            time_end_input.setGeometry(180, 475, 120, 25)
            time_end_input.setDisplayFormat("yyyy-MM-dd HH:mm")
            time_end_input.setDateTime(default_date)
            time_end_input.show()

            output_lineEdit = QtWidgets.QLineEdit(self.entry_widget)
            output_lineEdit.setGeometry(310, 475, 80, 25)
            output_lineEdit.setAlignment(Qt.AlignCenter)
            output_lineEdit.setStyleSheet("background-color: white; border: 1px solid black")
            output_lineEdit.show()

            self.plus_icon = ClickableLabel(self.entry_widget)
            self.plus_icon.setGeometry(390, 475, 25, 25)
            self.plus_icon.setPixmap(QtGui.QIcon('plus.png').pixmap(25, 25))
            self.plus_icon.setCursor(Qt.PointingHandCursor)
            self.plus_icon.clicked.connect(add_time)
            self.plus_icon.show()

            self.reset_icon = ClickableLabel(self.entry_widget)
            self.reset_icon.setGeometry(425, 475, 25, 25)
            self.reset_icon.setPixmap(QtGui.QIcon('reset.png').pixmap(20, 20))
            self.reset_icon.setCursor(Qt.PointingHandCursor)
            self.reset_icon.clicked.connect(reset_table)
            self.reset_icon.show()

        def print_file():

            try:
                from openpyxl.styles import Font
                from openpyxl.styles import Alignment

                selected = self.extruder_table.selectedItems()

                process_id = selected[0].text()

                self.cursor.execute(f"""
                            SELECT * FROM extruder 
                            WHERE process_id = '{process_id}';

                            """)
                items = self.cursor.fetchall()[0]
                # Unpack the Items
                machine_number = items[1]
                quantity_order = items[2]
                customer = items[4]
                code = items[6]
                total_input = items[29]
                total_time = items[8]
                time_start = items[9]
                time_end = items[10]
                outputPerHour = items[27]
                total_output = items[3]
                outputPercent = items[11]
                loss = items[12]
                lossPercent = items[13]
                purging = items[14]
                resin = items[15]
                remarks = items[16]
                operator = items[21]
                supervisor = items[22]
                outputs = items[26]
                materials = items[23]
                lot_number = items[30]

                wb = load_workbook(r"\\mbpi-server-01\IT\AMIEL\Extruder System\dist\Extruder Template.xlsx")
                worksheet = wb.active

                font = Font(size=8, bold=True, name='Arial')
                center_Alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                worksheet["F5"] = "Extruder Machine No. " + machine_number[-1]
                worksheet["A8"] = machine_number[-1]
                worksheet["B8"] = quantity_order  # quantity order
                worksheet["C8"].font = font
                worksheet["C8"].alignment = center_Alignment
                worksheet["C8"] = customer  # customer

                worksheet["F8"] = code  # product code
                worksheet["G9"] = total_input  # total input
                worksheet["H9"] = total_time  # total time used
                worksheet["I9"] = outputPerHour  # output Per Hour
                worksheet["K9"] = total_output  # total Output
                worksheet["L9"] = outputPercent  # Total Output Percentage
                worksheet["M9"] = loss
                worksheet["N9"] = lossPercent

                total_sec = timedelta()
                for row in range(len(time_start)):
                    worksheet["A" + str(12 + row)] = time_start[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["D" + str(12 + row)] = time_end[row].strftime("%d-%b-%Y %H:%M")
                    worksheet["F" + str(12 + row)] = time_end[row] - time_start[row]
                    worksheet["G" + str(12 + row)] = outputs[row]
                    total_sec = total_sec + (time_end[row] - time_start[row])

                try:
                    hour = str(int(total_sec.total_seconds() // 3600))
                    minute = str((int(total_sec.total_seconds() % 3600) // 60))
                    total_time_used = time(int(hour), int(minute))

                    worksheet["F25"] = total_time_used
                except ValueError:
                    worksheet["F25"] = hour + ":" + minute

                for key in list(materials.keys()):
                    worksheet["I" + str(12 + list(materials.keys()).index(key))] = key
                    worksheet["K" + str(12 + list(materials.keys()).index(key))] = materials[key]

                for ln in range(len(lot_number)):
                    worksheet["M" + str(12 + ln)] = lot_number[ln]

                worksheet["B27"] = purging
                worksheet["B29"] = resin
                worksheet["G26"] = remarks
                worksheet["M28"] = operator
                worksheet["M29"] = supervisor

                # Open QFileDialog

                filename, _ = QFileDialog.getSaveFileName(self.production_widget, "Save File",
                                                          r"C:\Users\Administrator\Documents",
                                                          "Excel Files (*.xlsx)", options=QFileDialog.Options())

                if filename:
                    if not filename.lower().endswith('.xlsx'):
                        filename += '.xlsx'

                    wb.save(filename)
                    QMessageBox.information(self.production_widget, "Export Success", "File Successfully Exported!")

            except Exception as e:
                print(e)
                QMessageBox.information(self.production_widget, "ERROR", "No Selected Items")

        def show_tables():

            try:
                main_time_table.clear()
                material_table.clear()
                lotNumber_table.clear()
            except Exception as e:
                print(e)
            selected = self.extruder_table.selectedItems()
            if selected:
                items = [item.text() for item in selected]
                process_id = items[0]

            try:
                self.cursor.execute(f"""
                            SELECT time_start, time_end, outputs, materials, lot_number
                            FROM extruder
                            WHERE process_id = {process_id};

                            """)
                result = self.cursor.fetchall()
            except:
                QMessageBox.information(self.production_widget, "ERROR", "No Selected Item")
                self.extruder_table.itemSelectionChanged.connect(show_tables)
                return

            t_start = result[0][0]
            t_end = result[0][1]
            outputs = result[0][2]
            materials = result[0][3]
            lotNumber = result[0][4]

            # Set the Header Labels again
            main_time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])
            material_table.setHorizontalHeaderLabels(["Materials", "Quantity(kg)"])
            lotNumber_table.setHorizontalHeaderLabels(["Lot Number"])

            main_time_table.setRowCount(len(t_start))

            # Populating Time Table
            for i in range(len(t_start)):
                item1 = QTableWidgetItem(str(t_start[i].strftime("%b-%d-%Y %H:%M")))
                item1.setTextAlignment(Qt.AlignCenter)
                item1.setFlags(item1.flags() & ~Qt.ItemIsEditable)
                item2 = QTableWidgetItem(str(t_end[i].strftime("%b-%d-%Y %H:%M")))
                item2.setTextAlignment(Qt.AlignCenter)
                item2.setFlags(item2.flags() & ~Qt.ItemIsEditable)
                item3 = QTableWidgetItem(str(outputs[i]))
                item3.setTextAlignment(Qt.AlignCenter)
                item3.setFlags(item3.flags() & ~Qt.ItemIsEditable)
                main_time_table.setItem(i, 0, item1)
                main_time_table.setItem(i, 1, item2)
                main_time_table.setItem(i, 2, item3)

            if len(materials) > material_table.rowCount():
                material_table.setRowCount(len(materials))

            for i in list(materials.keys()):
                key = QTableWidgetItem(str(i))
                key.setFlags(key.flags() & ~Qt.ItemIsEditable)
                value = QTableWidgetItem(str(round((materials[i]), 4)))
                value.setFlags(value.flags() & ~Qt.ItemIsEditable)
                material_table.setItem(list(materials.keys()).index(i), 0, key)
                material_table.setItem(list(materials.keys()).index(i), 1, value)

            for i in range(len(lotNumber)):
                item = QTableWidgetItem(str(lotNumber[i]))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                lotNumber_table.setItem(i, 0, item)

        def import_statistics():

            # Query For Getting the avg of output_per_hour, purge_duration, resin_quantity and yield per machine per product code
            self.cursor.execute(f"""
                SELECT  product_code, machine, ROUND(AVG(output_per_hour), 4) AS average_output_per_hour,
                ROUND(AVG(purge_duration), 2) AS average_cleaning_time, ROUND(AVG(resin_quantity), 2) as average_cleaning_material,
                ROUND(AVG(output_percent), 4) AS ave_yield

                FROM public.extruder
                WHERE time_start[1]::DATE BETWEEN '{date1.text()}' AND '{date2.text()}'

                GROUP BY product_code, machine
                ORDER BY product_code, machine


            """)

            result = self.cursor.fetchall()
            df = pd.DataFrame(result)

            column_names = ["Product Code", "Machine", "average_output_per_hour", "average_cleaning_time",
                            "Average_cleaning_material_used", "ave_yield"]

            try:
                filename, _ = QtWidgets.QFileDialog.getSaveFileName(self.production_widget, "Save File",
                                                                    r"C:\Users\Administrator\Documents",
                                                                    'Excel Files (*.xlsx)',
                                                                    options=QtWidgets.QFileDialog.Options())

                if filename:
                    # Ensuring the file name ends with .xlsx
                    if not filename.lower().endswith('.xlsx'):
                        filename += '.xlsx'

                    # Print or use the file name
                    df.to_excel(excel_writer=filename, index=False,
                                header=column_names)
                    QMessageBox.information(self.production_widget, "File Imported", "Successfully Imported Data")

            except PermissionError:
                QMessageBox.critical(self.production_widget, "Permission Error", "Unable to Export the File. \n "
                                                                                 "Someone is using blank.xlsx")

            wb = load_workbook(fr'{filename}')
            worksheet = wb.active

            # Query for getting the Standard Deviation of average_output_per_hour, average_cleaning_time, average_cleaning_material and ave_yield.
            self.cursor.execute(f"""
            WITH extruder_stats AS (
                SELECT  product_code, machine, ROUND(AVG(output_per_hour), 4) AS average_output_per_hour,
                                ROUND(AVG(purge_duration), 2) AS average_cleaning_time, ROUND(AVG(resin_quantity), 2) as average_cleaning_material,
                                ROUND(AVG(output_percent), 4) AS ave_yield

                                FROM public.extruder
                                WHERE time_start[1]::DATE BETWEEN '{date1.text()}' AND '{date2.text()}'

                                GROUP BY product_code, machine
                                ORDER BY product_code, machine
                )

                SELECT ROUND(STDDEV_POP(average_output_per_hour), 2), ROUND(STDDEV_POP(average_cleaning_time), 2),
                       ROUND(STDDEV_POP(average_cleaning_material), 2), ROUND(STDDEV_POP(ave_yield), 2)
                FROM extruder_stats
            """)
            result = self.cursor.fetchall()
            result = result[0]
            worksheet.column_dimensions['H'].width = 20
            worksheet.column_dimensions['I'].width = 20

            worksheet['H2'] = "Output Per Hour"
            worksheet['H3'] = "Cleaning Time"
            worksheet['H4'] = "Cleaning Material QTY"
            worksheet['H5'] = "Yield"

            worksheet['I1'] = "STANDARD DEVIATION"
            worksheet['I2'] = result[0]
            worksheet['I3'] = result[1]
            worksheet['I4'] = result[2]
            worksheet['I5'] = result[3]

            # Save the Work Book
            wb.save(filename)

        def import_extruder():
            # Setting the column names
            column_names = ['machine', 'date', 'product_code', 'total_input', 'total_output', 'output_percent',
                            'output_per_hour', 'loss', 'loss_percent', 'total_time', 'purge_duration']

            # Query For Getting the raw data of extruder
            self.cursor.execute(f"""
            SELECT machine, TO_CHAR(DATE(time_start[1]), 'MM/DD/YYYY'), product_code,
            total_input, total_output, output_percent, output_per_hour, loss, loss_percent, 
            total_time, purge_duration  
            FROM extruder
            WHERE time_start[1]::DATE BETWEEN '{date1.text()}' AND '{date2.text()}'

            ORDER BY machine, time_start[1]

            """)

            result = self.cursor.fetchall()

            df = pd.DataFrame(result)

            try:
                filename, _ = QtWidgets.QFileDialog.getSaveFileName(self.production_widget, "Save File",
                                                                    r"C:\Users\Administrator\Documents",
                                                                    'Excel Files (*.xlsx)',
                                                                    options=QtWidgets.QFileDialog.Options())

                if filename:
                    # Ensuring the file name ends with .xlsx
                    if not filename.lower().endswith('.xlsx'):
                        filename += '.xlsx'

                    # Print or use the file name
                    df.to_excel(excel_writer=filename, index=False,
                                header=column_names)
                    QMessageBox.information(self.production_widget, "File Imported", "Successfully Imported Data")
            except PermissionError:
                QMessageBox.critical(self.production_widget, "Permission Error", "Unable to Export the File. \n "
                                                                                 "Someone is using blank.xlsx")

        def compounding():

            def save_entry():
                # Compute for the Time Difference to get the Duration

                mix_duration = timedelta()
                try:
                    mix_start = datetime.strptime(time_start_input.text(), "%H:%M")
                    mix_end = datetime.strptime(time_end_input.text(), "%H:%M")

                    mix_duration = mix_end - mix_start
                    hours = int(str(mix_duration).split(':')[0])
                    minutes = int(str(mix_duration).split(':')[1])
                    mix_duration = (hours * 60) + minutes
                except Exception as e:
                    print(e)

                print(mix_duration)
                try:
                    self.cursor.execute(f"""
                                    INSERT INTO mixer
                                    VALUES({production_id_input.text()}, {machine_input.currentText()}::INTEGER, '{product_code_input.currentText()}', {mix_duration},
                                            '{operator_input.currentText()}', {material_output.text()}, {material_input.text()},
                                            '{remarks_box.toPlainText()}', '{date_input.text()}')              
                """)
                    self.conn.commit()
                    QMessageBox.information(self.compounding_widget, "SUCCESS", "Data Successfully Added.")
                    clear_inputs()
                except Exception as e:
                    print(e)

            def clear_inputs():
                product_code_input.clear()
                operator_input.clear()
                material_input.clear()
                material_output.clear()
                remarks_box.clear()
                production_id_input.clear()
                product_code_input.clear()

            self.compounding_widget = QWidget(self.main_widget)
            self.compounding_widget.setGeometry(0, 0, 991, 700)
            self.compounding_widget.setStyleSheet("background-color: white;")
            self.compounding_widget.show()

            buttons_widget = QWidget(self.compounding_widget)
            buttons_widget.setGeometry(0, 0, 991, 30)
            buttons_widget.setStyleSheet('background-color: rgb(92, 154, 255)')
            buttons_widget.show()

            # Extruder Tab Button
            extruder_btn = QPushButton(buttons_widget)
            extruder_btn.setGeometry(50, 0, 75, 30)
            extruder_btn.setText('EXTRUDER')
            extruder_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            extruder_btn.setStyleSheet('color: white; border: none;')
            extruder_btn.setCursor(Qt.PointingHandCursor)
            extruder_btn.clicked.connect(self.production)
            extruder_btn.show()

            compounding_btn = QPushButton(buttons_widget)
            compounding_btn.setGeometry(150, 0, 50, 30)
            compounding_btn.setText('MIXER')
            compounding_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            compounding_btn.setStyleSheet(
                'color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;')
            compounding_btn.setCursor(Qt.PointingHandCursor)
            compounding_btn.clicked.connect(compounding)
            compounding_btn.show()

            title_widget = QWidget(self.compounding_widget)
            title_widget.setGeometry(0, 30, 991, 50)
            title_widget.show()

            title_label = QLabel(title_widget)
            title_label.setGeometry(493, 10, 260, 30)
            title_label.setText('MIXER REPORT')
            title_label.setStyleSheet('color: rgb(92, 154, 255)')
            title_label.setFont(QtGui.QFont('Rockwell Extra Bold', 20))
            title_label.setAlignment(Qt.AlignCenter)
            title_label.show()

            input_widget = QWidget(self.compounding_widget)
            input_widget.setGeometry(0, 80, 320, 571)
            input_widget.setStyleSheet('border: 1px solid black;')

            mixer_table = QTableWidget(self.compounding_widget)
            mixer_table.setGeometry(320, 80, 671, 621)
            mixer_table.setColumnCount(8)
            mixer_table.setRowCount(25)
            mixer_table.verticalHeader().setVisible(False)
            mixer_table.setHorizontalHeaderLabels(
                ['Date', 'Production ID', 'Machine', 'Product Code', 'Total Time', 'Processed By', 'Output', 'Remarks'])
            mixer_table.horizontalHeader().setStyleSheet("""
                QHeaderView::section{
                font-weight: bold;
                background-color: rgb(0, 109, 189);
                color: white;
                }

        """)

            self.cursor.execute("""
                SELECT date, production_id, machine, product_code, time_consumed, processed_by, output, remarks 
                FROM mixer

            """)
            result = self.cursor.fetchall()

            # Populate table with data
            for i in range(len(result)):
                for j in range(8):
                    item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    mixer_table.setItem(i, j, item)

            mixer_table.setColumnWidth(1, 90)
            mixer_table.setColumnWidth(2, 70)
            mixer_table.setColumnWidth(4, 70)
            mixer_table.setColumnWidth(7, 120)

            mixer_table.show()

            font = QtGui.QFont('Arial', 11)

            production_id_label = QLabel()
            production_id_label.setText('Production ID')
            production_id_label.setFixedHeight(25)
            production_id_label.setFixedWidth(120)
            production_id_label.setFont(font)
            production_id_label.setStyleSheet('border: none')

            machine_number_label = QLabel()
            machine_number_label.setText('Machine No.')
            machine_number_label.setFixedHeight(25)
            machine_number_label.setFixedWidth(120)
            machine_number_label.setFont(font)
            machine_number_label.setStyleSheet('border: none')

            product_code_label = QLabel()
            product_code_label.setText('Product Code')
            product_code_label.setFixedHeight(25)
            product_code_label.setFixedWidth(120)
            product_code_label.setFont(font)
            product_code_label.setStyleSheet('border: none')

            time_start_label = QLabel()
            time_start_label.setText('Time Start')
            time_start_label.setFixedHeight(25)
            time_start_label.setFixedWidth(120)
            time_start_label.setFont(font)
            time_start_label.setStyleSheet('border: none')

            time_end_label = QLabel()
            time_end_label.setText('Time End')
            time_end_label.setFixedHeight(25)
            time_end_label.setFixedWidth(120)
            time_end_label.setFont(font)
            time_end_label.setStyleSheet('border: none')

            operator_label = QLabel()
            operator_label.setText('Operator')
            operator_label.setFixedHeight(25)
            operator_label.setFixedWidth(120)
            operator_label.setFont(font)
            operator_label.setStyleSheet('border: none')

            input_label = QLabel()
            input_label.setText('Input')
            input_label.setFixedHeight(25)
            input_label.setFixedWidth(120)
            input_label.setFont(font)
            input_label.setStyleSheet('border: none')

            output_label = QLabel()
            output_label.setText('Output')
            output_label.setFixedHeight(25)
            output_label.setFixedWidth(120)
            output_label.setFont(font)
            output_label.setStyleSheet('border: none')

            date_label = QLabel()
            date_label.setText('Date')
            date_label.setFixedHeight(25)
            date_label.setFixedWidth(120)
            date_label.setFont(font)
            date_label.setStyleSheet('border: none')

            remarks_label = QLabel()
            remarks_label.setText('Remarks')
            remarks_label.setFixedHeight(25)
            remarks_label.setFixedWidth(120)
            remarks_label.setFont(font)
            remarks_label.setStyleSheet('border: none')

            background_color = 'background-color: rgb(255, 255, 0)'

            production_id_input = QLineEdit()
            production_id_input.setFixedHeight(25)
            production_id_input.setStyleSheet(background_color)
            production_id_input.setAlignment(Qt.AlignCenter)

            machine_input = QComboBox()
            machine_input.setStyleSheet(background_color)
            machine_input.setFixedHeight(25)
            machine_input.addItem('1')
            machine_input.addItem('2')
            machine_input.addItem('3')
            machine_input.addItem('4')
            machine_input.addItem('5')
            machine_input.addItem('6')
            machine_input.setFont(QtGui.QFont('Arial', 10))

            product_code_input = QComboBox()
            product_code_input.setStyleSheet(background_color)
            product_code_input.setEditable(True)
            product_code_input.setFixedHeight(25)
            time_start_input = QTimeEdit()
            time_start_input.setStyleSheet(background_color)
            time_start_input.setFixedHeight(25)
            time_start_input.setDisplayFormat('HH:mm')
            time_start_input.setAlignment(Qt.AlignCenter)
            time_end_input = QTimeEdit()
            time_end_input.setStyleSheet(background_color)
            time_end_input.setFixedHeight(25)
            time_end_input.setDisplayFormat('HH:mm')
            time_end_input.setAlignment(Qt.AlignCenter)
            operator_input = QComboBox()
            operator_input.setStyleSheet(background_color)
            operator_input.setEditable(True)
            operator_input.setFixedHeight(25)
            material_input = QLineEdit()
            material_input.setStyleSheet(background_color)
            material_input.setFixedHeight(25)
            material_output = QLineEdit()
            material_output.setStyleSheet(background_color)
            material_output.setFixedHeight(25)
            date_input = QDateEdit()
            date_input.setStyleSheet(background_color)
            date_input.setFixedHeight(25)
            date_input.setAlignment(Qt.AlignCenter)
            date_input.setDisplayFormat('MM-dd-yyyy')
            remarks_box = QTextEdit()
            remarks_box.setStyleSheet(background_color)
            remarks_box.setFixedHeight(100)

            form_layout = QFormLayout(input_widget)

            form_layout.addRow(production_id_label, production_id_input)
            form_layout.addRow(machine_number_label, machine_input)
            form_layout.addRow(product_code_label, product_code_input)
            form_layout.addRow(date_label, date_input)
            form_layout.addRow(time_start_label, time_start_input)
            form_layout.addRow(time_end_label, time_end_input)
            form_layout.addRow(operator_label, operator_input)
            form_layout.addRow(input_label, material_input)
            form_layout.addRow(output_label, material_output)
            form_layout.addRow(remarks_label, remarks_box)

            input_widget.show()

            save_btn = QPushButton(input_widget)
            save_btn.setGeometry(80, 540, 60, 25)
            save_btn.setText('SAVE')
            save_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                   "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            save_btn.setCursor(Qt.PointingHandCursor)
            save_btn.clicked.connect(save_entry)
            save_btn.show()

            clear_btn = QPushButton(input_widget)
            clear_btn.setGeometry(190, 540, 60, 25)
            clear_btn.setText('CLEAR')
            clear_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                    "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            clear_btn.setCursor(Qt.PointingHandCursor)
            clear_btn.clicked.connect(clear_inputs)
            clear_btn.show()

        def search_lot_number():
            self.cursor.execute(f"""
                SELECT 
                process_id, TO_CHAR(DATE(time_start[1]), 'MM/DD/YYYY') as date,machine, customer, product_code, 
                        qty_order, total_output, output_per_hour,total_time, formula_id
                FROM extruder
                WHERE '{search_bar.text()}' = ANY(lot_number) OR product_code ILIKE '%{search_bar.text()}%'
                OR customer ILIKE '%{search_bar.text()}%' OR machine ILIKE '%{search_bar.text()}%'
                ORDER BY process_id DESC
            """)
            result = self.cursor.fetchall()
            self.extruder_table.clearContents()

            # Populate table with data
            for i in range(len(result)):
                for j in range(len(column_names)):
                    item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                    # Set Alignment for specific columns
                    if j == 2 or j == 6 or j == 3 or j == 4 or j == 7 or j == 5:
                        item.setTextAlignment(Qt.AlignCenter)
                    else:
                        pass
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.extruder_table.setItem(i, j, item)

        def delete_entry():

            if extruder_access:
                pass
            else:
                QMessageBox.critical(self.production_widget, 'Restricted Access',
                                     "You Dont Have Permission. \n Contact the Admin.")
                return

            selected = self.extruder_table.selectedItems()

            if selected:
                id = selected[0].text()

                delete = QMessageBox.question(self.production_widget, 'Delete Item',
                                              f"Do you want to Delete \n process id {id}?",
                                              QMessageBox.Yes | QMessageBox.No,
                                              QMessageBox.No)

                if delete == QMessageBox.Yes:
                    self.cursor.execute(f"""
                        DELETE FROM extruder
                        WHERE process_id = {id}

                                    """)
                    self.conn.commit()

                else:
                    print("User chose No.")
                    return

            else:
                QMessageBox.critical(self.production_widget, 'Error', "No Item Selected!")
                return

        try:
            if self.mixer_btn_clicked == True:
                compounding()
                self.mixer_btn_clicked = False
                return
            else:
                pass

        except:
            pass

        def auto_search():

            self.cursor.execute(f"""
                SELECT 
                process_id, TO_CHAR(DATE(time_start[1]), 'MM/DD/YYYY') as date,machine, customer, product_code, 
                        qty_order, total_output, output_per_hour,total_time, formula_id
                FROM extruder
                WHERE '{search_bar.text()}' = ANY(lot_number) OR product_code ILIKE '%{search_bar.text()}%'
                OR customer ILIKE '%{search_bar.text()}%' OR machine ILIKE '%{search_bar.text()}%'
                ORDER BY time_start[1] DESC

            """)

            result = self.cursor.fetchall()
            self.extruder_table.clearContents()

            # Populate table with data
            for i in range(len(result)):
                for j in range(len(column_names)):
                    item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                    # Set Alignment for specific columns
                    if j == 2 or j == 6 or j == 3 or j == 4 or j == 7 or j == 5:
                        item.setTextAlignment(Qt.AlignCenter)
                    else:
                        pass
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                    self.extruder_table.setItem(i, j, item)

            self.extruder_table.verticalScrollBar().setValue(0)

        self.production_widget = QtWidgets.QWidget(self.main_widget)
        self.production_widget.setGeometry(0, 0, 991, 700)
        self.production_widget.setStyleSheet("background-color: rgb(245, 246, 248);")
        self.production_widget.show()

        buttons_widget = QWidget(self.production_widget)
        buttons_widget.setGeometry(0, 0, 991, 30)
        buttons_widget.setStyleSheet('background-color: rgb(92, 154, 255)')
        buttons_widget.show()

        search_bar = QLineEdit(self.production_widget)
        search_bar.setGeometry(705, 35, 150, 25)
        search_bar.setStyleSheet('background-color: rgb(255, 255, 17); border: 1px solid rgb(171, 173, 179)')
        search_bar.setPlaceholderText('Lot Number')
        search_bar.setFocus()
        search_bar.textChanged.connect(auto_search)
        search_bar.show()

        search_button = QPushButton(self.production_widget)
        search_button.setGeometry(860, 35, 60, 25)
        search_button.setText('Search')
        search_button.setStyleSheet('border: 1px solid rgb(171, 173, 179)')
        search_button.setShortcut('Return')
        search_button.clicked.connect(search_lot_number)
        search_button.show()

        # Extruder Tab Button
        extruder_btn = QPushButton(buttons_widget)
        extruder_btn.setGeometry(50, 0, 75, 30)
        extruder_btn.setText('EXTRUDER')
        extruder_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
        extruder_btn.setStyleSheet('color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;')
        extruder_btn.setCursor(Qt.PointingHandCursor)
        extruder_btn.clicked.connect(self.production)
        extruder_btn.show()

        compounding_btn = QPushButton(buttons_widget)
        compounding_btn.setGeometry(150, 0, 50, 30)
        compounding_btn.setText('MIXER')
        compounding_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
        compounding_btn.setStyleSheet('color: white; border: none')
        compounding_btn.setCursor(Qt.PointingHandCursor)
        compounding_btn.clicked.connect(compounding)
        compounding_btn.show()

        self.extruder_table = QtWidgets.QTableWidget(self.production_widget)
        self.extruder_table.setGeometry(QtCore.QRect(20, 70, 900, 375))
        self.extruder_table.verticalHeader().setVisible(False)
        self.extruder_table.setSortingEnabled(True)

        self.cursor.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE TABLE_NAME = 'extruder';
        """)

        column_names = ["process_id", "date", "machine", "customer", "product_code", "qty_order", "total_output",
                        "output_per_hour", "total time(hr)",
                        "formula_id"]

        try:
            self.cursor.execute("""SELECT 
                        process_id, TO_CHAR(DATE(time_start[1]), 'MM/DD/YYYY') as date,machine, customer, product_code, 
                        qty_order, total_output, output_per_hour,total_time, formula_id
                        FROM extruder
                        ORDER BY date DESC;
                        """)
            result = self.cursor.fetchall()
        except Exception as e:
            self.cursor.execute("""
                        SELECT 
                        process_id, machine, customer, qty_order, total_output, output_per_hour, formula_id, product_code, total_time
                        FROM extruder
                        ORDER BY process_id DESC
                        ; 

                        """)
            result = self.cursor.fetchall()

        # Set Column Count
        self.extruder_table.setColumnCount(len(column_names))
        # Set Row Count
        self.extruder_table.setRowCount(len(result))

        self.extruder_table.setStyleSheet("""

        color : black;
        background-color: rgb(255, 255, 255);
        """)

        # Populate table with data
        for i in range(len(result)):
            for j in range(len(column_names)):
                item = QtWidgets.QTableWidgetItem(str(result[i][j]))  # Convert to string
                if j == 3:
                    pass
                else:
                    item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make the cells unable to be edited
                self.extruder_table.setItem(i, j, item)

        bold_font = QtGui.QFont()
        bold_font.setBold(True)
        self.extruder_table.horizontalHeader().setFont(bold_font)

        self.extruder_table.horizontalHeader().setStyleSheet("""
        QHeaderView::section{
        font-weight: bold;
        background-color: rgb(0, 109, 189);
        color: white;
        }

        """)

        # Set Column Width
        self.extruder_table.setColumnWidth(0, 70)
        self.extruder_table.setColumnWidth(1, 90)
        self.extruder_table.setColumnWidth(2, 80)
        self.extruder_table.setColumnWidth(3, 150)
        self.extruder_table.setColumnWidth(9, 90)

        self.extruder_table.setHorizontalHeaderLabels([col.upper() for col in column_names])  # Set column names
        # Set selection mode to select entire rows and disable single item selection
        self.extruder_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.extruder_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.extruder_table.itemSelectionChanged.connect(show_tables)
        self.extruder_table.show()

        date1 = QDateEdit(self.production_widget)
        date1.setGeometry(50, 660, 100, 25)
        date1.setDate(date.today())
        date1.setDisplayFormat('yyyy-MM-dd')
        date1.show()

        date2 = QDateEdit(self.production_widget)
        date2.setGeometry(170, 660, 100, 25)
        date2.setDate(date.today())
        date2.setDisplayFormat('yyyy-MM-dd')
        date2.show()

        import_extruder_btn = ClickableLabel(self.production_widget)
        import_extruder_btn.setGeometry(280, 660, 20, 20)
        import_extruder_btn.setPixmap(QtGui.QIcon('export.png').pixmap(20, 20))
        import_extruder_btn.setCursor(Qt.PointingHandCursor)
        import_extruder_btn.setToolTip('Import to Excel')
        import_extruder_btn.clicked.connect(import_extruder)
        import_extruder_btn.show()

        statsButtonImport = ClickableLabel(self.production_widget)
        statsButtonImport.setGeometry(310, 660, 20, 20)
        statsButtonImport.setPixmap(QtGui.QIcon('statistics-icon.png').pixmap(20, 20))
        statsButtonImport.setCursor(Qt.PointingHandCursor)
        statsButtonImport.setToolTip('Export Extruder Data')
        statsButtonImport.clicked.connect(import_statistics)
        statsButtonImport.show()

        self.view_btn = QtWidgets.QPushButton(self.production_widget)
        self.view_btn.setGeometry(540, 660, 60, 25)
        self.view_btn.setText("View")
        self.view_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                    "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
        self.view_btn.clicked.connect(show_form)
        self.view_btn.setCursor(Qt.PointingHandCursor)
        self.view_btn.show()

        self.add_btn = QtWidgets.QPushButton(self.production_widget)
        self.add_btn.setGeometry(605, 660, 60, 25)
        self.add_btn.setText("Add")
        self.add_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                   " border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
        self.add_btn.clicked.connect(add_entry)
        self.add_btn.setCursor(Qt.PointingHandCursor)
        self.add_btn.show()

        self.update_btn = QtWidgets.QPushButton(self.production_widget)
        self.update_btn.setGeometry(670, 660, 60, 25)
        self.update_btn.setText("Update")
        self.update_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                      "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
        self.update_btn.clicked.connect(update_entry)
        self.update_btn.setCursor(Qt.PointingHandCursor)
        self.update_btn.show()

        self.delete_btn = QtWidgets.QPushButton(self.production_widget)
        self.delete_btn.setGeometry(735, 660, 60, 25)
        self.delete_btn.setText("Delete")
        self.delete_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                      "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
        self.delete_btn.clicked.connect(delete_entry)
        self.delete_btn.setCursor(Qt.PointingHandCursor)
        self.delete_btn.show()

        self.print_btn = QtWidgets.QPushButton(self.production_widget)
        self.print_btn.setGeometry(800, 660, 60, 25)
        self.print_btn.setText("Print")
        self.print_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                     "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
        self.print_btn.clicked.connect(print_file)
        self.print_btn.setCursor(Qt.PointingHandCursor)
        self.print_btn.show()

        main_time_table = QtWidgets.QTableWidget(self.production_widget)
        main_time_table.setGeometry(20, 455, 460, 195)
        main_time_table.setColumnCount(3)
        main_time_table.setRowCount(6)
        main_time_table.setHorizontalHeaderLabels(["Time Start", "Time End", "Output"])
        main_time_table.verticalHeader().setVisible(False)
        main_time_table.horizontalHeader().setStyleSheet("""
                QHeaderView::section{
                font-weight: bold;
                background-color: rgb(0, 109, 189);
                color: white;
                }

                """)
        main_time_table.setColumnWidth(0, 180)
        main_time_table.setColumnWidth(1, 180)
        main_time_table.setColumnWidth(2, 98)
        main_time_table.setStyleSheet('background-color: white')
        main_time_table.show()

        material_table = QtWidgets.QTableWidget(self.production_widget)
        material_table.setGeometry(495, 455, 245, 195)
        material_table.setColumnCount(2)
        material_table.setRowCount(13)
        material_table.setHorizontalHeaderLabels(["Material", "Value(Kg)"])
        material_table.verticalHeader().setVisible(False)
        material_table.horizontalHeader().setStyleSheet("""
                QHeaderView::section{
                font-weight: bold;
                background-color: rgb(0, 109, 189);
                color: white;
                }

                """)
        material_table.setStyleSheet('background-color: white;')
        material_table.setColumnWidth(0, 120)
        material_table.setColumnWidth(1, 100)
        material_table.show()

        lotNumber_table = QtWidgets.QTableWidget(self.production_widget)
        lotNumber_table.setGeometry(758, 455, 162, 195)
        lotNumber_table.setColumnCount(1)
        lotNumber_table.setRowCount(6)
        lotNumber_table.setHorizontalHeaderLabels(["Lot Number"])
        lotNumber_table.verticalHeader().setVisible(False)
        lotNumber_table.horizontalHeader().setStyleSheet("""
                QHeaderView::section{
                font-weight: bold;
                background-color: rgb(0, 109, 189);
                color: white;
                }

                """)
        lotNumber_table.setStyleSheet('background-color: white;')
        lotNumber_table.setColumnWidth(0, 162)
        lotNumber_table.show()

    def quality_control(self):

        import pandas as pd
        import matplotlib.pyplot as plt
        from psycopg2 import sql

        try:
            self.production_widget.deleteLater()
            body_widget.deleteLater()

        except Exception as e:
            print(e)

        def exportBtn_clicked():
            date_from = date1.text()
            date_to = date2.text()
            self.cursor.execute(f"""
            SELECT * FROM quality_control
            WHERE evaluated_on::DATE BETWEEN '{date_from}' AND '{date_to}'
            ORDER BY id DESC

            """)

            result = self.cursor.fetchall()

            df = pd.DataFrame(result)
            print(df.dtypes)
            # Get the Column Names
            self.cursor.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE TABLE_NAME = 'quality_control'

            """)
            result = self.cursor.fetchall()
            column_names = ['id', 'lot_number', 'product_code', 'customer', 'status', 'remarks', 'action',
                            'original_lot', 'evaluated_by', 'evaluated_on', 'encoded_on', 'updated_by',
                            'updated_on', 'time_endorsed', 'edited', 'qc_type', 'formula_id', 'status_changed']

            try:

                filename, _ = QtWidgets.QFileDialog.getSaveFileName(self.qc_widget, "Save File",
                                                                    r"C:\Users\Administrator\Documents",
                                                                    'Excel Files (*.xlsx)',
                                                                    options=QtWidgets.QFileDialog.Options())

                if filename:
                    # Ensuring the file name ends with .xlsx
                    if not filename.lower().endswith('.xlsx'):
                        filename += '.xlsx'

                    # Print or use the file name
                    df.to_excel(excel_writer=filename, index=False,
                                header=column_names)
                    QMessageBox.information(self.qc_widget, "File Imported", "Successfully Imported Data")

            except PermissionError:
                QMessageBox.critical(self.qc_widget, "Permission Error", "Unable to Export the File. \n "
                                                                         "Someone is using blank.xlsx")

        def edited_checkbox_changed():
            # Select Only Edited Data
            if edited_checkbox.isChecked():
                self.qc_table.itemSelectionChanged.disconnect(show_items)
                self.qc_table.clearSelection()
                self.qc_table.clear()
                self.qc_table.setHorizontalHeaderLabels(['ID', 'Lot Number', 'Customer', 'Product Code',
                                                         'Status', 'Remarks', 'Action Taken'])
                self.cursor.execute("""
                SELECT id, lot_number, customer, product_code, status, remarks, action
                FROM quality_control
                WHERE edited = 'True'

                """)

                result = self.cursor.fetchall()

                # Populate the table
                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.qc_table.setItem(i, j, item)
                self.qc_table.itemSelectionChanged.connect(show_items)
                self.qc_table.show()
            else:
                # Select ALL
                self.qc_table.itemSelectionChanged.disconnect(show_items)
                self.qc_table.clearSelection()
                self.qc_table.clear()
                self.qc_table.setHorizontalHeaderLabels(['ID', 'Lot Number', 'Customer', 'Product Code',
                                                         'Status', 'Remarks', 'Action Taken'])
                self.cursor.execute("""
                                SELECT id, lot_number, customer, product_code, status, remarks, action 
                                FROM quality_control

                                """)

                result = self.cursor.fetchall()

                # Populate the table
                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.qc_table.setItem(i, j, item)

                self.qc_table.itemSelectionChanged.connect(show_items)
                self.qc_table.show()

        def auto_search():
            self.cursor.execute(f"""
                SELECT id, lot_number, customer, product_code, status, remarks, action
                FROM quality_control
                WHERE lot_number ILIKE '%{search_bar.text()}%' OR product_code ILIKE '%{search_bar.text()}%'
                OR customer ILIKE '%{search_bar.text()}%' OR status ILIKE '%{search_bar.text()}%'
                ORDER BY id DESC

                                        """)
            result = self.cursor.fetchall()
            self.qc_table.clearContents()

            # Populate the table
            for i in range(len(result)):
                for j in range(len(result[i])):
                    item = QTableWidgetItem(str(result[i][j]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                    if result[i][4] == 'Failed':
                        item.setBackground(QtGui.QColor(255, 128, 128))
                    else:
                        pass

                    self.qc_table.setItem(i, j, item)
            self.qc_table.verticalScrollBar().setValue(0)

        def evaluation_entry():

            # Check if User have permission to enter QC data.
            if qc_access:
                pass
            else:
                QMessageBox.critical(self.qc_widget, 'Restricted Access',
                                     "You Dont Have Permission. \n Contact the Admin.")
                return

            # Getting every single Lot Number in a Multiple Lot Number
            new_lot_list = []
            old_lot_list = []

            # Getting the new Lot List
            def multiple_lotNumber():
                try:
                    new_lot_list.clear()
                    lotNumbers_board.clear()
                    if '-' in lotNumber_input.text() and lotNumber_input.text()[-1] != ')':
                        start_lot = re.findall(r'\d+', lotNumber_input.text())[0]
                        end_lot = re.findall(r'\d+', lotNumber_input.text())[1]
                        string_code = re.findall(r'[a-zA-Z]+', lotNumber_input.text())[0]

                        for i in range(int(start_lot), int(end_lot) + 1):
                            while len(str(i)) != 4:
                                i = '0' + str(i)
                            new_lot_list.append((str(i) + string_code).upper())

                        lotNumbers_board.setText("\n".join(new_lot_list))
                    else:
                        new_lot_list.append(lotNumber_input.text().strip())
                        lotNumbers_board.clear()

                except:
                    print("INVALID")

            # For getting multiple old Lot Numbers from correction input
            def get_old_lotNumbers():

                try:
                    # Clear Old Lot List
                    old_lot_list.clear()

                    correction = correction_input.text().strip()
                    if ',' in correction:
                        correction = correction.split(',')
                        for i in correction:
                            if '-' in i and i[-1] != ')':  # Check if it is a multiple Lot
                                start_lot = re.findall(r'\d+', i)[0]
                                end_lot = re.findall(r'\d+', i)[1]
                                string_code = re.findall('[a-zA-Z]+', i)[0]

                                for num in range(int(start_lot), int(end_lot) + 1):
                                    old_lot_list.append(str(num) + string_code)
                            else:
                                old_lot_list.append(i)

                    elif '-' in correction_input.text() and correction_input.text()[-1] != ')':
                        start_lot = re.findall(r'\d+', correction_input.text())[0]
                        end_lot = re.findall(r'\d+', correction_input.text())[1]
                        string_code = re.findall('[a-zA-Z]+', correction_input.text())[0]

                        for i in range(int(start_lot), int(end_lot) + 1):
                            old_lot_list.append(str(i) + string_code)

                    else:
                        old_lot_list.append(correction_input.text().strip())
                except:
                    pass

            def saveBtn_clicked():
                def clear_entries():
                    lotNumber_input.clear()
                    productCode_dropdown.setCurrentIndex(0)
                    formulaID_input.clear()
                    result_dropdown.setCurrentIndex(0)
                    remarks_box.clear()

                if len(new_lot_list) > 100:
                    QMessageBox.information(self.qc_widget, 'ERROR', 'Too Much Lot Number')
                    return

                try:
                    # For saving Multiple Lot Number in quality_control_tbl2

                    self.cursor.execute("SELECT MAX(id) FROM quality_control")
                    qc_ID = self.cursor.fetchone()[0]

                    if qcType_dropdown.currentText() == "NEW":

                        self.cursor.execute(f"""
                        SELECT lot_number 
                        FROM quality_control
                        WHERE lot_number = '{lotNumber_input.text()}'

                        """)
                        result = self.cursor.fetchone()

                        # This if statement break the cancel the saving if the lot number is already in the Database
                        if result:
                            QMessageBox.information(self.qc_widget, "Data Exist", "Data is already Entered.")
                            return

                        self.cursor.execute(f"""
                                           INSERT INTO quality_control
                                           (lot_number, product_code, customer, status, remarks, action, original_lot, evaluated_by,
                                           evaluated_on, encoded_on, time_endorsed, qc_type, formula_id)
                                           VALUES('{lotNumber_input.text().strip()}', '{productCode_dropdown.currentText().strip()}', '{customers_box.currentText()}',
                                           '{result_dropdown.currentText()}', '{remarks_box.toPlainText()}', '{actionTake_box.toPlainText()}',
                                           '{lotNumber_input.text()}', '{evaluatedBy_dropdown.currentText()}', '{date_started_input.text()}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}',
                                           '{time_endorsed_input.text()}', '{qcType_dropdown.currentText()}', '{formulaID_input.text()}' )
                                           """)

                        # For saving Multiple Lot Number in quality_control_tbl2
                        self.cursor.execute("SELECT MAX(id) FROM quality_control")
                        qc_ID = self.cursor.fetchone()[0]

                        if "-" in lotNumber_input.text() and lotNumber_input.text()[-1] != ')':
                            for lot in new_lot_list:
                                self.cursor.execute(f"""
                                        INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot,
                                         status, product_code, qc_type, formula_id, date_endorsed, encoded_on)
                                        VALUES('{qc_ID}','{lot.strip()}', '{date_started_input.text()}', '{lot.strip()}',
                                        '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}', 
                                        '{qcType_dropdown.currentText()}', '{formulaID_input.text()}', '{time_endorsed_input.text()}',
                                        '{datetime.now().strftime("%Y-%m-%d %H:%M")}')
                                                    """)
                            self.conn.commit()
                            print("Query Successful")
                            clear_entries()  # Clear the entries after successful entry
                        else:
                            self.cursor.execute(f"""
                                    INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot, status, 
                                    product_code, qc_type, formula_id, date_endorsed, encoded_on)
                                    VALUES('{qc_ID}', '{lotNumber_input.text().strip()}', '{date_started_input.text()}', '{lotNumber_input.text()}',
                                    '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}', '{qcType_dropdown.currentText()}', 
                                    '{formulaID_input.text()}', '{time_endorsed_input.text()}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}')    """)
                            self.conn.commit()
                            clear_entries()

                        QMessageBox.information(self.body_widget.setStyleSheet("border: none;"), "Query Success",
                                                "QC Entry Added")

                        self.cursor.execute("""
                        SELECT MAX(id) FROM quality_control

                        """)
                        # Update the QC ID
                        current_id = self.cursor.fetchone()[0]
                        qcControl_input.setText(str(current_id + 1))
                        qcControl_input.show()

                        new_lot_list.clear()
                        old_lot_list.clear()

                        lotNumber_input.setFocus()

                    else:  # CORRECTION
                        # SAVE TO THE FIRST QC TABLE 1
                        self.cursor.execute(f""" SELECT original_lot FROM quality_control
                                            WHERE original_lot = '{correction_input.text()}'
                                       """)
                        result = self.cursor.fetchone()
                        try:
                            if result == None:
                                orig_lot = correction_input.text().strip()
                            else:
                                orig_lot = result[0]
                        except:
                            orig_lot = correction_input.text().strip()

                        self.cursor.execute(f"""
                                       INSERT INTO quality_control
                                           (lot_number, product_code, customer, status, remarks, action, original_lot, evaluated_by,
                                           evaluated_on, encoded_on, time_endorsed, qc_type, updated_on, formula_id)
                                           VALUES('{lotNumber_input.text().strip()}', '{productCode_dropdown.currentText().strip()}', '{customers_box.currentText()}',
                                           '{result_dropdown.currentText()}', '{remarks_box.toPlainText()}', '{actionTake_box.toPlainText()}',
                                           '{orig_lot.strip()}', '{evaluatedBy_dropdown.currentText()}', '{date_started_input.text()}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}',
                                          ' {time_endorsed_input.text()}', '{qcType_dropdown.currentText()}', 
                                            '{datetime.now().strftime("%Y-%m-%d %H:%M")}',
                                          '{formulaID_input.text()}')

                                       """)

                        # For saving Multiple Lot Number in quality_control_tbl2

                        self.cursor.execute("SELECT MAX(id) FROM quality_control")
                        qc_ID = self.cursor.fetchone()[0]

                        # Save To quality_control_tbl2 DB
                        if len(old_lot_list) == len(new_lot_list):
                            for i in range(len(old_lot_list)):
                                self.cursor.execute(f"""
                                SELECT original_lot FROM quality_control_tbl2
                                WHERE lot_number = '{old_lot_list[i]}'

                                """)

                                result = self.cursor.fetchall()
                                try:
                                    orig_lot = result[0][0]
                                except IndexError:
                                    orig_lot = old_lot_list[i]

                                self.cursor.execute(f"""
                                    INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot, status,
                                    product_code, qc_type, formula_id, date_endorsed, encoded_on)
                                    VALUES({qc_ID}, '{new_lot_list[i].strip()}', '{date_started_input.text()}', '{orig_lot.strip()}',
                                    '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}',
                                    '{qcType_dropdown.currentText()}', '{formulaID_input.text()}', '{time_endorsed_input.text()}', '{datetime.now().strftime("%Y-%m-%d %H:%M")}')
                                                        """)
                                self.conn.commit()
                                QMessageBox.information(self.body_widget.setStyleSheet("border: none;"),
                                                        "Query Success", "QC Entry Added")

                        elif len(new_lot_list) > len(
                                old_lot_list):  # IF NEW LOT IS HAVE MORE LOT THAN THE CORRECTED LOT
                            while len(old_lot_list) != len(new_lot_list):
                                old_lot_list.append(old_lot_list[-1])

                            for i in range(len(new_lot_list)):
                                self.cursor.execute(f"""
                                SELECT original_lot FROM quality_control_tbl2
                                WHERE lot_number = '{old_lot_list[i]}'

                                """)

                                result = self.cursor.fetchall()

                                try:
                                    orig_lot = result[0][0]
                                except IndexError:
                                    orig_lot = old_lot_list[i]

                                self.cursor.execute(f"""
                                    INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot, status,
                                    product_code, qc_type, formula_id, date_endorsed, encoded_on)
                                    VALUES({qc_ID}, '{new_lot_list[i].strip()}', '{date_started_input.text()}', '{orig_lot.strip()}',
                                    '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}',
                                    '{qcType_dropdown.currentText()}', '{formulaID_input.text()}', '{time_endorsed_input.text()}',
                                    '{datetime.now().strftime("%Y-%m-%d %H:%M")}')
                                                        """)
                            self.conn.commit()
                            QMessageBox.information(self.body_widget.setStyleSheet("border: none;"),
                                                    "Query Success", "QC Entry Added")

                        elif len(new_lot_list) < len(
                                old_lot_list):  # IF NEW LOT IS HAVE LESS LOT THAN THE CORRECTED LOT
                            while len(new_lot_list) != len(old_lot_list):
                                new_lot_list.append(new_lot_list[-1])
                            print(new_lot_list, old_lot_list)
                            for i in range(len(new_lot_list)):
                                self.cursor.execute(f"""
                                SELECT original_lot FROM quality_control_tbl2
                                WHERE lot_number = '{old_lot_list[i]}'

                                """)

                                result = self.cursor.fetchall()
                                try:
                                    orig_lot = result[0][0]
                                except IndexError:
                                    orig_lot = old_lot_list[i]

                                self.cursor.execute(f"""
                                    INSERT INTO quality_control_tbl2(id, lot_number, evaluation_date, original_lot, status,
                                    product_code, qc_type, formula_id, date_endorsed, encoded_on)
                                    VALUES({qc_ID}, '{new_lot_list[i].strip()}', '{date_started_input.text()}', '{orig_lot.strip()}',
                                    '{result_dropdown.currentText()}', '{productCode_dropdown.currentText().strip()}',
                                    '{qcType_dropdown.currentText()}', '{formulaID_input.text()}', '{time_endorsed_input.text()}',
                                    '{datetime.now().strftime("%Y-%m-%d %H:%M")}')
                                                        """)
                            self.conn.commit()
                            QMessageBox.information(self.body_widget.setStyleSheet("border: none;"),
                                                    "Query Success", "QC Entry Added")

                        else:
                            self.conn.rollback()
                            QMessageBox.critical(self.body_widget.setStyleSheet("border: none;"),
                                                 "ERROR",
                                                 "CORRECTION AND LOT NUMBER SHOULD BE EQUAL RANGE")
                            return
                        clear_entries()

                except Exception as e:
                    print(e)
                    QMessageBox.critical(self.body_widget, "ERROR", "test")
                    self.conn.rollback()

            # Enables the User to input in Lot number correction input
            def correction_enabled():
                if qcType_dropdown.currentText() == "CORRECTION":
                    correction_input.setEnabled(True)
                    correction_input.setStyleSheet(
                        "background-color: rgb(255, 255, 0); border: 1px solid rgb(171, 173, 179);")
                else:
                    correction_input.setEnabled(False)
                    correction_input.setStyleSheet(
                        "background-color: rgb(240, 240, 240); border: 1px solid rgb(171, 173, 179);")
                    correction_input.clear()

            def correction_auto_input():

                self.cursor.execute(f"""
                SELECT product_code, customer, evaluated_by, formula_id FROM quality_control
                WHERE lot_number = '{correction_input.text()}'

                """)
                result = self.cursor.fetchone()
                if result == None:
                    print("no result")
                    QMessageBox.information(self.qc_widget, "", "NO RESULT")
                else:
                    customers_box.setCurrentText(result[1])
                    productCode_dropdown.setCurrentText(result[0])
                    evaluatedBy_dropdown.setCurrentText(result[2])
                    formulaID_input.setText(str(result[3]))

            def autofill():
                # Auto upper case the lotnumber input
                lotNumber_input.setText(lotNumber_input.text().upper())

                if qcType_dropdown.currentText() == "NEW":
                    self.cursor.execute(f"""
                        SELECT t_fid as formula_id, t_prodcode, t_customer  
                        FROM production_merge
                        WHERE t_lotnum = '{lotNumber_input.text()}' AND t_deleted = 'false'              

                                    """)
                    result = self.cursor.fetchone()
                    if result != None:
                        customers_box.setCurrentText(result[2])
                        productCode_dropdown.setCurrentText(result[1])
                        formulaID_input.setText(str(result[0]))

                    else:
                        # For lot numbers in between the original lots
                        try:
                            if '-' in lotNumber_input.text() and lotNumber_input.text()[-1] != ')':
                                lot_number = lotNumber_input.text().strip()
                                code = re.findall(r'[A-Z]+', lot_number)[0]
                                num1 = re.findall(r'(\d+)', lot_number)[0]
                                num2 = re.findall(r'(\d+)', lot_number)[1]

                                self.cursor.execute(f"""
                                                    SELECT t_lotnum, t_customer, t_prodcode, t_fid, range1, range2 FROM 
                                                            (SELECT t_lotnum, t_customer, t_prodcode, t_fid, LEFT(t_lotnum, 4)::INTEGER AS range1,
                                                               CASE 
                                                                   WHEN LENGTH(t_lotnum) = 13 THEN SUBSTRING(t_lotnum FROM 8 FOR 4)::INTEGER
                                                                   WHEN LENGTH(t_lotnum) = 11 THEN SUBSTRING(t_lotnum FROM 7 FOR 4)::INTEGER
                                                                   ELSE NULL
                                                               END AS range2
                                                        FROM public.production_merge
                                                        WHERE t_lotnum LIKE '%-%' 
                                                          AND t_lotnum ~* '\d{code}$'
                                                          AND t_deleted = 'false'

                                                        ORDER BY t_prodid::INTEGER DESC
                                                        )
                                                        WHERE {num1} >= range1 AND {num2} <= range2
                                                                                            """)
                                result = self.cursor.fetchone()
                                customers_box.setCurrentText(result[1])
                                productCode_dropdown.setCurrentText(result[2])
                                formulaID_input.setText(str(result[3]))

                            else:

                                lot_number = lotNumber_input.text().strip()
                                code = re.findall(r'[a-zA-Z]+', lot_number)[0]
                                num1 = int(lot_number[:4])

                                self.cursor.execute(f"""
                                    SELECT t_lotnum, t_customer, t_prodcode, t_fid, range1, range2 FROM 
                                            (SELECT t_lotnum, t_customer, t_prodcode, t_fid, LEFT(t_lotnum, 4)::INTEGER AS range1,
                                               CASE 
                                                   WHEN LENGTH(t_lotnum) = 13 THEN SUBSTRING(t_lotnum FROM 8 FOR 4)::INTEGER
                                                   WHEN LENGTH(t_lotnum) = 11 THEN SUBSTRING(t_lotnum FROM 7 FOR 4)::INTEGER
                                                   ELSE NULL
                                               END AS range2
                                        FROM public.production_merge
                                        WHERE t_lotnum LIKE '%-%' 
                                          AND t_lotnum LIKE '%{code}%' 
                                          AND t_deleted = false
                                        ORDER BY t_prodid::INTEGER DESC
                                        )
                                        WHERE {num1} >= range1 AND {num1} <= range2
                                    """)

                                result = self.cursor.fetchone()
                                if result:
                                    productCode_dropdown.setCurrentText(result[2])
                                    formulaID_input.setText(str(result[3]))
                                    customers_box.setCurrentText(result[1])
                                else:
                                    QMessageBox.critical(self.qc_widget, "Not Found", "LOT Number Does not Exist!")
                                    return

                        except:
                            QMessageBox.critical(self.qc_widget, "Not Found", "LOT Number Does not Exist!")
                            return

                else:
                    pass

            def clear_field():
                evaluation_entry()

            try:
                self.qc_widget.deleteLater()
            except:
                pass

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 700)
            self.qc_widget.setStyleSheet("background-color: white;")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.setStyleSheet('background-color: rgb(92, 154, 255)')
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(55, 0, 70, 30)
            self.qc_TableBtn.setText("QC DATA")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_TableBtn.setStyleSheet(
                "color: white; border: none;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(145, 0, 100, 30)
            self.qc_addEntryBtn.setText("QC ENTRY")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet(
                "color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(270, 0, 70, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet("color: white; border: none;")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(370, 0, 85, 30)
            self.dashboardBtn.setText("DASHBOARD")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("color: white; border: none")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(0, 30, 991, 721)
            self.body_widget.setStyleSheet(
                "background-color: rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(500, 0, 70, 30)
            self.qc_returns.setText("RETURNS")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet("color: white; border: none;")
            self.qc_returns.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            label1 = QtWidgets.QLabel(self.body_widget)
            label1.setGeometry(0, 0, 300, 15)
            label1.setText("QC EVALUATION AND RESULT")
            label1.setStyleSheet("border-top: none")
            label1.setFont(QtGui.QFont("Arial", 9))
            label1.show()

            label2 = QtWidgets.QLabel(self.body_widget)
            label2.setGeometry(0, 15, 250, 13)
            label2.setText("USER CAN VIEW, ADD AND UPDATE QC RECORDS")
            label2.setFont(QtGui.QFont("Arial", 8))
            label2.setStyleSheet("border-top: none")
            label2.show()

            widget1 = QtWidgets.QWidget(self.body_widget)
            widget1.setGeometry(0, 33, 440, 395)
            widget1.setStyleSheet("background-color: rgb(239, 243, 254); border: none;")

            font = QtGui.QFont("Arial", 10)

            # Create widgets to put in widget1
            qcType_label = QLabel()
            qcType_label.setText("QC TYPE")
            qcType_label.setFixedWidth(110)
            qcType_label.setFixedHeight(20)
            qcType_label.setFont(font)
            qcType_label.setStyleSheet("border: none;")

            qcControl_label = QLabel()
            qcControl_label.setText("QC Control")
            qcControl_label.setFont(font)
            qcControl_label.setStyleSheet("border: none;")

            customer_label = QLabel()
            customer_label.setText("CUSTOMER")
            customer_label.setFixedWidth(110)
            customer_label.setFixedHeight(20)
            customer_label.setFont(font)
            customer_label.setStyleSheet("border: none;")

            productCode_label = QLabel()
            productCode_label.setText("PRODUCT CODE")
            productCode_label.setFixedWidth(110)
            productCode_label.setFixedHeight(20)
            productCode_label.setFont(font)
            productCode_label.setStyleSheet("border: none;")

            evaluatedBy_label = QLabel()
            evaluatedBy_label.setText("EVALUATED BY")
            evaluatedBy_label.setFixedWidth(110)
            evaluatedBy_label.setFixedHeight(20)
            evaluatedBy_label.setFont(font)
            evaluatedBy_label.setStyleSheet("border: none;")

            date_started_label = QLabel()
            date_started_label.setText("DATE STARTED")
            date_started_label.setFixedWidth(110)
            date_started_label.setFixedHeight(20)
            date_started_label.setFont(font)
            date_started_label.setStyleSheet("border: none;")

            lotNumber_label = QLabel()
            lotNumber_label.setText("LOT NUMBER")
            lotNumber_label.setFixedWidth(110)
            lotNumber_label.setFixedHeight(20)
            lotNumber_label.setFont(font)
            lotNumber_label.setStyleSheet("border: none;")

            time_started_label = QLabel()
            time_started_label.setText("TIME STARTED")
            time_started_label.setFixedWidth(110)
            time_started_label.setFixedHeight(20)
            time_started_label.setFont(font)
            time_started_label.setStyleSheet("border: none;")

            time_endorsed_label = QLabel()
            time_endorsed_label.setText("TIME ENDORSED")
            time_endorsed_label.setFixedWidth(110)
            time_endorsed_label.setFixedHeight(20)
            time_endorsed_label.setFont(font)
            time_endorsed_label.setStyleSheet("border: none;")

            result_label = QLabel()
            result_label.setText("TEST RESULT")
            result_label.setFixedWidth(110)
            result_label.setFixedHeight(20)
            result_label.setFont(font)
            result_label.setStyleSheet("border: none;")

            # Right Side Widgets
            qcType_dropdown = QtWidgets.QComboBox()
            qcType_dropdown.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            qcType_dropdown.addItem("NEW")
            qcType_dropdown.addItem("CORRECTION")
            qcType_dropdown.setFixedHeight(25)
            qcType_dropdown.setFixedWidth(296)
            qcType_dropdown.currentTextChanged.connect(correction_enabled)

            qcControl_input = QtWidgets.QLineEdit()
            qcControl_input.setStyleSheet(
                "background-color: rgb(238, 238, 238); border: 1px solid rgb(171, 173, 179); color: rgb(0, 128, 192);")
            qcControl_input.setFixedHeight(35)
            qcControl_input.setEnabled(False)
            qcControl_input.setFont(QtGui.QFont('Arial Black', 20))
            self.cursor.execute("""
            SELECT MAX(id) FROM quality_control

            """)
            current_id = self.cursor.fetchone()

            if current_id[0] != None:
                qcControl_input.setText(str(current_id[0] + 1))
            else:
                qcControl_input.setText("0")

            qcControl_input.setFixedWidth(296)

            customers_box = QtWidgets.QComboBox()
            customers_box.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            customers_box.setFixedHeight(25)
            customers_box.setFixedWidth(296)
            # Add Customer List
            self.cursor.execute("""
            SELECT customers FROM customer
            ORDER BY customers

            """)
            result = self.cursor.fetchall()
            for customer in result:
                customers_box.addItem(customer[0])

            productCode_dropdown = QtWidgets.QComboBox()
            productCode_dropdown.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow; ")
            productCode_dropdown.setFixedHeight(25)
            productCode_dropdown.setFixedWidth(296)
            productCode_dropdown.setEditable(True)

            evaluatedBy_dropdown = QtWidgets.QComboBox()
            evaluatedBy_dropdown.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            evaluatedBy_dropdown.setFixedHeight(25)
            evaluatedBy_dropdown.setFixedWidth(296)
            evaluatedBy_dropdown.addItem("Chelsea")
            evaluatedBy_dropdown.addItem("Linzy Jam")
            evaluatedBy_dropdown.addItem("Jinky")
            evaluatedBy_dropdown.addItem("Ana")

            date_started_input = QDateTimeEdit()
            date_started_input.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            date_started_input.setFixedHeight(25)
            date_started_input.setFont(font)
            date_started_input.setFixedWidth(296)
            date_now = datetime.now()
            date_started_input.setDateTime(date_now)
            date_started_input.setDisplayFormat("MM-dd-yyyy HH:mm")

            lotNumber_input = QtWidgets.QLineEdit()
            lotNumber_input.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            lotNumber_input.setFixedHeight(25)
            lotNumber_input.setFixedWidth(296)
            lotNumber_input.editingFinished.connect(multiple_lotNumber)
            lotNumber_input.editingFinished.connect(autofill)

            time_started_input = QTimeEdit()
            time_started_input.setStyleSheet("border: 1px solid rgb(171, 173, 179);")
            time_started_input.setFixedHeight(25)
            time_started_input.setFixedWidth(120)
            time_started_input.setDisplayFormat("HH:mm")
            time_started_input.setFixedWidth(296)

            time_endorsed_input = QDateTimeEdit()
            time_endorsed_input.setStyleSheet(
                "border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 0)")
            time_endorsed_input.setFixedHeight(25)
            time_endorsed_input.setFixedWidth(296)
            time_endorsed_input.setFont(font)
            time_endorsed_input.setDateTime(date_now)
            time_endorsed_input.setDisplayFormat("MM-dd-yyyy HH:mm")

            remarks_label = QLabel(self.body_widget)
            remarks_label.setGeometry(0, 393, 100, 25)
            remarks_label.setText("   REMARKS")
            remarks_label.setFont(QtGui.QFont("Arial", 9))
            remarks_label.setStyleSheet("border: none;")
            remarks_label.show()

            remarks_box = QtWidgets.QTextEdit(self.body_widget)
            remarks_box.setGeometry(125, 393, 595, 104)
            remarks_box.setStyleSheet("background-color: rgb(255, 255, 255); border: 1px solid rgb(171, 173, 179)")
            remarks_box.show()

            result_dropdown = QComboBox()
            result_dropdown.setFixedWidth(296)
            result_dropdown.setFixedHeight(25)
            result_dropdown.setStyleSheet("background-color: rgb(255, 255, 0); border: 1px solid rgb(171, 173, 179)")
            result_dropdown.addItem("Passed")
            result_dropdown.addItem("Failed")

            correction_label = QLabel()
            correction_label.setText("CORRECTION")
            correction_label.setFixedWidth(110)
            correction_label.setFixedHeight(20)
            correction_label.setFont(font)
            correction_label.setStyleSheet("border: none;")

            correction_input = QLineEdit()
            correction_input.setStyleSheet(
                "background-color: rgb(240, 240, 240); border: 1px solid rgb(171, 173, 179);")
            correction_input.setFixedHeight(25)
            correction_input.setEnabled(False)
            correction_input.setFixedWidth(296)
            correction_input.editingFinished.connect(correction_auto_input)
            correction_input.editingFinished.connect(get_old_lotNumbers)

            formulaID_label = QLabel()
            formulaID_label.setText("FORMULA ID")
            formulaID_label.setFixedWidth(110)
            formulaID_label.setFixedHeight(20)
            formulaID_label.setFont(font)
            formulaID_label.setStyleSheet("border: none;")

            formulaID_input = QLineEdit()
            formulaID_input.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: yellow;")
            formulaID_input.setFixedHeight(25)
            formulaID_input.setFixedWidth(296)

            actionTaken_label = QLabel(self.body_widget)
            actionTaken_label.setGeometry(0, 500, 100, 25)
            actionTaken_label.setText("   ACTION TAKEN")
            actionTaken_label.setFont(QtGui.QFont("Arial", 9))
            actionTaken_label.setStyleSheet("border: none;")
            actionTaken_label.show()

            actionTake_box = QTextEdit(self.body_widget)
            actionTake_box.setGeometry(125, 500, 595, 104)
            actionTake_box.setStyleSheet("background-color: rgb(255, 255, 255); border: 1px solid rgb(171, 173, 179)")
            actionTake_box.show()

            lot_number_board_label = QLabel(self.body_widget)
            lot_number_board_label.setGeometry(470, 170, 100, 30)
            lot_number_board_label.setText('LOT NUMBERS')
            lot_number_board_label.setFont(QtGui.QFont('Arial', 10))
            lot_number_board_label.setStyleSheet('border: none; ')
            lot_number_board_label.show()

            lotNumbers_board = QtWidgets.QTextEdit(self.body_widget)
            lotNumbers_board.setGeometry(470, 200, 250, 180)
            lotNumbers_board.setStyleSheet("border: 1px solid black; background-color: rgb(255, 255, 0)")
            lotNumbers_board.setFont(font)
            lotNumbers_board.show()

            label3 = QLabel(self.body_widget)
            label3.setText(" WORKSTATION AND USERNAME")
            label3.setGeometry(0, 620, 194, 25)
            label3.setFont(QtGui.QFont("Arial", 9))
            label3.setStyleSheet("border: none;")
            label3.show()

            user_input = QLineEdit(self.body_widget)
            user_input.setGeometry(194, 620, 221, 25)
            user_input.setEnabled(False)
            user_input.setText(os.environ['COMPUTERNAME'])
            user_input.setStyleSheet("border: 1px solid rgb(177, 206, 237); background-color: rgb(192, 192, 192); ")
            user_input.show()

            save_btn = QPushButton(self.body_widget)
            save_btn.setGeometry(534, 618, 60, 25)
            save_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                   "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            save_btn.setText("SAVE")
            save_btn.clicked.connect(saveBtn_clicked)
            save_btn.setShortcut("Return")
            save_btn.show()

            clear_btn = QPushButton(self.body_widget)
            clear_btn.setGeometry(596, 618, 60, 25)
            clear_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                    "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            clear_btn.setText("CLEAR")
            clear_btn.clicked.connect(clear_field)
            clear_btn.show()

            close_btn = QPushButton(self.body_widget)
            close_btn.setGeometry(658, 618, 60, 25)
            close_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                    "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            close_btn.setText("CLOSE")
            close_btn.show()

            topFormLayout = QFormLayout(widget1)
            topFormLayout.addRow(qcType_label, qcType_dropdown)
            topFormLayout.addRow(qcControl_label, qcControl_input)
            topFormLayout.addRow(correction_label, correction_input)
            topFormLayout.addRow(customer_label, customers_box)
            topFormLayout.addRow(lotNumber_label, lotNumber_input)
            topFormLayout.addRow(productCode_label, productCode_dropdown)
            topFormLayout.addRow(formulaID_label, formulaID_input)
            topFormLayout.addRow(evaluatedBy_label, evaluatedBy_dropdown)
            topFormLayout.addRow(date_started_label, date_started_input)
            topFormLayout.addRow(time_endorsed_label, time_endorsed_input)
            topFormLayout.addRow(result_label, result_dropdown)

            widget1.show()

        def show_qc_data():
            self.qc_widget.deleteLater()

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 751)
            self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.setStyleSheet('background-color: rgb(92, 154, 255)')
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(55, 0, 70, 30)
            self.qc_TableBtn.setText("QC DATA")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_TableBtn.setStyleSheet(
                "color: white; border: none;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(145, 0, 100, 30)
            self.qc_addEntryBtn.setText("QC ENTRY")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet(
                "color: white; border: none;")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(270, 0, 70, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet(
                "color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(370, 0, 85, 30)
            self.dashboardBtn.setText("DASHBOARD")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("color: white; border: none")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(500, 0, 70, 30)
            self.qc_returns.setText("RETURNS")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet("color: white; border: none;")
            self.qc_returns.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(0, 30, 991, 721)
            self.body_widget.setStyleSheet(
                "background-color: rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            qc_data_table = QTableWidget(self.body_widget)
            qc_data_table.setGeometry(50, 20, 890, 590)
            qc_data_table.setStyleSheet("border: 1px solid black; ")
            qc_data_table.horizontalHeader().setStyleSheet("""
                    QHeaderView::section{
                    font-weight: bold;
                    background-color: rgb(0, 109, 189);
                    color: white;
                    }

                    """)

            ph_holiday = hd.country_holidays('PH')
            self.cursor.execute("""
            SELECT original_lot, MIN(date_endorsed)::DATE as min_date, MAX(evaluation_date)::DATE as max_date
            FROM qc_num_days  
            GROUP BY original_lot

            """)
            result = self.cursor.fetchall()

            dayoff = []
            original_lot = []
            for entry in result:

                prod_code = entry[0]
                min_date = entry[1]
                max_date = entry[2]

                date_range = pd.date_range(start=f'{min_date}', end=f'{max_date}').strftime('%m-%d-%Y')

                holidays = []
                sundays = []
                for i in list(date_range):
                    if i in ph_holiday:
                        holidays.append(i)
                    if datetime.strptime(i, '%m-%d-%Y').weekday() == 6:
                        sundays.append(i)
                no_operation = holidays + sundays
                no_operation = len(set(no_operation))
                dayoff.append(no_operation)
                original_lot.append(prod_code)

            data = [(x, y) for x, y in zip(original_lot, dayoff)]

            # DELETE THE TABLE TO CLEAR THE CONTENT FOR UPDATE
            self.cursor.execute("""
            DELETE FROM qc_dayoff

            """)
            self.conn.commit()

            insert_query = sql.SQL("""
                        INSERT INTO qc_dayoff
                        VALUES(%s, %s)
            """)

            self.cursor.executemany(insert_query, data)
            self.conn.commit()

            # get the data from the Database
            self.cursor.execute("""
                SELECT id, lot_number, evaluated_on, time_endorsed, status, product_code, evaluated_on - time_endorsed as qc_days
                FROM quality_control
                ORDER BY id DESC

                """)

            result = self.cursor.fetchall()

            # Set Row Count
            qc_data_table.setRowCount(len(result))
            qc_data_table.setColumnCount(7)

            for i in range(len(result)):
                if result[i][4] == "Failed":
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setBackground(QtGui.QColor(252, 3, 28))
                        item.setForeground(QtGui.QColor(0, 0, 0))
                        qc_data_table.setItem(i, j, item)
                else:
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        qc_data_table.setItem(i, j, item)

            qc_data_table.setColumnWidth(1, 150)
            qc_data_table.setColumnWidth(2, 150)
            qc_data_table.setColumnWidth(3, 150)
            qc_data_table.setColumnWidth(6, 120)
            qc_data_table.verticalHeader().setVisible(False)
            qc_data_table.setHorizontalHeaderLabels(
                ["QC ID", "LOT NUMBER", "EVALUATION DATE", "TIME ENDORSED", "STATUS",
                 "PRODUCT CODE", "QC DAYS"])
            qc_data_table.show()

            # self.cursor.execute("""
            # WITH numbered_row AS (SELECT * , ROW_NUMBER() OVER (PARTITION BY original_lot order by evaluation_date) AS rn
            # FROM quality_control_tbl2)
            # SELECT numbered_row.original_lot, numbered_row.status, product_code, numbered_row.rn
            # FROM numbered_row
            #
            # """)
            #
            # result = self.cursor.fetchall()
            # df = pd.DataFrame(result)
            # try:
            #     df.columns = ["original_lot", "status", "product_code", "row_number"]
            # except:
            #     QMessageBox.information(self.qc_widget, "No Data", "No Data Found")
            #     return
            # passToFail_counter = {}
            # firstTry_failed = {}
            #
            # # Getting the number of Lot Number with Passed Status and then become Failed Later
            # for index, row in df.iterrows():
            #     original_lot = row['original_lot']
            #     status = row['status']
            #     product_code = row['product_code']
            #     row_number = row['row_number']
            #     if status == "Passed" and row_number == 1:
            #         try:
            #             if df.iat[index + 1, 1] == 'Failed' and df.iat[index + 1, 0] == original_lot:
            #
            #                 if product_code not in passToFail_counter.keys():
            #                     passToFail_counter[product_code] = 1
            #                 else:
            #                     passToFail_counter[product_code] += 1
            #         except Exception as e:
            #             print(e)
            #
            #     # For getting the Failed Lot on the First Qc
            #     elif status == "Failed" and row_number == 1:
            #         if product_code not in firstTry_failed.keys():
            #             firstTry_failed[product_code] = 1
            #         else:
            #             firstTry_failed[product_code] += 1
            #
            # passToFail_x = []
            # passToFail_y = []
            #
            # # Query For Getting the total Amount of original_lot per Product Code
            # self.cursor.execute("""
            # SELECT product_code, COUNT(*) AS total_quantity
            # FROM (SELECT DISTINCT ON (product_code, original_lot) *
            #       FROM quality_control_tbl2
            #       ORDER BY product_code, original_lot, evaluation_date ) AS distinct_lots
            # GROUP BY product_code
            # ORDER BY product_code;
            #
            # """)
            # result = self.cursor.fetchall()
            # total_productcode = {}
            # for i in result:
            #     total_productcode[i[0]] = i[1]
            #
            # # Getting the percentage of Pass to Fail of Product Codes
            # passToFail_percentage = {}
            # for key in passToFail_counter.keys():
            #     passToFail_percentage[key] = passToFail_counter[key] / total_productcode[key]
            #
            # # Get the DISTINCT OF PRODUCT CODE
            # self.cursor.execute("""
            # SELECT DISTINCT product_code
            # FROM quality_control_tbl2
            #
            # """)
            # result = self.cursor.fetchall()
            # prod_code_list = [i[0] for i in result] # Parse the data
            #
            # # Add the other product code and Set the Other Value to 0
            # for i in prod_code_list:
            #     if i not in passToFail_percentage.keys():
            #         passToFail_percentage[i] = 0
            #
            # # unpacking the dictionary to list
            # for key, value in passToFail_percentage.items():
            #     passToFail_x.append(key)
            #     passToFail_y.append(value)
            #
            # # Table For Showing Average QC days per Product Code
            # aggregated_products_table = QTableWidget(self.body_widget)
            # aggregated_products_table.setGeometry(50, 390, 300, 300)
            # aggregated_products_table.setColumnCount(3)
            # aggregated_products_table.setRowCount(10)
            # aggregated_products_table.verticalHeader().setVisible(False)
            # aggregated_products_table.setHorizontalHeaderLabels(["Product Code", "Average QC days", "Pass to Fail"])
            # # aggregated_products_table.show()

        def show_dashboards():

            def get_data():
                # This is for the Other Graphs that cant make it in the First Six

                matplotlib.use('Qt5Agg')

                def change_graph():

                    if graph5_selections.currentText() == 'Supervisor':
                        self.cursor.execute(f"""
                            WITH splitted_lot AS (
                        SELECT  *,
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_max,
                        (regexp_match(lot_number[1], '[A-Z]+'))[1] as first_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_max,
                        (regexp_match(lot_number[2], '[A-Z]+'))[1] as second_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_max,
                        (regexp_match(lot_number[3], '[A-Z]+'))[1] as third_lot_code
                        FROM extruder
                    )

                SELECT t2.supervisor, COUNT(*) FROM returns t1
                JOIN splitted_lot t2
                    ON (((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.first_lot_min AND t2.first_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.second_lot_min AND t2.second_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = second_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.third_lot_min AND t2.third_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = third_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER IN (t2.first_lot_min, t2.second_lot_min, t2.third_lot_min)
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code)
                       )
                WHERE return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                GROUP BY supervisor
                                             """)
                        result = self.cursor.fetchall()
                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)
                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0, 110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest AVG Returns By Supervisor", fontsize=15)
                        self.canvas.draw()

                    elif graph5_selections.currentText() == 'Operator':

                        self.cursor.execute(f"""
                                            SELECT operator, COUNT(operator)
                                            FROM
                                            (SELECT t1.lot_number, t2.operator, t2.supervisor
                                            FROM returns t1
                                            JOIN extruder t2 ON t1.origin_lot = ANY(t2.lot_number)
                                            WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                                            GROUP BY operator

                                        """)
                        result = self.cursor.fetchall()
                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0, 110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest AVG Returns By Operator", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                        self.canvas.draw()

                    elif graph5_selections.currentText() == 'QC Analyst':

                        self.cursor.execute(f"""
                        SELECT evaluated_by, COUNT(evaluated_by)
                        FROM (SELECT t1.*, t2.evaluated_by
                        FROM returns t1
                        JOIN quality_control t2 ON t1.origin_lot = t2.lot_number
                        WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                        GROUP BY evaluated_by

                        """)
                        result = self.cursor.fetchall()

                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0, 110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest Returns By QC Analyst", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                        self.canvas.draw()

                    elif graph5_selections.currentText() == 'Extruder':

                        self.cursor.execute(f"""
                        SELECT machine, COUNT(machine)
                        FROM
                        (SELECT t1.lot_number, t2.machine
                        FROM returns t1
                        JOIN extruder t2 ON t1.origin_lot = ANY(t2.lot_number)
                        WHERE t1.return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}')
                        GROUP BY machine
                        """)

                        result = self.cursor.fetchall()

                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_yticks(np.arange(0, 110, 10))
                        self.graph5.set_ylim(0, 110)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Highest Returns By Extruder", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                        self.canvas.draw()

                    elif graph5_selections.currentText() == 'Total Kg':

                        self.cursor.execute(f"""
                        SELECT product_code, SUM(quantity) as total_kg
                        FROM returns
                        WHERE return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                        GROUP BY product_code

                        ORDER BY total_kg
                        """)

                        result = self.cursor.fetchall()

                        x = []
                        y = []
                        for i, j in result:
                            x.append(i)
                            y.append(j)

                        self.graph5.clear()
                        self.graph5.bar(x, y)
                        self.graph5.set_xticklabels(x, rotation=30)
                        self.graph5.set_title("Total Returns(kg) By Product Code", fontsize=15)
                        self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                        self.canvas.draw()

                date1 = dateFrom_widget.currentIndex() + 1
                date2 = dateTo_widget.currentIndex() + 1

                # Get the last day of Month
                ph_holiday = hd.country_holidays('PH')
                # Query For Getting the MIN AND MAX Date of original lot
                self.cursor.execute(f"""
                            SELECT original_lot, MIN(evaluation_date)::DATE as min_date, MAX(evaluation_date)::DATE as max_date
                            FROM qc_num_days  
                            WHERE evaluation_date::DATE BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                            GROUP BY original_lot

                            """)
                result = self.cursor.fetchall()

                dayoff = []
                original_lot = []

                for entry in result:
                    prod_code = entry[0]
                    min_date = entry[1]
                    max_date = entry[2]

                    date_range = pd.date_range(start=f'{min_date}', end=f'{max_date}').strftime('%m-%d-%Y')

                    holidays = []
                    sundays = []
                    for i in list(date_range):
                        if i in ph_holiday:
                            holidays.append(i)
                        if datetime.strptime(i, '%m-%d-%Y').weekday() == 6:
                            sundays.append(i)

                    no_operation = holidays + sundays
                    no_operation = len(set(no_operation))
                    dayoff.append(no_operation)
                    original_lot.append(prod_code)

                data = [(x, y) for x, y in zip(original_lot, dayoff)]

                # DELETE THE TABLE TO CLEAR THE CONTENT FOR UPDATE
                self.cursor.execute("""DELETE FROM qc_dayoff
                     """)
                self.conn.commit()
                insert_query = sql.SQL("""
                INSERT INTO qc_dayoff
                VALUES(%s, %s)
                """)

                self.cursor.executemany(insert_query, data)
                self.conn.commit()

                # Query For Getting the AVERAGE qc_days PER PRODUCT_CODE
                self.cursor.execute(f"""
                 SELECT product_code, ROUND(AVG(dayoff), 2) as avg_dayoff
                    FROM(SELECT product_code, qc_day - dayoff AS dayoff
                    FROM (SELECT t1.product_code, t1.original_lot, dayoff ,(MAX(evaluation_date::DATE) - MIN(date_endorsed::DATE)) + 1   as qc_day 
                    FROM quality_control_tbl2 t1
                    JOIN qc_dayoff t2 ON t1.original_lot = t2.original_lot
                    WHERE evaluation_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                    GROUP BY product_code, t1.original_lot, dayoff))
                    GROUP BY product_code
                    ORDER BY avg_dayoff DESC
                    LIMIT 20
                """)

                layout = QHBoxLayout(self.body_widget)

                self.figure = plt.figure(figsize=(2, 2), dpi=60)
                self.canvas = FigureCanvas(self.figure)
                layout.addWidget(self.canvas)
                self.figure.patch.set_facecolor("#eff3fe")

                font = QtGui.QFont("Arial", 14)

                dateRangeLabel = QLabel(self.body_widget)
                dateRangeLabel.setGeometry(450, 25, 320, 30)
                dateRangeLabel.setAlignment(Qt.AlignCenter)
                if dateTo_widget.currentText() == dateFrom_widget.currentText():
                    dateRangeLabel.setText(dateFrom_widget.currentText())
                else:
                    dateRangeLabel.setText(dateFrom_widget.currentText() + " - " + dateTo_widget.currentText())
                dateRangeLabel.setFont(font)
                dateRangeLabel.setStyleSheet("color: rgb(0, 109, 189); border: none;")
                dateRangeLabel.show()

                self.graph1 = self.figure.add_subplot(231)
                self.graph2 = self.figure.add_subplot(232)
                self.graph3 = self.figure.add_subplot(233)
                self.graph4 = self.figure.add_subplot(234)
                self.graph5 = self.figure.add_subplot(235)
                self.graph6 = self.figure.add_subplot(236)

                result = self.cursor.fetchall()
                x = []
                y = []
                # Unpack the List of tuple
                for item, value in result:
                    x.append(item)
                    y.append(value)

                # slice to only top 5
                x = x[:5]
                y = y[:5]
                self.graph1.bar(x, y)
                self.graph1.set_xticklabels(x, rotation=30)
                self.graph1.set_ylim(0, 35)
                self.graph1.set_yticks(np.arange(0, 35, 5))

                self.graph1.set_title("Highest AVG QC days", fontsize=15)

                # Get the Total Product Code runs from data x to date y
                self.cursor.execute(f"""
                SELECT product_code, COUNT(*) AS total_quantity
            FROM (SELECT DISTINCT ON (product_code, original_lot) *
            FROM quality_control_tbl2
		    WHERE evaluation_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
            ORDER BY product_code, original_lot, evaluation_date ) AS distinct_lots
            GROUP BY product_code
            ORDER BY COUNT(*) DESC

                """)
                result = self.cursor.fetchall()

                total_productCodes = {}
                for i in result:
                    total_productCodes[i[0].strip()] = i[1]

                self.cursor.execute(f"""
                SELECT product_code, COUNT(*)
                FROM quality_control
                WHERE status_changed = true AND evaluated_on BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                GROUP BY product_code
                ORDER BY COUNT(*) DESC
                LIMIT 5

                """)

                result = self.cursor.fetchall()

                prodouctCodeList = []
                prodouctCodeCount = []
                for items in result:
                    prodouctCodeList.append(items[0])
                    prodouctCodeCount.append(items[1])

                self.graph2.bar(prodouctCodeList, prodouctCodeCount)
                try:
                    self.graph2.set_yticks(np.arange(0, prodouctCodeCount[0] + 20, 5))
                    self.graph2.set_ylim(0, prodouctCodeCount[0] + 20)
                except IndexError:
                    self.graph2.set_yticks(np.arange(0, 50, 5))
                    self.graph2.set_ylim(0, 50)

                self.graph2.set_xticklabels(prodouctCodeList, rotation=30)
                self.graph2.set_title("Most Change", fontsize=15)

                # Visual 3

                # Query For Getting the Failed First Run
                self.cursor.execute(f"""
                    SELECT product_code , COUNT(*) FROM public.quality_control_tbl2
                    WHERE qc_type = 'NEW' AND status = 'Failed' AND evaluation_date BETWEEN '2024-{date1}-01' AND 
                    '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                    GROUP BY product_code
                    ORDER BY count
                    LIMIT 5

                """)

                result = self.cursor.fetchall()
                x = []
                y = []

                for item in result:
                    x.append(item[0])
                    y.append(item[1])

                self.graph3.bar(x, y)
                self.graph3.set_xticklabels(x, rotation=30)
                self.graph3.set_yticks(np.arange(0, 110, 10))
                self.graph3.set_ylim(0, 110)
                self.graph3.set_title("Failed First Run", fontsize=15)

                # Visual 4

                self.cursor.execute(f"""
                SELECT product_code, COUNT(*) 
                FROM returns
				WHERE return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                GROUP BY product_code			
                ORDER BY count DESC	
                LIMIT 5
                """)

                result = self.cursor.fetchall()
                productCodeReturns = {}

                for i, j in result:
                    productCodeReturns[i] = j

                # Getting the Percentage of Product Code returns for each Product Codes

                for key in productCodeReturns.keys():
                    print(key, total_productCodes)
                    productCodeReturns[key] = (productCodeReturns[key] / total_productCodes[key]) * 100

                x = []
                y = []
                # Getting the x and y from the Returns dictionary
                for key, value in productCodeReturns.items():
                    x.append(key)
                    y.append(value)

                # Plot the Visual 4

                self.graph4.bar(x, y)
                self.graph4.set_yticks(np.arange(0, 110, 10))
                self.graph4.set_ylim(0, 110)
                self.graph4.set_xticklabels(x, rotation=30)
                self.graph4.set_title("Highest Return Percentage", fontsize=15)
                self.graph4.set_position([0.125, 0.08, 0.228, 0.35])

                # Getting the data For Visualization 5

                # Getting the Data for how many times an operator have a Returned product
                self.cursor.execute(f"""
                     WITH splitted_lot AS (
                        SELECT  *,
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_max,
                        (regexp_match(lot_number[1], '[A-Z]+'))[1] as first_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_max,
                        (regexp_match(lot_number[2], '[A-Z]+'))[1] as second_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_max,
                        (regexp_match(lot_number[3], '[A-Z]+'))[1] as third_lot_code
                        FROM extruder
                    )

                SELECT t2.operator, COUNT(*) FROM returns t1
                JOIN splitted_lot t2
                    ON (((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.first_lot_min AND t2.first_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.second_lot_min AND t2.second_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = second_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.third_lot_min AND t2.third_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = third_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER IN (t2.first_lot_min, t2.second_lot_min, t2.third_lot_min)
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code)
                       )
                WHERE return_date BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                GROUP BY operator

                """)
                result = self.cursor.fetchall()
                x = []
                y = []
                for i, j in result:
                    x.append(i)
                    y.append(j)

                # Set Page for Graph 5
                graph5_page = 1

                self.graph5.bar(x, y)
                self.graph5.set_yticks(np.arange(0, 110, 10))
                self.graph5.set_ylim(0, 110)
                self.graph5.set_xticklabels(x, rotation=30)
                self.graph5.set_title("Highest AVG Returns By Operator", fontsize=15)
                self.graph5.set_position([0.4, 0.08, 0.228, 0.35])

                graph5_selections = QComboBox(self.body_widget)
                graph5_selections.setGeometry(650, 632, 100, 20)

                graph5_selections.addItem('Operator')
                graph5_selections.addItem('Supervisor')
                graph5_selections.addItem('Total Kg')
                graph5_selections.addItem("QC Analyst")
                graph5_selections.addItem('Extruder')
                graph5_selections.currentIndexChanged.connect(change_graph)
                graph5_selections.show()

                self.cursor.execute(f"""
                    SELECT t_matcode, SUM(t_wt) as total_weight FROM tbl_prod02
                    WHERE t_deleted = false AND t_proddate BETWEEN '2024-{date1}-01' AND '2024-{date2}-{calendar.monthrange(2024, date2)[1]}'
                    GROUP BY t_matcode
                    ORDER BY total_weight DESC
                    LIMIT 5;

                    """)

                result = self.cursor.fetchall()
                x = []
                y = []

                for items in result:
                    x.append(items[0])
                    y.append(items[1] / 1000)

                self.graph6.bar(x, y)
                self.graph6.set_ylabel('Tonnes', fontsize=15)
                self.graph6.set_xticklabels(x, rotation=30)
                self.graph6.set_title("Highest Materials Used", fontsize=15)
                self.graph6.set_position([0.672, 0.078, 0.228, 0.35])

                import_button = QPushButton(self.body_widget)
                import_button.setGeometry(850, 35, 100, 25)
                import_button.setText("IMPORT")
                import_button.setStyleSheet("border: none; background-color: rgb(227, 227, 227)")
                import_button.setCursor(Qt.PointingHandCursor)
                import_button.clicked.connect(lambda: save_to_excel(date1, date2))
                import_button.show()

            try:
                self.qc_widget.deleteLater()
            except:
                pass

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 700)
            self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.setStyleSheet('background-color: rgb(92, 154, 255)')
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(55, 0, 70, 30)
            self.qc_TableBtn.setText("QC DATA")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_TableBtn.setStyleSheet(
                "color: white; border: none;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(145, 0, 100, 30)
            self.qc_addEntryBtn.setText("QC ENTRY")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet(
                "color: white; border: none; padding-bottom: 5px;")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(270, 0, 70, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet("color: white; border: none;")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(370, 0, 85, 30)
            self.dashboardBtn.setText("DASHBOARD")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("color: white; border: none; border-bottom: 2px solid white;")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(500, 0, 70, 30)
            self.qc_returns.setText("RETURNS")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet("color: white; border: none;")
            self.qc_returns.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(-120, 30, 1200, 671)
            self.body_widget.setStyleSheet(
                "background-color: rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            dateFrom_widget = QtWidgets.QComboBox(self.body_widget)
            dateFrom_widget.setGeometry(200, 30, 100, 20)
            dateFrom_widget.setStyleSheet("background-color: rgb(137, 137, 161)")
            dateFrom_widget.addItem("JANUARY")
            dateFrom_widget.addItem("FEBRUARY")
            dateFrom_widget.addItem("MARCH")
            dateFrom_widget.addItem("APRIL")
            dateFrom_widget.addItem("MAY")
            dateFrom_widget.addItem("JUNE")
            dateFrom_widget.addItem("JULY")
            dateFrom_widget.addItem("AUGUST")
            dateFrom_widget.addItem("SEPTEMBER")
            dateFrom_widget.addItem("OCTOBER")
            dateFrom_widget.addItem("NOVEMBER")
            dateFrom_widget.addItem("DECEMBER")
            dateFrom_widget.show()

            dateTo_widget = QtWidgets.QComboBox(self.body_widget)
            dateTo_widget.setGeometry(350, 30, 100, 20)
            dateTo_widget.setStyleSheet("background-color: rgb(137, 137, 161)")
            dateTo_widget.addItem("JANUARY")
            dateTo_widget.addItem("FEBRUARY")
            dateTo_widget.addItem("MARCH")
            dateTo_widget.addItem("APRIL")
            dateTo_widget.addItem("MAY")
            dateTo_widget.addItem("JUNE")
            dateTo_widget.addItem("JULY")
            dateTo_widget.addItem("AUGUST")
            dateTo_widget.addItem("SEPTEMBER")
            dateTo_widget.addItem("OCTOBER")
            dateTo_widget.addItem("NOVEMBER")
            dateTo_widget.addItem("DECEMBER")
            dateTo_widget.activated.connect(get_data)
            dateTo_widget.show()

        def save_to_excel(date1, date2):
            show_qc_data()

            # Query for getting the Average QC days For Each Product Code
            self.cursor.execute(f"""
                SELECT product_code, ROUND(AVG(qc_days), 2) as avg_qc_days
                    FROM(SELECT product_code, qc_day - dayoff AS qc_days
                    FROM (SELECT t1.product_code, t1.original_lot, dayoff ,(MAX(encoded_on::DATE) - MIN(date_endorsed::DATE)) + 1   as qc_day 
                    FROM quality_control_tbl2 t1
                    JOIN qc_dayoff t2 ON t1.original_lot = t2.original_lot
					WHERE evaluation_date BETWEEN '{date1}' AND '{date2}'
                    GROUP BY product_code, t1.original_lot, dayoff))
                    GROUP BY product_code
                    ORDER BY avg_qc_days DESC
                    LIMIT 20

                            """)
            result = self.cursor.fetchall()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "1)	How long is QC (days)"
            center_Alignment = Alignment(horizontal='center', vertical='center')
            ws1.column_dimensions['B'].width = 20
            ws1.column_dimensions['A'].width = 20

            title_color = Font(color='FF0000')

            ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

            ws1['A2'].alignment = center_Alignment
            ws1['B2'].alignment = center_Alignment
            ws1['A2'] = "Product Code"
            ws1['B2'] = "Average QC Days"

            ws1['A1'] = "Average QC Days Per Product Code"
            ws1['A1'].alignment = center_Alignment
            ws1['A1'].font = title_color

            for i in range(len(result)):
                ws1[f"A{3 + i}"] = result[i][0]
                ws1[f"B{3 + i}"] = result[i][1]
                ws1[f"A{3 + i}"].alignment = center_Alignment
                ws1[f"B{3 + i}"].alignment = center_Alignment

            # Query For The Highest QC day For Each UNIQUE product Code with lot_number and FN
            self.cursor.execute(f"""
                                WITH LotQcDays AS (
                    SELECT t1.original_lot, product_code, formula_id, qc_days - dayoff AS qc_days
                FROM (SELECT
                        original_lot,
                        product_code,
                        formula_id,
                        (MAX(encoded_on::DATE) - MIN(date_endorsed::DATE)) + 1 AS qc_days  -- Adjust +1 for inclusive date range
                    FROM
                        quality_control_tbl2 
                    WHERE evaluation_date BETWEEN '{date1}' AND '{date2}'
                    GROUP BY
                        original_lot, product_code, formula_id

                     ) t1
                JOIN qc_dayoff t2 ON t2.original_lot = t1.original_lot
                ),
                MaxQcDays AS (
                    SELECT
                        product_code,
                        MAX(qc_days) AS max_qc_days
                    FROM
                        LotQcDays
                    GROUP BY
                        product_code
                ),
                RankedQcDays AS (
                    SELECT
                        l.product_code,
                        t.lot_number,
                        l.qc_days,
                        l.formula_id,
                        ROW_NUMBER() OVER (PARTITION BY l.product_code ORDER BY l.qc_days DESC) AS rn
                    FROM
                        LotQcDays l
                    JOIN
                        MaxQcDays m
                    ON
                        l.product_code = m.product_code
                        AND l.qc_days = m.max_qc_days
                    JOIN
                        quality_control_tbl2 t
                    ON
                        l.original_lot = t.original_lot
                )
                SELECT
                    product_code,
                    lot_number,
                    qc_days,
                    formula_id
                FROM
                    RankedQcDays
                WHERE
                    rn = 1
                ORDER BY qc_days DESC
                LIMIT 20

                                """)
            result = self.cursor.fetchall()

            ws1["D2"] = "Product Code"
            ws1['E2'] = "Lot Number"
            ws1['F2'] = "QC Days"
            ws1['G2'] = "Formula ID"

            cell_pointer = 3
            for item in result:
                ws1[f"D{cell_pointer}"] = item[0]
                ws1[f"E{cell_pointer}"] = item[1]
                ws1[f"F{cell_pointer}"] = item[2]
                ws1[f"G{cell_pointer}"] = item[3]
                ws1[f"D{cell_pointer}"].alignment = center_Alignment
                ws1[f"E{cell_pointer}"].alignment = center_Alignment
                ws1[f"F{cell_pointer}"].alignment = center_Alignment
                ws1[f"G{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws1.column_dimensions['D'].width = 15
            ws1.column_dimensions['E'].width = 15
            ws1.column_dimensions['F'].width = 15
            ws1.column_dimensions['G'].width = 15

            ws1.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)
            ws1['D1'] = "Product Code Highest QC Days"
            ws1['D1'].alignment = center_Alignment
            ws1['D1'].font = title_color

            # Getting the Average QC days of all lot number, Sunday and Holidays are excluded.
            self.cursor.execute(f"""
                SELECT ROUND(avg(qc_day), 2)
                FROM (SELECT t1.original_lot, MAX(encoded_on::DATE) - MIN(date_endorsed::DATE) + 1 as qc_day
                FROM quality_control_tbl2 t1
                JOIN qc_dayoff t2 ON t1.original_lot = t2.original_lot
                WHERE evaluation_date BETWEEN '{date1}' AND '{date2}'
                GROUP BY t1.original_lot)

            """)
            result = self.cursor.fetchall()

            ws1["I2"] = "AVG QC Days"

            cell_pointer = 3
            for item in result:
                ws1[f"I{cell_pointer}"] = item[0]
                ws1[f"I{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws1.column_dimensions['I'].width = 15

            # Create a new Worksheet
            ws2 = wb.create_sheet("2) QC evaluation changes")

            # Query For Getting the Count of Every Decision Changed.
            self.cursor.execute(f"""
            WITH lot_range as (
                SELECT * ,  (regexp_matches(SPLIT_PART(lot_number, '-', 1), '(\d+)[A-Z]', 'g'))[1]::INTEGER as first_lot, 
                (regexp_matches(SPLIT_PART(lot_number, '-', 2), '(\d+)[A-Z]', 'g'))[1]::INTEGER as last_lot
                FROM quality_control
				WHERE evaluated_on BETWEEN '{date1}' AND '{date2}'
                ),

                decision_changed AS (SELECT product_code, 
                        CASE 
                            WHEN last_lot IS NULL THEN 1
                            ELSE (last_lot - first_lot) + 1
                            END AS lot_count
                        FROM lot_range	
                                WHERE status_changed = true)

                SELECT product_code, SUM(lot_count) as change
                FROM decision_changed
                GROUP BY product_code
                ORDER BY change DESC
                LIMIT 20
            """)

            status_changed = self.cursor.fetchall()

            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            ws2['A1'] = "Top 20 Product Code With The Highest Evaluation Change"
            ws2['A1'].font = title_color

            ws2['A2'] = "Product Code"
            ws2['B2'] = "Decision Changed"
            ws2['A2'].alignment = center_Alignment
            ws2['B2'].alignment = center_Alignment

            ws2.column_dimensions['A'].width = 17
            ws2.column_dimensions['B'].width = 17
            ws2.column_dimensions['C'].width = 15

            cell_pointer = 3
            for items in status_changed:
                ws2[f"A{cell_pointer}"] = items[0]
                ws2[f"B{cell_pointer}"] = items[1]
                ws2[f"A{cell_pointer}"].alignment = center_Alignment
                ws2[f"B{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws4 = wb.create_sheet("3) CUSTOMER REJECT")

            # Query for getting the total customer return by product code
            self.cursor.execute(f"""
                        SELECT product_code, COUNT(*) 
                            FROM returns
            				WHERE return_date BETWEEN '{date1}' AND '{date2}'
                            GROUP BY product_code				
                        """)

            result = self.cursor.fetchall()

            ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

            ws4['A1'] = "Highest Return Per Code"
            ws4['A1'].font = title_color
            ws4['A1'].alignment = center_Alignment

            ws4["A2"] = "Product Code"
            ws4['B2'] = "RETURNS"
            ws4['A2'].alignment = center_Alignment
            ws4['B2'].alignment = center_Alignment

            ws4.column_dimensions['B'].width = 15
            ws4.column_dimensions['A'].width = 15

            cell_pointer = 3
            for item in result:
                ws4[f"A{cell_pointer}"] = item[0]
                ws4[f"B{cell_pointer}"] = item[1]
                ws4[f"A{cell_pointer}"].alignment = center_Alignment
                ws4[f"B{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Returns By Formula ID
            self.cursor.execute(f"""
                        SELECT product_code, formula_id, COUNT(*)
                        FROM returns
                        WHERE return_date BETWEEN '{date1}' AND '{date2}'
                        GROUP BY product_code, formula_id
                        ORDER BY product_code

                        """)
            result = self.cursor.fetchall()

            ws4.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
            ws4['D1'] = "Highest Return Per (Code, FN)"
            ws4['D1'].alignment = center_Alignment
            ws4['D1'].font = title_color

            ws4["D2"] = "Product Code"
            ws4['E2'] = "Formula ID"
            ws4['F2'] = "RETURNS"
            ws4['D2'].alignment = center_Alignment
            ws4['E2'].alignment = center_Alignment
            ws4['F2'].alignment = center_Alignment

            ws4.column_dimensions['D'].width = 13
            ws4.column_dimensions['E'].width = 15
            ws4.column_dimensions['F'].width = 10

            cell_pointer = 3
            for item in result:
                ws4[f"D{cell_pointer}"] = item[0]
                ws4[f"E{cell_pointer}"] = item[1]
                ws4[f"F{cell_pointer}"] = item[2]
                ws4[f"D{cell_pointer}"].alignment = center_Alignment
                ws4[f"E{cell_pointer}"].alignment = center_Alignment
                ws4[f"F{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Returns BY QC Analyst
            self.cursor.execute(f"""
                        SELECT evaluated_by, COUNT(*)
                                FROM (SELECT t1.*, t2.evaluated_by
                                FROM returns t1
                                JOIN quality_control t2 ON t1.origin_lot = t2.lot_number
                                WHERE t1.return_date BETWEEN '{date1}' AND '{date2}')
                                GROUP BY evaluated_by

                        """)

            result = self.cursor.fetchall()

            ws4['K1'] = "QC Analyst Most Returns"
            ws4.merge_cells(start_row=1, start_column=11, end_row=1, end_column=12)
            ws4['K1'].alignment = center_Alignment
            ws4['K1'].font = title_color

            ws4["K2"] = "QC Analyst"
            ws4['L2'] = "Returns"
            ws4['K2'].alignment = center_Alignment
            ws4['L2'].alignment = center_Alignment

            ws4.column_dimensions['K'].width = 17
            ws4.column_dimensions['L'].width = 15

            cell_pointer = 3
            for item in result:
                ws4[f"K{cell_pointer}"] = item[0]
                ws4[f"L{cell_pointer}"] = item[1]
                ws4[f"K{cell_pointer}"].alignment = center_Alignment
                ws4[f"L{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Query for getting the Machine's total customer return
            self.cursor.execute(f"""
                WITH splitted_lot AS (

                                    SELECT  *,
                                    (regexp_match(SPLIT_PART(lot_number[1], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_min, 
                                    (regexp_match(SPLIT_PART(lot_number[1], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_max,
                                    (regexp_match(lot_number[1], '[A-Z]+'))[1] as first_lot_code,
                                    (regexp_match(SPLIT_PART(lot_number[2], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_min, 
                                    (regexp_match(SPLIT_PART(lot_number[2], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_max,
                                    (regexp_match(lot_number[2], '[A-Z]+'))[1] as second_lot_code,
                                    (regexp_match(SPLIT_PART(lot_number[3], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_min, 
                                    (regexp_match(SPLIT_PART(lot_number[3], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_max,
                                    (regexp_match(lot_number[3], '[A-Z]+'))[1] as third_lot_code
                                    FROM extruder
                                    )

                SELECT t2.machine, COUNT(*) FROM returns t1
                JOIN splitted_lot t2
                ON (((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.first_lot_min AND t2.first_lot_max 
                                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code) OR
                                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.second_lot_min AND t2.second_lot_max 
                                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = second_lot_code) OR
                                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.third_lot_min AND t2.third_lot_max 
                                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = third_lot_code) OR
                                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER IN (t2.first_lot_min, t2.second_lot_min, t2.third_lot_min)
                                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code)
                                   )
                WHERE return_date BETWEEN '{date1}' AND '{date2}'
                GROUP BY machine


                                            """)

            result = self.cursor.fetchall()

            ws4.merge_cells(start_row=1, start_column=14, end_row=1, end_column=15)
            ws4['N1'] = "Highest Return Per Machine"
            ws4['N1'].alignment = center_Alignment
            ws4['N1'].font = title_color

            ws4["N2"] = "Machine"
            ws4['O2'] = "RETURNS"
            ws4['N2'].alignment = center_Alignment
            ws4['O2'].alignment = center_Alignment

            ws4.column_dimensions['N'].width = 15
            ws4.column_dimensions['O'].width = 15

            cell_pointer = 2
            for item in result:
                ws4[f"N{cell_pointer}"] = item[0]
                ws4[f"O{cell_pointer}"] = item[1]
                ws4[f"N{cell_pointer}"].alignment = center_Alignment
                ws4[f"O{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Query for getting the total kg of Customer Returns by product code
            self.cursor.execute(f"""
                SELECT product_code, SUM(quantity) as total_kg
                FROM returns
                WHERE return_date BETWEEN '{date1}' AND '{date2}'
                GROUP BY product_code
                ORDER BY total_kg
                                                """)

            result = self.cursor.fetchall()

            ws4.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)
            ws4['H1'] = "Highest Return(Kg) Per Code"
            ws4['H1'].alignment = center_Alignment
            ws4['H1'].font = title_color

            ws4["H2"] = "Product Code"
            ws4['I2'] = "RETURNS(kg)"
            ws4['H2'].alignment = center_Alignment
            ws4['I2'].alignment = center_Alignment

            ws4.column_dimensions['H'].width = 15
            ws4.column_dimensions['I'].width = 12

            cell_pointer = 3
            for item in result:
                ws4[f"H{cell_pointer}"] = item[0]
                ws4[f"I{cell_pointer}"] = item[1]
                ws4[f"H{cell_pointer}"].alignment = center_Alignment
                ws4[f"I{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Query for getting the Returns By Each Supervisor
            self.cursor.execute(f"""
                    WITH splitted_lot AS (
                        SELECT  *,
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_max,
                        (regexp_match(lot_number[1], '[A-Z]+'))[1] as first_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_max,
                        (regexp_match(lot_number[2], '[A-Z]+'))[1] as second_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_max,
                        (regexp_match(lot_number[3], '[A-Z]+'))[1] as third_lot_code
                        FROM extruder
                    )

                SELECT t2.supervisor, COUNT(*) FROM returns t1
                JOIN splitted_lot t2
                    ON (((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.first_lot_min AND t2.first_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.second_lot_min AND t2.second_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = second_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.third_lot_min AND t2.third_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = third_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER IN (t2.first_lot_min, t2.second_lot_min, t2.third_lot_min)
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code)
                       )
                WHERE return_date BETWEEN '{date1}' AND '{date2}'
                GROUP BY supervisor

                                            """)

            result = self.cursor.fetchall()

            ws4["V1"] = "Supervisor"
            ws4['W1'] = "RETURNS"
            ws4['V1'].alignment = center_Alignment
            ws4['W1'].alignment = center_Alignment

            ws4.column_dimensions['V'].width = 15
            ws4.column_dimensions['W'].width = 12

            cell_pointer = 2
            for item in result:
                ws4[f"V{cell_pointer}"] = item[0]
                ws4[f"W{cell_pointer}"] = item[1]
                ws4[f"V{cell_pointer}"].alignment = center_Alignment
                ws4[f"W{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Query for getting the Returns By Each Operator
            self.cursor.execute(f"""
                            WITH splitted_lot AS (
                        SELECT  *,
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[1], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_max,
                        (regexp_match(lot_number[1], '[A-Z]+'))[1] as first_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[2], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_max,
                        (regexp_match(lot_number[2], '[A-Z]+'))[1] as second_lot_code,
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_min, 
                        (regexp_match(SPLIT_PART(lot_number[3], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_max,
                        (regexp_match(lot_number[3], '[A-Z]+'))[1] as third_lot_code
                        FROM extruder
                    )

                SELECT t2.operator, COUNT(*) FROM returns t1
                JOIN splitted_lot t2
                    ON (((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.first_lot_min AND t2.first_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.second_lot_min AND t2.second_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = second_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.third_lot_min AND t2.third_lot_max 
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = third_lot_code) OR
                        ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER IN (t2.first_lot_min, t2.second_lot_min, t2.third_lot_min)
                    AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code)
                       )
                WHERE return_date BETWEEN '{date1}' AND '{date2}'
                GROUP BY operator

                                            """)
            result = self.cursor.fetchall()

            ws4.merge_cells(start_row=1, start_column=17, end_row=1, end_column=18)
            ws4['Q1'] = 'Highest Returns By Operator'
            ws4['Q1'].font = title_color
            ws4['Q1'].alignment = center_Alignment

            ws4["Q2"] = "Operator"
            ws4['R2'] = "RETURNS"
            ws4['Q2'].alignment = center_Alignment
            ws4['R2'].alignment = center_Alignment

            ws4.column_dimensions['Q'].width = 15
            ws4.column_dimensions['R'].width = 15

            cell_pointer = 3
            for item in result:
                ws4[f"Q{cell_pointer}"] = item[0]
                ws4[f"R{cell_pointer}"] = item[1]
                ws4[f"Q{cell_pointer}"].alignment = center_Alignment
                ws4[f"R{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Worksheet 3
            ws3 = wb.create_sheet("4) INTERNAL REJECT")

            # Query for getting the Highest Failed First Run by product code
            self.cursor.execute(f"""
            SELECT product_code, COUNT(*) as failed_count
            FROM quality_control_tbl2
            WHERE status = 'Failed' AND qc_type = 'NEW' AND evaluation_date BETWEEN '{date1}' AND 
                    '{date2}'
            GROUP BY product_code
            ORDER BY failed_count DESC

            """)
            result = self.cursor.fetchall()

            ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws3['A1'] = "Highest Failed First Run"
            ws3['A1'].font = title_color

            ws3["A2"] = "Product Code"
            ws3['B2'] = "Count"
            ws3['A2'].alignment = center_Alignment
            ws3['B2'].alignment = center_Alignment

            ws3.column_dimensions['B'].width = 10
            ws3.column_dimensions['A'].width = 15

            cell_pointer = 3
            for item in result:
                ws3[f"A{cell_pointer}"] = item[0]
                ws3[f"B{cell_pointer}"] = item[1]
                ws3[f"A{cell_pointer}"].alignment = center_Alignment
                ws3[f"B{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            # Query for getting the total failed by product code
            self.cursor.execute(f"""
                SELECT product_code, COUNT(*) as failed_count
                FROM quality_control_tbl2
                WHERE status = 'Failed' AND evaluation_date BETWEEN '{date1}' AND 
                    '{date2}'
                GROUP BY product_code
                ORDER BY failed_count ASC
                LIMIT 20

            """)
            result = self.cursor.fetchall()

            ws3["D2"] = "Product Code"
            ws3['E2'] = "Count"

            cell_pointer = 3
            for item in result:
                ws3[f"D{cell_pointer}"] = item[0]
                ws3[f"E{cell_pointer}"] = item[1]
                ws3[f"D{cell_pointer}"].alignment = center_Alignment
                ws3[f"E{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws3.column_dimensions['D'].width = 15
            ws3.column_dimensions['E'].width = 10

            ws3.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
            ws3['D1'] = "LOWEST Failed First Run"
            ws3['D1'].alignment = center_Alignment
            ws3['D1'].font = title_color

            # Query for getting the number of failed qc by product code
            self.cursor.execute(f"""
                SELECT product_code, COUNT(*) as failed_count
                FROM quality_control_tbl2
                WHERE status = 'Failed' AND evaluation_date BETWEEN '{date1}' AND 
                    '{date2}'
                GROUP BY product_code
                ORDER BY failed_count DESC
            """)

            result = self.cursor.fetchall()

            ws3["G2"] = "Product Code"
            ws3['H2'] = "Count"

            cell_pointer = 3
            for item in result:
                ws3[f"G{cell_pointer}"] = item[0]
                ws3[f"H{cell_pointer}"] = item[1]
                ws3[f"G{cell_pointer}"].alignment = center_Alignment
                ws3[f"H{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws3.column_dimensions['G'].width = 15
            ws3.column_dimensions['H'].width = 10

            ws3.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
            ws3['G1'] = "Highest Lot Failed"
            ws3['G1'].alignment = center_Alignment
            ws3['G1'].font = title_color

            # Query or getting the most used formula id
            self.cursor.execute(f"""
                            SELECT product_code, formula_id, COUNT(formula_id) FROM quality_control_tbl2
                            WHERE status = 'Failed' AND evaluation_date BETWEEN '{date1}' AND 
                            '{date2}'
                            GROUP BY product_code, formula_id
                            ORDER BY count DESC, product_code 
                        """)

            result = self.cursor.fetchall()

            ws3["J2"] = "Product Code"
            ws3['K2'] = "FN"
            ws3['L2'] = "Count"

            cell_pointer = 3
            for item in result:
                ws3[f"J{cell_pointer}"] = item[0]
                ws3[f"K{cell_pointer}"] = item[1]
                ws3[f"L{cell_pointer}"] = item[2]
                ws3[f"J{cell_pointer}"].alignment = center_Alignment
                ws3[f"K{cell_pointer}"].alignment = center_Alignment
                ws3[f"L{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws3.column_dimensions['J'].width = 15
            ws3.column_dimensions['K'].width = 10

            ws3.merge_cells(start_row=1, start_column=10, end_row=1, end_column=12)
            ws3['J1'] = "FN per Product Code Fails"
            ws3['J1'].alignment = center_Alignment
            ws3['J1'].font = title_color

            # Failed First Run by Operator
            self.cursor.execute(f"""
                WITH splitted_lot AS (

                    SELECT  *,
                    (regexp_match(SPLIT_PART(lot_number[1], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_min, 
                    (regexp_match(SPLIT_PART(lot_number[1], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_max,
                    (regexp_match(lot_number[1], '[A-Z]+'))[1] as first_lot_code,
                    (regexp_match(SPLIT_PART(lot_number[2], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_min, 
                    (regexp_match(SPLIT_PART(lot_number[2], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_max,
                    (regexp_match(lot_number[2], '[A-Z]+'))[1] as second_lot_code,
                    (regexp_match(SPLIT_PART(lot_number[3], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_min, 
                    (regexp_match(SPLIT_PART(lot_number[3], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_max,
                    (regexp_match(lot_number[3], '[A-Z]+'))[1] as third_lot_code
                    FROM extruder
                    )
                    SELECT operator, SUM(lot_count) as total_ffr FROM 

                (WITH lot_range as (
                                SELECT * ,  (regexp_matches(SPLIT_PART(lot_number, '-', 1), '(\d+)[A-Z]', 'g'))[1]::INTEGER as first_lot, 
                                (regexp_matches(SPLIT_PART(lot_number, '-', 2), '(\d+)[A-Z]', 'g'))[1]::INTEGER as last_lot

                                FROM quality_control
                                WHERE evaluated_on BETWEEN '{date1}' AND '{date2}'

                                )

                                SELECT *, 
                                CASE 
                                    WHEN last_lot IS NULL THEN 1
                                    ELSE (last_lot - first_lot) + 1
                                    END AS lot_count
                                FROM lot_range	) t1


                JOIN splitted_lot t2
                ON (((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.first_lot_min AND t2.first_lot_max 
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code) OR
                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.second_lot_min AND t2.second_lot_max 
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = second_lot_code) OR
                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.third_lot_min AND t2.third_lot_max 
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = third_lot_code) OR
                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER IN (t2.first_lot_min, t2.second_lot_min, t2.third_lot_min)
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code)
                   )

                WHERE status = 'Failed' AND qc_type = 'NEW'
                GROUP BY operator
            """)

            result = self.cursor.fetchall()

            ws3["N2"] = "Operator"
            ws3['O2'] = "Failed First Run"

            cell_pointer = 3
            for item in result:
                ws3[f"N{cell_pointer}"] = item[0]
                ws3[f"O{cell_pointer}"] = item[1]
                ws3[f"N{cell_pointer}"].alignment = center_Alignment
                ws3[f"O{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws3.column_dimensions['N'].width = 15
            ws3.column_dimensions['O'].width = 15

            ws3.merge_cells(start_row=1, start_column=14, end_row=1, end_column=15)
            ws3['N1'] = "Failed First Run by Operators"
            ws3['N1'].alignment = center_Alignment
            ws3['N1'].font = title_color

            # Query for getting the Failed First Run by Machine
            self.cursor.execute(f"""
                WITH splitted_lot AS (

                    SELECT  *,
                    (regexp_match(SPLIT_PART(lot_number[1], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_min, 
                    (regexp_match(SPLIT_PART(lot_number[1], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as first_lot_max,
                    (regexp_match(lot_number[1], '[A-Z]+'))[1] as first_lot_code,
                    (regexp_match(SPLIT_PART(lot_number[2], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_min, 
                    (regexp_match(SPLIT_PART(lot_number[2], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as second_lot_max,
                    (regexp_match(lot_number[2], '[A-Z]+'))[1] as second_lot_code,
                    (regexp_match(SPLIT_PART(lot_number[3], '-', 1), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_min, 
                    (regexp_match(SPLIT_PART(lot_number[3], '-', 2), '(\d+)[A-Z]'))[1]::INTEGER as third_lot_max,
                    (regexp_match(lot_number[3], '[A-Z]+'))[1] as third_lot_code
                    FROM extruder
                    )

                    SELECT machine, SUM(lot_count) as total_ffr FROM 


                    (WITH lot_range as (
                                    SELECT * ,  (regexp_matches(SPLIT_PART(lot_number, '-', 1), '(\d+)[A-Z]', 'g'))[1]::INTEGER as first_lot, 
                                    (regexp_matches(SPLIT_PART(lot_number, '-', 2), '(\d+)[A-Z]', 'g'))[1]::INTEGER as last_lot

                                    FROM quality_control
                                    WHERE evaluated_on BETWEEN '{date1}' AND '{date2}'

                                    )

                                    SELECT *, 
                                    CASE 
                                        WHEN last_lot IS NULL THEN 1
                                        ELSE (last_lot - first_lot) + 1
                                        END AS lot_count
                                    FROM lot_range	) t1


                    JOIN splitted_lot t2
                ON (((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.first_lot_min AND t2.first_lot_max 
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code) OR
                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.second_lot_min AND t2.second_lot_max 
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = second_lot_code) OR
                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER BETWEEN t2.third_lot_min AND t2.third_lot_max 
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = third_lot_code) OR
                    ((regexp_match(t1.lot_number, '(\d+)[A-Z]+'))[1]::INTEGER IN (t2.first_lot_min, t2.second_lot_min, t2.third_lot_min)
                AND (regexp_match(t1.lot_number, '[A-Z]+'))[1] = first_lot_code)
                   )

                    WHERE status = 'Failed' AND qc_type = 'NEW'
                    GROUP BY machine

            """)
            result = self.cursor.fetchall()

            ws3["Q2"] = "Machine"
            ws3['R2'] = "Failed First Run"

            cell_pointer = 3
            for item in result:
                ws3[f"Q{cell_pointer}"] = item[0]
                ws3[f"R{cell_pointer}"] = item[1]
                ws3[f"Q{cell_pointer}"].alignment = center_Alignment
                ws3[f"R{cell_pointer}"].alignment = center_Alignment
                cell_pointer += 1

            ws3.column_dimensions['Q'].width = 15
            ws3.column_dimensions['R'].width = 15

            ws3.merge_cells(start_row=1, start_column=17, end_row=1, end_column=18)
            ws3['Q1'] = "Failed First Run by Machine"
            ws3['Q1'].alignment = center_Alignment
            ws3['Q1'].font = title_color

            # Open QFileDialog
            filename, _ = QFileDialog.getSaveFileName(self.qc_widget, "Save File",
                                                      r"C:\Users\Administrator\Desktop",
                                                      "Excel Files (*.xlsx)", options=QFileDialog.Options())

            if filename:
                if not filename.lower().endswith('.xlsx'):
                    filename += '.xlsx'

                wb.save(filename)
                QMessageBox.information(self.qc_widget, "Export Success", "File Successfully Exported!")

        def show_items():
            try:
                # Querying the selected item in the database
                id = self.qc_table.selectedItems()[0].text()
                self.cursor.execute(f"""
                            SELECT * FROM quality_control
                            WHERE id = '{id}'

                            """)
                result = self.cursor.fetchall()[0]

                # Unpacking the items
                lot_num, product_code, customer, status, remarks, action = result[1:7]
                evaluated_by, evaluated_date, encoded_on, updated_by, updated_on = result[8:13]
                time_endorsed = result[13]
                qc_type = result[15]
                formula_id = result[16]

                # showing the Selected Items
                customer_selected.setText(customer)
                productCode_selected.setText(product_code)
                result_selected.setText(status)
                evaluatedBy_selected.setText(evaluated_by)
                evaluatedDate_selected.setText(str(evaluated_date.strftime("%m-%d-%Y %H:%M:%S")))
                encodedDate_selected.setText(str(encoded_on.strftime("%m-%d-%Y %H:%M:%S")))
                remarks_box.setText(remarks)

                updatedBy_val1.setText(updated_by)
                time_endorsed_val.setText(str(time_endorsed.strftime("%m-%d-%Y %H:%M:%S")))
                qc_type_selected.setText(qc_type)
            except:
                QMessageBox.information(self.qc_widget, "Selection Error", "No Items Selected")
                self.qc_table.clearSelection()

        def qc_returns():

            def insert_entry():
                # Check if the lot number is not entered yet in the Database
                self.cursor.execute(f"""
                SELECT * FROM returns
                WHERE  lot_number = '{lot_input.text()}'

                """)
                result = self.cursor.fetchall()

                try:
                    if len(result) == 1:
                        QMessageBox.information(self.body_widget, "Data Exist", "Lot Number already exist.")
                        lot_input.clear()
                        quantity_input.clear()
                        remarks_input.clear()
                        lot_input.setFocus()
                        origin_lot.clear()
                        return
                    else:
                        self.cursor.execute(f"""
                                            INSERT INTO returns (lot_number, quantity, product_code, customer, formula_id,
                                            remarks, return_date, origin_lot, return_id)
                                            VALUES('{lot_input.text()}', '{quantity_input.text()}', '{product_code_input.text()}', 
                                            '{customer_input.text()}','{formulaID_input.text()}', '{remarks_input.toPlainText()}',
                                            '{date_return.text()}', '{origin_lot.text()}', {return_id_input.text()})
                                            """)
                        self.conn.commit()
                        # Clear the Widgets
                        lot_input.clear()
                        quantity_input.clear()
                        remarks_input.clear()
                        origin_lot.clear()
                        QtWidgets.QMessageBox.information(self.qc_widget, "SUCCESS", "Successfully Inserted Data")
                        lot_input.setFocus()
                        show_table()

                except Exception as e:
                    self.conn.rollback()
                    QMessageBox.critical(self.qc_widget, "Insert Error", str(e))

            def autofill():

                self.cursor.execute(f"""
                SELECT * FROM quality_control_tbl2
                WHERE lot_number = '{lot_input.text()}'
                """)
                result = self.cursor.fetchone()

                if result != None:
                    self.cursor.execute(f"""
                    SELECT t2.lot_number as origin_lot, t1.lot_number, t1.product_code, t1.formula_id, t2.customer
                    FROM quality_control_tbl2 t1
                    JOIN quality_control t2 ON t1.id = t2.id
                    WHERE t1.lot_number = '{lot_input.text()}'

                    """)
                    result = self.cursor.fetchall()[0]

                    # Set the Text
                    origin_lot.setText(result[0])
                    product_code_input.setText(result[2])
                    formulaID_input.setText(str(result[3]))
                    customer_input.setText(result[4])
                else:
                    QMessageBox.information(self.qc_widget, "Not Found", "LOT NUMBER Does not Exist!")
                    product_code_input.clear()
                    formulaID_input.clear()
                    customer_input.clear()

            def show_table():

                self.cursor.execute("""
                SELECT return_id, lot_number, quantity, product_code, customer, formula_id, origin_lot, return_date, remarks  
                FROM returns
                ORDER BY return_id DESC

                """)
                result = self.cursor.fetchall()

                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        returns_table.setItem(i, j, item)

            def search_table():

                returns_table.clear()
                returns_table.setHorizontalHeaderLabels(
                    ["Return ID", "Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                     "Origin Lot", "Return Date", "Remarks"])

                self.cursor.execute(f"""

                SELECT return_id, lot_number, quantity, product_code, customer, formula_id, return_date, remarks  
                FROM returns
                WHERE lot_number ILIKE '%{search_bar.text()}%'


                """)
                result = self.cursor.fetchall()

                for i in range(len(result)):
                    for j in range(len(result[i])):
                        item = QTableWidgetItem(str(result[i][j]))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        returns_table.setItem(i, j, item)

                returns_table.show()

            def filter_data():
                if masterbatch_checkbox.isChecked() == True and dryColor_checkbox.isChecked() == False:
                    self.cursor.execute("""
                    SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks
                    FROM returns
                    WHERE RIGHT(lot_number, 2) ~ '^[A-Za-z]';
                    """)
                    result = self.cursor.fetchall()

                    if len(result) > 20:
                        returns_table.setColumnCount(len(result))

                    returns_table.clear()
                    returns_table.setHorizontalHeaderLabels(
                        ["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                         "Return Date", "Remarks"])

                    for i in range(len(result)):
                        for j in range(len(result[i])):
                            item = QTableWidgetItem(str(result[i][j]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            returns_table.setItem(i, j, item)

                    returns_table.show()
                elif masterbatch_checkbox.isChecked() == False and dryColor_checkbox.isChecked() == True:
                    self.cursor.execute("""
                                        SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks
                                        FROM returns
                                        WHERE RIGHT(lot_number, 2) !~ '^[A-Za-z]';
                                        """)
                    result = self.cursor.fetchall()

                    if len(result) > 20:
                        returns_table.setColumnCount(len(result))

                    returns_table.clear()
                    returns_table.setHorizontalHeaderLabels(
                        ["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                         "Return Date", "Remarks"])

                    for i in range(len(result)):
                        for j in range(len(result[i])):
                            item = QTableWidgetItem(str(result[i][j]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            returns_table.setItem(i, j, item)

                    returns_table.show()
                else:
                    self.cursor.execute("""
                                        SELECT lot_number, quantity, product_code, customer, formula_id, return_date, remarks
                                        FROM returns
                                        """)
                    result = self.cursor.fetchall()

                    if len(result) > 20:
                        returns_table.setColumnCount(len(result))

                    returns_table.clear()
                    returns_table.setHorizontalHeaderLabels(
                        ["Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                         "Return Date", "Remarks"])

                    for i in range(len(result)):
                        for j in range(len(result[i])):
                            item = QTableWidgetItem(str(result[i][j]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            returns_table.setItem(i, j, item)

                    returns_table.show()

            def clear_entry():
                lot_input.clear()
                quantity_input.clear()
                remarks_input.clear()

            def update_returns():

                def save_button_clicked():
                    self.cursor.execute(f"""
                                    UPDATE returns
                                    SET lot_number = '{lot_input.text()}', quantity = {quantity_input.text()}, product_code = '{product_code_input.text()}',
                                    customer = '{customer_input.text()}', remarks = '{remarks_input.toPlainText()}', return_date = '{date_return.text()}',
                                    edited = true, last_edited = '{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                                    WHERE lot_number = '{items[0]}'

                                    """)
                    self.conn.commit()

                items = [item.text() for item in returns_table.selectedItems()]

                self.edit_returns_widget = QWidget()
                self.edit_returns_widget.setGeometry(300, 100, 340, 500)
                self.edit_returns_widget.setWindowTitle('Edit')
                self.edit_returns_widget.show()

                widget1 = QWidget(self.edit_returns_widget)
                widget1.setGeometry(0, 0, 340, 450)
                widget1.show()

                form_layout = QFormLayout(widget1)
                form_layout.setVerticalSpacing(20)

                # Set Font
                font = QtGui.QFont("Arial", 9)

                lot_label = QLabel()
                lot_label.setText("LOT NUMBER")
                lot_label.setFont(font)

                lot_input = QLineEdit()
                lot_input.setStyleSheet("background-color: rgb(255, 255, 0)")
                lot_input.setAlignment(Qt.AlignCenter)
                lot_input.setFixedHeight(25)
                lot_input.setText(items[0])

                quantity_label = QLabel()
                quantity_label.setText("QUANTITY")
                quantity_label.setFont(font)

                quantity_input = QLineEdit()
                quantity_input.setStyleSheet("background-color: rgb(255, 255, 0)")
                quantity_input.setAlignment(Qt.AlignCenter)
                quantity_input.setFixedHeight(25)
                quantity_input.setText(items[1])

                product_code_label = QLabel()
                product_code_label.setText("PRODUCT CODE")
                product_code_label.setFont(font)

                product_code_input = QLineEdit()
                product_code_input.setStyleSheet("background-color: rgb(240, 240, 240)")
                product_code_input.setAlignment(Qt.AlignCenter)
                product_code_input.setFixedHeight(25)
                product_code_input.setText(items[2])

                formulaID_label = QLabel()
                formulaID_label.setText("FORMULA ID")
                formulaID_label.setFont(font)

                formulaID_input = QLineEdit()
                formulaID_input.setStyleSheet("background-color: rgb(255, 255, 0)")
                formulaID_input.setAlignment(Qt.AlignCenter)
                formulaID_input.setFixedHeight(25)
                formulaID_input.setEnabled(False)
                formulaID_input.setText(items[4])

                customer_label = QLabel()
                customer_label.setText("CUSTOMER")
                customer_label.setFont(font)

                customer_input = QLineEdit()
                customer_input.setStyleSheet("background-color: rgb(255, 255, 0)")
                customer_input.setFixedHeight(25)
                customer_input.setText(items[3])
                customer_input.setAlignment(Qt.AlignCenter)

                remarks_label = QLabel()
                remarks_label.setText("REMARKS")
                remarks_label.setFont(font)
                remarks_label.setAlignment(Qt.AlignRight)

                remarks_input = QTextEdit()
                remarks_input.setFont(font)
                remarks_input.setFixedHeight(100)
                remarks_input.setText(items[7])

                date_label = QLabel()
                date_label.setText("RETURN DATE")
                date_label.setFont(font)
                date_label.setAlignment(Qt.AlignRight)

                date_return = QDateEdit()
                date_return.setStyleSheet("background-color: rgb(255, 255, 0)")
                date_return.setFont(font)
                date_return.setFixedHeight(25)
                date_return.setDate(datetime.strptime(items[6], '%Y-%m-%d'))
                date_return.setDisplayFormat("MM-dd-yyyy")

                origin_lot_label = QLabel()
                origin_lot_label.setText("ORIGIN LOT")
                origin_lot_label.setFont(font)
                origin_lot_label.setAlignment(Qt.AlignRight)

                origin_lot = QLineEdit()
                origin_lot.setStyleSheet("background-color: rgb(255, 255, 0)")
                origin_lot.setFixedHeight(25)
                origin_lot.setEnabled(False)
                origin_lot.setText(items[5])
                origin_lot.setAlignment(Qt.AlignCenter)

                form_layout.addRow(lot_label, lot_input)
                form_layout.addRow(quantity_label, quantity_input)
                form_layout.addRow(product_code_label, product_code_input)
                form_layout.addRow(formulaID_label, formulaID_input)
                form_layout.addRow(customer_label, customer_input)
                form_layout.addRow(origin_lot_label, origin_lot)
                form_layout.addRow(date_label, date_return)
                form_layout.addRow(remarks_label, remarks_input)

                save_button = QPushButton(self.edit_returns_widget)
                save_button.setGeometry(90, 455, 60, 25)
                save_button.setText('SAVE')
                save_button.clicked.connect(save_button_clicked)
                save_button.show()

                cancel_button = QPushButton(self.edit_returns_widget)
                cancel_button.setGeometry(190, 455, 60, 25)
                cancel_button.setText('CANCEL')
                cancel_button.clicked.connect(lambda: self.edit_returns_widget.close())
                cancel_button.show()

            try:
                self.qc_widget.deleteLater()
            except:
                pass

            self.qc_widget = QtWidgets.QWidget(self.main_widget)
            self.qc_widget.setGeometry(0, 0, 991, 700)
            self.qc_widget.setStyleSheet("background-color: rgb(240,240,240);")
            self.qc_widget.show()

            self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
            self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
            self.qcBtn_topBorder.setStyleSheet('background-color: rgb(92, 154, 255)')
            self.qcBtn_topBorder.show()

            self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_TableBtn.setGeometry(55, 0, 70, 30)
            self.qc_TableBtn.setText("QC DATA")
            self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
            self.qc_TableBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_TableBtn.setStyleSheet(
                "color: white; border: none;")
            self.qc_TableBtn.clicked.connect(self.quality_control)
            self.qc_TableBtn.setShortcut("F1")
            self.qc_TableBtn.show()

            self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_addEntryBtn.setGeometry(145, 0, 100, 30)
            self.qc_addEntryBtn.setText("QC ENTRY")
            self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
            self.qc_addEntryBtn.setStyleSheet(
                "color: white; border: none;")
            self.qc_addEntryBtn.clicked.connect(evaluation_entry)
            self.qc_addEntryBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_addEntryBtn.setShortcut("F2")
            self.qc_addEntryBtn.show()

            self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_dataBtn.setGeometry(270, 0, 70, 30)
            self.qc_dataBtn.setText("QC Data")
            self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
            self.qc_dataBtn.setStyleSheet(
                "color: white; border: none; ")
            self.qc_dataBtn.clicked.connect(show_qc_data)
            self.qc_dataBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_dataBtn.setShortcut("F3")
            self.qc_dataBtn.show()

            self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.dashboardBtn.setGeometry(370, 0, 85, 30)
            self.dashboardBtn.setText("DASHBOARD")
            self.dashboardBtn.setCursor(Qt.PointingHandCursor)
            self.dashboardBtn.setStyleSheet("color: white; border: none")
            self.dashboardBtn.clicked.connect(show_dashboards)
            self.dashboardBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.dashboardBtn.setShortcut("F4")
            self.dashboardBtn.show()

            self.body_widget = QtWidgets.QWidget(self.qc_widget)
            self.body_widget.setGeometry(0, 30, 991, 721)
            self.body_widget.setStyleSheet(
                "background-color:rgb(239, 243, 254); border-top : 1px solid rgb(160, 160, 160);")
            self.body_widget.show()

            self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
            self.qc_returns.setGeometry(500, 0, 70, 30)
            self.qc_returns.setText("RETURNS")
            self.qc_returns.setCursor(Qt.PointingHandCursor)
            self.qc_returns.setStyleSheet(
                "color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;")
            self.qc_returns.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
            self.qc_returns.clicked.connect(qc_returns)
            self.qc_returns.setShortcut("F5")
            self.qc_returns.show()

            header_widget = QWidget(self.body_widget)
            header_widget.setGeometry(0, 0, 991, 60)
            header_widget.setStyleSheet('border: 1px solid black; background-color:rgb(239, 243, 254)')
            header_widget.show()

            page_title = QtWidgets.QLabel(header_widget)
            page_title.setGeometry(520, 10, 200, 40)
            page_title.setText('RETURNS')
            page_title.setFont(QtGui.QFont('Arial Black', 22))
            page_title.setAlignment(Qt.AlignCenter)
            page_title.setStyleSheet('border: none; color: red')
            page_title.show()

            edited_checkbox = QCheckBox(header_widget)
            edited_checkbox.move(15, 5)
            edited_checkbox.setStyleSheet("border: none;")
            edited_checkbox.show()

            edited_label = QLabel(header_widget)
            edited_label.setGeometry(30, 4, 90, 15)
            edited_label.setStyleSheet("border: none;")
            edited_label.setText("EDITED RECORDS")
            edited_label.setFont(QtGui.QFont("Arial", 8))
            edited_label.show()

            masterbatch_checkbox = QCheckBox(header_widget)
            masterbatch_checkbox.move(125, 5)
            masterbatch_checkbox.setStyleSheet('border:none')
            masterbatch_checkbox.stateChanged.connect(filter_data)
            masterbatch_checkbox.show()

            masterbatch_label = QLabel(header_widget)
            masterbatch_label.setGeometry(140, 4, 90, 15)
            masterbatch_label.setStyleSheet("border: none;")
            masterbatch_label.setText("MASTERBATCH")
            masterbatch_label.setFont(QtGui.QFont("Arial", 8))
            masterbatch_label.show()

            dryColor_checkbox = QCheckBox(header_widget)
            dryColor_checkbox.move(235, 5)
            dryColor_checkbox.setStyleSheet('border:none')
            dryColor_checkbox.stateChanged.connect(filter_data)
            dryColor_checkbox.show()

            drycolor_label = QLabel(header_widget)
            drycolor_label.setGeometry(250, 4, 90, 15)
            drycolor_label.setStyleSheet("border: none;")
            drycolor_label.setText("DRYCOLOR")
            drycolor_label.setFont(QtGui.QFont("Arial", 8))
            drycolor_label.show()

            search_bar = QtWidgets.QLineEdit(header_widget)
            search_bar.setGeometry(770, 5, 150, 25)
            search_bar.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 17);")
            search_bar.setFont(QtGui.QFont("Arial", 9))
            search_bar.setPlaceholderText("Lot Number")
            search_bar.show()

            search_btn = QtWidgets.QPushButton(header_widget)
            search_btn.setGeometry(925, 5, 60, 25)
            search_btn.setStyleSheet("border: 1px solid rgb(171, 173, 179);")
            search_btn.setText("Search")
            search_btn.setShortcut("Ctrl+Return")
            search_btn.clicked.connect(search_table)
            search_btn.show()

            self.body_widget.setStyleSheet("border: none")
            entry_widget = QtWidgets.QWidget(self.body_widget)
            entry_widget.setGeometry(0, 60, 340, 550)
            entry_widget.setStyleSheet("background-color: white;")
            entry_widget.show()

            form_layout = QFormLayout(entry_widget)
            form_layout.setVerticalSpacing(10)

            # Set Font
            font = QtGui.QFont("Arial", 9)

            return_id_label = QLabel()
            return_id_label.setText('RETURN ID')
            return_id_label.setFont(font)

            return_id_input = QLineEdit()
            return_id_input.setStyleSheet("background-color: rgb(255, 255, 0); border: 1px solid black")
            return_id_input.setAlignment(Qt.AlignCenter)
            return_id_input.setFixedHeight(25)

            lot_label = QLabel()
            lot_label.setText("LOT NUMBER")
            lot_label.setFont(font)

            lot_input = QLineEdit()
            lot_input.setStyleSheet("background-color: rgb(255, 255, 0); border: 1px solid black")
            lot_input.setAlignment(Qt.AlignCenter)
            lot_input.setFixedHeight(25)
            lot_input.editingFinished.connect(autofill)

            quantity_label = QLabel()
            quantity_label.setText("QUANTITY")
            quantity_label.setFont(font)

            quantity_input = QLineEdit()
            quantity_input.setStyleSheet("background-color: rgb(255, 255, 0); border: 1px solid black")
            quantity_input.setAlignment(Qt.AlignCenter)
            quantity_input.setFixedHeight(25)

            product_code_label = QLabel()
            product_code_label.setText("PRODUCT CODE")
            product_code_label.setFont(font)

            product_code_input = QLineEdit()
            product_code_input.setStyleSheet("background-color: rgb(240, 240, 240)")
            product_code_input.setAlignment(Qt.AlignCenter)
            product_code_input.setFixedHeight(25)
            product_code_input.setEnabled(False)

            formulaID_label = QLabel()
            formulaID_label.setText("FORMULA ID")
            formulaID_label.setFont(font)

            formulaID_input = QLineEdit()
            formulaID_input.setStyleSheet("background-color: rgb(240, 240, 240)")
            formulaID_input.setAlignment(Qt.AlignCenter)
            formulaID_input.setFixedHeight(25)
            formulaID_input.setEnabled(False)

            customer_label = QLabel()
            customer_label.setText("CUSTOMER")
            customer_label.setFont(font)

            customer_input = QLineEdit()
            customer_input.setStyleSheet("background-color: rgb(240, 240, 240)")
            customer_input.setFixedHeight(25)
            customer_input.setEnabled(False)
            customer_input.setAlignment(Qt.AlignCenter)

            remarks_label = QLabel()
            remarks_label.setText("REMARKS")
            remarks_label.setFont(font)
            remarks_label.setAlignment(Qt.AlignRight)

            remarks_input = QTextEdit()
            remarks_input.setStyleSheet("background-color: rgb(255, 255, 0); border: 1px solid black")
            remarks_input.setFont(font)
            remarks_input.setFixedHeight(160)

            date_label = QLabel()
            date_label.setText("RETURN DATE")
            date_label.setFont(font)
            date_label.setAlignment(Qt.AlignRight)

            date_return = QDateEdit()
            date_return.setStyleSheet("background-color: rgb(255, 255, 0); border: 1px solid black")
            date_return.setFont(font)
            date_return.setFixedHeight(25)
            today = date.today()
            date_return.setDate(today)
            date_return.setDisplayFormat("MM-dd-yyyy")

            origin_lot_label = QLabel()
            origin_lot_label.setText("ORIGIN LOT")
            origin_lot_label.setFont(font)
            origin_lot_label.setAlignment(Qt.AlignRight)

            origin_lot = QLineEdit()
            origin_lot.setStyleSheet("background-color: rgb(240, 240, 240)")
            origin_lot.setFixedHeight(25)
            origin_lot.setEnabled(False)
            origin_lot.setAlignment(Qt.AlignCenter)

            form_layout.addRow(return_id_label, return_id_input)
            form_layout.addRow(lot_label, lot_input)
            form_layout.addRow(quantity_label, quantity_input)
            form_layout.addRow(product_code_label, product_code_input)
            form_layout.addRow(formulaID_label, formulaID_input)
            form_layout.addRow(customer_label, customer_input)
            form_layout.addRow(origin_lot_label, origin_lot)
            form_layout.addRow(date_label, date_return)
            form_layout.addRow(remarks_label, remarks_input)

            # Create Return Table

            returns_table = QTableWidget(self.body_widget)
            returns_table.setGeometry(340, 60, 651, 550)
            returns_table.setColumnCount(9)
            returns_table.setRowCount(20)
            returns_table.verticalHeader().setVisible(False)
            returns_table.setHorizontalHeaderLabels(
                ["Return ID", "Lot Number", "Quantity", "Product Code", "Customer", "Formula ID",
                 "Origin Lot", "Return Date", "Remarks"])
            returns_table.setColumnWidth(1, 120)
            returns_table.setColumnWidth(3, 120)
            returns_table.setColumnWidth(4, 200)
            returns_table.setColumnWidth(5, 115)
            returns_table.setColumnWidth(6, 200)

            returns_table.horizontalHeader().setStyleSheet("""
            QHeaderView::section{
            font-weight: bold;
            background-color: rgb(0, 109, 189);
            color: white;
                }  

            """)
            returns_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
            returns_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)

            show_table()
            returns_table.show()

            save_button = QPushButton(self.body_widget)
            save_button.setGeometry(50, 630, 60, 25)
            save_button.setText("Save")
            save_button.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                      "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            save_button.setCursor(Qt.PointingHandCursor)
            save_button.clicked.connect(insert_entry)
            save_button.setShortcut("Return")
            save_button.show()

            clear_button = QPushButton(self.body_widget)
            clear_button.setGeometry(150, 630, 60, 25)
            clear_button.setText("Clear")
            clear_button.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                       "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            clear_button.setCursor(Qt.PointingHandCursor)
            clear_button.clicked.connect(clear_entry)

            edit_button = QPushButton(self.body_widget)
            edit_button.setGeometry(250, 630, 60, 25)
            edit_button.setText("Edit")
            edit_button.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                      "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            edit_button.clicked.connect(update_returns)
            edit_button.show()

            clear_button.show()

        def update_entry():

            # Check if User have permission to enter QC data.
            if qc_access:
                pass
            else:
                QMessageBox.critical(self.qc_widget, 'Restricted Access',
                                     "You Dont Have Permission. \n Contact the Admin.")
                return

            def save_update():
                # Update quality_control table
                self.cursor.execute(f"""
                UPDATE quality_control 
                SET customer = '{customer_list.currentText()}', formula_id = '{formulaID_input.text()}',
                evaluated_on = '{date_started.text()}', status = '{test_result_dropdown.currentText()}', 
                time_endorsed = '{endorsed_date.text()}',
                remarks = '{remarks_box.toPlainText()}', edited = true, updated_on = '{datetime.now()}'::timestamp,
				status_changed = 
				    CASE 
				        WHEN (SELECT status_changed FROM quality_control WHERE lot_number = '{lot_number}') = 'true' THEN TRUE
                        WHEN (SELECT status FROM quality_control WHERE lot_number = '{lot_number}') != '{test_result_dropdown.currentText()}'
                        THEN true
                        else false
                        END
                WHERE lot_number = '{lot_number}'

                """)

                self.cursor.execute(f"""
                    INSERT INTO qc_logs
                    VALUES('{datetime.now().strftime('%Y-%m-%d %H:%M')}', 'UPDATE', {selected[0].text()}, '{selected[1].text()}', 
                            '{selected[2].text()}', '{selected[3].text()}', '{selected[4].text()}', '{selected[5].text()}')

                """)

                self.conn.commit()

                # Update quality_control_tbl2
                self.cursor.execute(f"""
                UPDATE quality_control_tbl2
                SET status = '{test_result_dropdown.currentText()}', evaluation_date = '{date_started.text()}',
                formula_id = '{formulaID_input.text()}', date_endorsed = '{endorsed_date.text()}'
                WHERE id = '{id}'

                """)

                self.conn.commit()
                self.update_qc_widget.close()
                self.quality_control()
                QMessageBox.information(self.qc_widget, "Success", "Update Successful")

            selected = self.qc_table.selectedItems()
            try:
                lot_number = selected[1].text()
            except IndexError:
                QMessageBox.information(self.qc_widget, '', 'No Selected Item.')
                return

            self.cursor.execute(f"""
            SELECT * FROM quality_control
            WHERE lot_number = '{lot_number}'

            """)
            result = self.cursor.fetchone()

            id = result[0]
            lot_number = result[1].strip()
            product_code = result[2].strip()
            customer = result[3]
            status = result[4]
            remarks = result[5].strip()
            date_evaluated = result[9]
            formula_id = result[-2]
            endorsed_date_val = result[13]

            if selected:
                self.update_qc_widget = QWidget()
                self.update_qc_widget.setGeometry(0, 0, 350, 550)
                self.update_qc_widget.setFixedSize(350, 550)
                self.update_qc_widget.setWindowModality(Qt.ApplicationModal)
                self.update_qc_widget.setStyleSheet("background-color: ")
                self.update_qc_widget.setWindowTitle("Edit Entry")

                widget1 = QWidget(self.update_qc_widget)
                widget1.setGeometry(0, 0, 350, 500)

                form_layout = QFormLayout(widget1)

                font = QtGui.QFont("Arial", 11)

                lotnumber_label = QLabel()
                lotnumber_label.setText("Lot Number")
                lotnumber_label.setFixedHeight(30)
                lotnumber_label.setFixedWidth(100)
                lotnumber_label.setFont(font)
                lotnumber_label.setStyleSheet("background-color: rgb(240, 240, 240)")

                lotnumber_input = QLineEdit()
                lotnumber_input.setFixedHeight(30)
                lotnumber_input.setText(lot_number)
                lotnumber_input.setEnabled(False)

                customer_label = QLabel()
                customer_label.setText("Customer")
                customer_label.setFont(font)
                customer_label.setFixedWidth(100)

                customer_list = QComboBox()
                customer_list.setFixedHeight(30)

                self.cursor.execute("""
                            SELECT customers FROM customer
                            ORDER BY customers
                            """)
                customers = self.cursor.fetchall()

                for i in customers:
                    customer_list.addItem(i[0])
                customer_list.setCurrentText(customer)

                productCode_label = QLabel()
                productCode_label.setText("Product Code")
                productCode_label.setFont(font)
                productCode_label.setFixedWidth(100)

                productCode_input = QLineEdit()
                productCode_input.setFixedHeight(30)
                productCode_input.setText(product_code)
                productCode_input.setEnabled(False)

                formulaID_label = QLabel()
                formulaID_label.setText("Formula ID")
                formulaID_label.setFixedWidth(100)
                formulaID_label.setFont(font)

                formulaID_input = QLineEdit()
                formulaID_input.setFixedHeight(30)
                formulaID_input.setText(str(formula_id))

                date_started_label = QLabel()
                date_started_label.setText("Date Evaluated")
                date_started_label.setFont(font)

                date_started = QDateTimeEdit()
                date_started.setDisplayFormat("MM-dd-yyyy HH:mm")
                date_started.setFixedHeight(30)
                date_started.setDateTime(date_evaluated)

                date_endorsed_label = QLabel()
                date_endorsed_label.setText("Date Endorsed")
                date_endorsed_label.setFont(font)

                endorsed_date = QDateTimeEdit()
                endorsed_date.setDisplayFormat("MM-dd-yyyy HH:mm")
                endorsed_date.setFixedHeight(30)
                endorsed_date.setDateTime(endorsed_date_val)

                test_result_label = QLabel()
                test_result_label.setText("Test Result")
                test_result_label.setFont(font)
                test_result_label.setFixedWidth(100)

                test_result_dropdown = QComboBox()
                test_result_dropdown.addItem("Passed")
                test_result_dropdown.addItem("Failed")
                test_result_dropdown.setFixedHeight(30)
                test_result_dropdown.setCurrentText(status)

                remarks_label = QLabel()
                remarks_label.setText("Remarks")
                remarks_label.setFont(font)
                remarks_label.setFixedWidth(100)

                remarks_box = QTextEdit()
                remarks_box.setFixedHeight(120)
                remarks_box.setText(remarks)

                form_layout.addRow(lotnumber_label, lotnumber_input)
                form_layout.addRow(customer_label, customer_list)
                form_layout.addRow(productCode_label, productCode_input)
                form_layout.addRow(formulaID_label, formulaID_input)
                form_layout.addRow(date_started_label, date_started)
                form_layout.addRow(date_endorsed_label, endorsed_date)
                form_layout.addRow(test_result_label, test_result_dropdown)
                form_layout.addRow(remarks_label, remarks_box)

                save_button = QPushButton(self.update_qc_widget)
                save_button.setGeometry(100, 500, 60, 25)
                save_button.setText("Save")
                save_button.clicked.connect(save_update)
                save_button.show()

                cancel_button = QPushButton(self.update_qc_widget)
                cancel_button.setGeometry(200, 500, 60, 25)
                cancel_button.setText("Cancel")
                cancel_button.show()

                self.update_qc_widget.show()

            else:
                QMessageBox.information(self.qc_widget, "No Selected", "No Item Selected")

        def delete_entry():

            selected = self.qc_table.selectedItems()

            if selected:
                id = selected[0].text()
                lot_number = selected[1].text()

                try:
                    delete = QMessageBox.question(self.qc_widget, 'Delete Item',
                                                  f"Do you want to Delete \n Lot Number {lot_number}?",
                                                  QMessageBox.Yes | QMessageBox.No,
                                                  QMessageBox.No)
                    if delete == QMessageBox.Yes:
                        self.cursor.execute(f"""
                                                            DELETE FROM quality_control
                                                            WHERE id = {id}
                                                            """)

                        self.cursor.execute(f"""
                                                            DELETE FROM quality_control_tbl2
                                                            WHERE id = {id}

                                                            """)

                        self.cursor.execute(f"""
                                                                INSERT INTO qc_logs
                                                                VALUES('{datetime.now().strftime('%Y-%m-%d %H:%M')}', 'DELETE', {selected[0].text()}, '{selected[1].text()}', 
                                                                        '{selected[2].text()}', '{selected[3].text()}', '{selected[4].text()}', '{selected[5].text()}')

                                                            """)

                        self.conn.commit()
                        QMessageBox.information(self.qc_widget, '', f"{lot_number} successfully deleted.")

                    else:
                        return





                except:
                    pass


            else:
                QMessageBox.critical(self.qc_widget, '', "No Selected Item.")
                return

        try:
            if self.qc_entry_clicked_status == True:
                evaluation_entry()
                self.qc_entry_clicked_status = False
                return
            elif self.dashboard_clicked_status == True:
                show_dashboards()
                self.dashboard_clicked_status = False
                return
            elif self.returns_clicked_status == True:
                qc_returns()
                self.returns_clicked_status = False
                return
            else:
                pass
        except Exception as e:
            print(e)

        self.qc_widget = QtWidgets.QWidget(self.main_widget)
        self.qc_widget.setGeometry(0, 0, 991, 700)
        self.qc_widget.setStyleSheet("background-color: white;")
        self.qc_widget.show()

        self.qcBtn_topBorder = QtWidgets.QWidget(self.qc_widget)
        self.qcBtn_topBorder.setGeometry(0, 0, 991, 30)
        self.qcBtn_topBorder.setStyleSheet('background-color: rgb(92, 154, 255)')
        self.qcBtn_topBorder.show()

        self.qc_topBorder = QtWidgets.QWidget(self.qc_widget)
        self.qc_topBorder.setGeometry(0, 30, 991, 60)
        self.qc_topBorder.setStyleSheet("background-color: white")
        self.qc_topBorder.show()

        self.qc_table = QtWidgets.QTableWidget(self.qc_widget)
        self.qc_table.setGeometry(0, 90, 991, 350)
        self.qc_table.setColumnCount(7)

        # Set Column Width
        self.qc_table.setColumnWidth(0, 80)
        self.qc_table.setColumnWidth(1, 120)
        self.qc_table.setColumnWidth(2, 300)
        self.qc_table.setColumnWidth(3, 120)
        self.qc_table.setColumnWidth(4, 80)
        self.qc_table.setColumnWidth(5, 170)

        self.qc_table.verticalHeader().setVisible(False)
        self.qc_table.horizontalHeader().setStyleSheet("""
        QHeaderView::section{
        font-weight: bold;
        background-color: rgb(0, 109, 189);
        color: white;
        }

        """)

        self.qc_table.setHorizontalHeaderLabels(
            ["ID", "LOT NUMBER", "CUSTOMER", "PRODDUCT CODE", "STATUS", "REMARKS", "ACTION TAKEN"])

        self.qc_TableBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_TableBtn.setGeometry(55, 0, 70, 30)
        self.qc_TableBtn.setText("QC DATA")
        self.qc_TableBtn.setCursor(Qt.PointingHandCursor)
        self.qc_TableBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
        self.qc_TableBtn.setStyleSheet(
            "color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;")
        self.qc_TableBtn.clicked.connect(self.quality_control)
        self.qc_TableBtn.setShortcut("F1")

        # Change the Row height of the table
        for i in range(self.qc_table.rowCount()):
            self.qc_table.setRowHeight(i, 22)

        table_header = self.qc_table.horizontalHeader()
        table_header.setFixedHeight(25)

        # Get the table Items from database
        self.cursor.execute("""
        SELECT id, lot_number, customer, product_code, status, remarks, action
        FROM quality_control
        ORDER BY 
		CASE 
        WHEN updated_on IS NOT NULL THEN updated_on
        ELSE encoded_on
        END DESC, id DESC

        """)

        result = self.cursor.fetchall()

        self.qc_table.setRowCount(len(result))

        # Populate the table
        for i in range(len(result)):
            for j in range(len(result[i])):
                item = QTableWidgetItem(str(result[i][j]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                if result[i][4] == 'Failed':
                    item.setBackground(QtGui.QColor(255, 128, 128))

                self.qc_table.setItem(i, j, item)

        self.qc_table.itemSelectionChanged.connect(show_items)

        self.qc_TableBtn.show()

        self.qc_addEntryBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_addEntryBtn.setGeometry(145, 0, 100, 30)
        self.qc_addEntryBtn.setText("QC ENTRY")
        self.qc_addEntryBtn.setCursor(Qt.PointingHandCursor)
        self.qc_addEntryBtn.setStyleSheet("color: white; border: none;")
        self.qc_addEntryBtn.clicked.connect(evaluation_entry)
        self.qc_addEntryBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
        self.qc_addEntryBtn.setShortcut("F2")
        self.qc_addEntryBtn.show()

        self.qc_dataBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_dataBtn.setGeometry(270, 0, 70, 30)
        self.qc_dataBtn.setText("QC Data")
        self.qc_dataBtn.setCursor(Qt.PointingHandCursor)
        self.qc_dataBtn.setStyleSheet("color: white; border: none;")
        self.qc_dataBtn.clicked.connect(show_qc_data)
        self.qc_dataBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
        self.qc_dataBtn.setShortcut("F3")
        self.qc_dataBtn.show()

        self.dashboardBtn = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.dashboardBtn.setGeometry(370, 0, 85, 30)
        self.dashboardBtn.setText("DASHBOARD")
        self.dashboardBtn.setCursor(Qt.PointingHandCursor)
        self.dashboardBtn.setStyleSheet("color: white; border: none")
        self.dashboardBtn.clicked.connect(show_dashboards)
        self.dashboardBtn.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
        self.dashboardBtn.setShortcut("F4")
        self.dashboardBtn.show()

        self.qc_returns = QtWidgets.QPushButton(self.qcBtn_topBorder)
        self.qc_returns.setGeometry(500, 0, 70, 30)
        self.qc_returns.setText("RETURNS")
        self.qc_returns.setCursor(Qt.PointingHandCursor)
        self.qc_returns.setStyleSheet("color: white; border: none;")
        self.qc_returns.setFont(QtGui.QFont("Berlin Sans FB Demi", 10))
        self.qc_returns.clicked.connect(qc_returns)
        self.qc_returns.setShortcut("F5")
        self.qc_returns.show()

        # Top Border Widgets
        evaluation_lbl = QLabel(self.qc_topBorder)
        evaluation_lbl.setGeometry(10, 30, 100, 30)
        evaluation_lbl.setFont(QtGui.QFont("Arial", 9))
        evaluation_lbl.setText("Evaluation For:")
        evaluation_lbl.setStyleSheet("border: none;")
        evaluation_lbl.show()

        evaluation_text = QLabel(self.qc_topBorder)
        evaluation_text.setGeometry(120, 30, 500, 30)
        evaluation_text.setStyleSheet("border: none;")
        evaluation_text.show()

        loadAll_checkbox = QCheckBox(self.qc_topBorder)
        loadAll_checkbox.move(5, 5)
        loadAll_checkbox.setStyleSheet("border: none")
        loadAll_checkbox.show()

        loadAll_label = QLabel(self.qc_topBorder)
        loadAll_label.setGeometry(20, 4, 80, 15)
        loadAll_label.setStyleSheet("border: none;")
        loadAll_label.setText("LOAD ALL DATA")
        loadAll_label.setFont(QtGui.QFont("Arial", 8))
        loadAll_label.show()

        edited_checkbox = QCheckBox(self.qc_topBorder)
        edited_checkbox.move(105, 5)
        edited_checkbox.setStyleSheet("border: none;")
        edited_checkbox.stateChanged.connect(edited_checkbox_changed)
        edited_checkbox.show()

        edited_label = QLabel(self.qc_topBorder)
        edited_label.setGeometry(120, 4, 90, 15)
        edited_label.setStyleSheet("border: none;")
        edited_label.setText("EDITED RECORDS")
        edited_label.setFont(QtGui.QFont("Arial", 8))
        edited_label.show()

        search_bar = QtWidgets.QLineEdit(self.qc_topBorder)
        search_bar.setGeometry(770, 5, 150, 25)
        search_bar.setStyleSheet("border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 17);")
        search_bar.setFont(QtGui.QFont("Arial", 9))
        search_bar.setPlaceholderText("Lot Number")
        search_bar.setFocus()
        search_bar.textChanged.connect(auto_search)
        search_bar.show()

        search_btn = QtWidgets.QPushButton(self.qc_topBorder)
        search_btn.setGeometry(925, 5, 60, 25)
        search_btn.setStyleSheet("border: 1px solid rgb(171, 173, 179);")
        search_btn.setText("Search")
        search_btn.show()

        # Bottom Widgets
        bottom_widget = QtWidgets.QWidget(self.qc_widget)
        bottom_widget.setGeometry(0, 440, 991, 271)
        bottom_widget.setStyleSheet("background-color : rgb(239, 243, 254)")
        bottom_widget.show()

        leftSide1_widget = QtWidgets.QWidget(bottom_widget)
        leftSide1_widget.setGeometry(0, 0, 140, 200)
        leftSide1_widget.show()

        leftSide2_widget = QtWidgets.QWidget(bottom_widget)
        leftSide2_widget.setGeometry(140, 0, 455, 200)
        leftSide2_widget.show()

        rightSide_widget = QtWidgets.QWidget(bottom_widget)
        rightSide_widget.setGeometry(595, 0, 130, 100)
        rightSide_widget.show()

        label_font = QtGui.QFont("Segoe UI", 11)

        # Create Vertical Box layout
        leftVBox1_layout = QVBoxLayout(leftSide1_widget)
        leftVBox2_layout = QVBoxLayout(leftSide2_widget)
        rightVBo1_layout = QVBoxLayout(rightSide_widget)

        # Left Side Labels
        customer_label = QLabel()
        customer_label.setText("Customer            :")
        customer_label.setFont(label_font)

        productCode_label = QLabel()
        productCode_label.setText("Product Code     :")
        productCode_label.setFont(label_font)

        result_label = QLabel()
        result_label.setText("Result                 :")
        result_label.setFont(label_font)

        evaluatedBy_label = QLabel()
        evaluatedBy_label.setText("Evaluated By      :")
        evaluatedBy_label.setFont(label_font)

        evaluatedDate_label = QLabel()
        evaluatedDate_label.setText("Evaluated On     :")
        evaluatedDate_label.setFont(label_font)

        encodedDate_label = QLabel()
        encodedDate_label.setText("Encoded On       :")
        encodedDate_label.setFont(label_font)

        # Left Side Outputs
        customer_selected = QLabel()
        productCode_selected = QLabel()
        result_selected = QLabel()
        evaluatedBy_selected = QLabel()
        evaluatedDate_selected = QLabel()
        encodedDate_selected = QLabel()

        # Right Side Labels
        updatedBy_label = QLabel()
        updatedBy_label.setText("Updated By :")
        updatedBy_label.setAlignment(Qt.AlignRight)
        updatedBy_label.setFont(label_font)

        remarks_label = QLabel(bottom_widget)
        remarks_label.setText("Remarks :")
        remarks_label.setAlignment(Qt.AlignRight)
        remarks_label.setFont(label_font)
        remarks_label.setGeometry(612, 100, 100, 30)
        remarks_label.show()

        remarks_box = QTextEdit(bottom_widget)
        remarks_box.setGeometry(722, 100, 200, 80)
        remarks_box.setStyleSheet("background-color: white")
        remarks_box.setEnabled(False)
        remarks_box.show()

        actionTaken_label = QLabel(bottom_widget)
        actionTaken_label.setText("Action Taken :")
        actionTaken_label.setAlignment(Qt.AlignRight)
        actionTaken_label.setFont(label_font)
        actionTaken_label.setGeometry(612, 180, 100, 30)
        actionTaken_label.show()

        actionTake_box = QTextEdit(bottom_widget)
        actionTake_box.setGeometry(722, 180, 200, 60)
        actionTake_box.setStyleSheet("background-color: white")
        actionTake_box.setEnabled(False)
        actionTake_box.show()

        time_endorsed = QLabel()
        time_endorsed.setText("Time Endorsed :")
        time_endorsed.setAlignment(Qt.AlignRight)
        time_endorsed.setFont(label_font)

        time_endorsed_val = QLabel(self.qc_widget)
        time_endorsed_val.setGeometry(721, 478, 120, 25)
        time_endorsed_val.setFont(QtGui.QFont("Arial", 9))
        time_endorsed_val.setStyleSheet("background-color: rgb(239, 243, 254)")
        time_endorsed_val.show()

        qc_type_label = QLabel()
        qc_type_label.setText("QC Type : ")
        qc_type_label.setAlignment(Qt.AlignRight)
        qc_type_label.setFont(label_font)

        qc_type_selected = QLabel(self.qc_widget)
        qc_type_selected.setGeometry(721, 505, 120, 25)
        qc_type_selected.setStyleSheet("background-color: rgb(239, 243, 254)")
        qc_type_selected.show()

        # Adding the widgets to the Layout
        leftVBox1_layout.addWidget(customer_label)
        leftVBox1_layout.addWidget(productCode_label)
        leftVBox1_layout.addWidget(result_label)
        leftVBox1_layout.addWidget(evaluatedBy_label)
        leftVBox1_layout.addWidget(evaluatedDate_label)
        leftVBox1_layout.addWidget(encodedDate_label)

        leftVBox2_layout.addWidget(customer_selected)
        leftVBox2_layout.addWidget(productCode_selected)
        leftVBox2_layout.addWidget(result_selected)
        leftVBox2_layout.addWidget(evaluatedBy_selected)
        leftVBox2_layout.addWidget(evaluatedDate_selected)
        leftVBox2_layout.addWidget(encodedDate_selected)

        rightVBo1_layout.addWidget(updatedBy_label)
        rightVBo1_layout.addWidget(time_endorsed)
        rightVBo1_layout.addWidget(qc_type_label)

        updatedBy_val1 = QtWidgets.QLineEdit(self.qc_widget)
        updatedBy_val1.setGeometry(721, 445, 130, 25)
        updatedBy_val1.setStyleSheet("background-color: white")
        updatedBy_val1.setEnabled(False)
        updatedBy_val1.show()

        updatedBy_val2 = QtWidgets.QLineEdit(self.qc_widget)
        updatedBy_val2.setGeometry(851, 445, 130, 25)
        updatedBy_val2.setEnabled(False)
        updatedBy_val2.setStyleSheet("background-color: white")
        updatedBy_val2.show()

        export_label = QLabel(self.qc_widget)
        export_label.setGeometry(10, 630, 100, 20)
        export_label.setText("Export Database")
        export_label.setStyleSheet("background-color: rgb(239, 243, 254);")
        export_label.show()

        label1 = QLabel(self.qc_widget)
        label1.setText("FROM")
        label1.setGeometry(30, 650, 70, 25)
        label1.setStyleSheet("background-color: rgb(239, 243, 254)")
        label1.setFont(label_font)
        label1.show()

        date1 = QDateEdit(self.qc_widget)
        date1.setGeometry(75, 650, 100, 25)
        date1.setStyleSheet("background-color: rgb(239, 243, 254)")
        date1.setDisplayFormat("yyyy-MM-dd")
        date1.show()

        label2 = QLabel(self.qc_widget)
        label2.setGeometry(185, 650, 30, 25)
        label2.setText("TO")
        label2.setFont(label_font)
        label2.setStyleSheet("background-color: rgb(239, 243, 254)")
        label2.show()

        now = QtCore.QDate.currentDate()  # get the current date

        date2 = QDateEdit(self.qc_widget)
        date2.setGeometry(215, 650, 100, 25)
        date2.setStyleSheet("background-color: rgb(239, 243, 254)")
        date2.setDisplayFormat("yyyy-MM-dd")
        date2.setDate(now)
        date2.show()

        export_btn = ClickableLabel(self.qc_widget)
        export_btn.setGeometry(320, 653, 20, 20)
        export_btn.setPixmap(QtGui.QIcon('export.png').pixmap(20, 20))
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(exportBtn_clicked)
        export_btn.setToolTip("Export")
        export_btn.show()

        get_statistics_btn = ClickableLabel(self.qc_widget)
        get_statistics_btn.setGeometry(345, 653, 20, 20)
        get_statistics_btn.setPixmap(QtGui.QIcon('statistics-icon.png').pixmap(20, 20))
        get_statistics_btn.setCursor(Qt.PointingHandCursor)
        get_statistics_btn.clicked.connect(lambda: save_to_excel(date1.text(), date2.text()))
        get_statistics_btn.setToolTip("Export")
        get_statistics_btn.show()

        refresh_btn = QtWidgets.QPushButton(self.qc_widget)
        refresh_btn.setGeometry(530, 663, 60, 25)
        refresh_btn.setStyleSheet(
            'border: none; border-radius: 5px; background-color: rgb(194, 232, 255); border: 1px solid rgb(92, 154, 255)')
        refresh_btn.setText("REFRESH")
        refresh_btn.clicked.connect(self.quality_control)
        refresh_btn.show()

        update_btn = QtWidgets.QPushButton(self.qc_widget)
        update_btn.setGeometry(595, 663, 60, 25)
        update_btn.setStyleSheet(
            'border: none; border-radius: 5px; background-color: rgb(194, 232, 255); border: 1px solid rgb(92, 154, 255)')
        update_btn.setText("UPDATE")
        update_btn.clicked.connect(update_entry)
        update_btn.show()

        delete_btn = QtWidgets.QPushButton(self.qc_widget)
        delete_btn.setGeometry(660, 663, 60, 25)
        delete_btn.setStyleSheet(
            'border: none; border-radius: 5px; background-color: rgb(194, 232, 255); border: 1px solid rgb(92, 154, 255)')
        delete_btn.setText("DELETE")
        delete_btn.clicked.connect(delete_entry)
        delete_btn.show()

        self.qc_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.qc_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.qc_table.show()

    def warehouse(self):
        def fg_incoming():
            def add_entry():

                # Check if User have permission to enter QC data.
                if warehouse_access:
                    pass
                else:
                    QMessageBox.critical(self.warehouse_widget, 'Restricted Access',
                                         "You Dont Have Permission. \n Contact the Admin.")
                    return

                def autofill():
                    # Set to Capital Letter
                    lot_number_box.setText(lot_number_box.text().upper())

                    self.cursor.execute(f"""
                    SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor
                    FROM tbl_prod01
                    WHERE t_lotnum = '{lot_number_box.text()}' AND t_deleted = false

                    """)
                    result = self.cursor.fetchone()

                    if result:
                        product_code_box.setText(result[2])

                        if '-' in product_code_box.text():
                            category_box.setCurrentText('DRYCOLOR')
                        else:
                            category_box.setCurrentText('MASTERBATCH')
                        return

                    else:
                        pass

                    try:
                        # for lot numbers that have cuts
                        if '-' in lot_number_box.text() and lot_number_box.text()[-1] != ')':
                            num1 = re.findall(r'\d{4}', lot_number_box.text().strip())[0]
                            num2 = re.findall(r'\d{4}', lot_number_box.text().strip())[1]
                            code_type = re.findall(r'[a-zA-Z]+', lot_number_box.text().strip())[0]

                            self.cursor.execute(f"""
                                                    SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor
                                                    FROM
                                                    (SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor, LEFT(t_lotnum, 4):: INT as range1, 
                                                    CASE 
                                                    WHEN LENGTH(t_lotnum) = 13 THEN SUBSTRING(t_lotnum FROM 8 FOR 4)::INTEGER
                                                    WHEN LENGTH(t_lotnum) = 11 THEN SUBSTRING(t_lotnum FROM 7 FOR 4)::INTEGER
                                                    ELSE NULL
                                                    END AS range2
                                                    FROM tbl_prod01
                                                    WHERE t_lotnum LIKE '%-%' 
                                                    AND t_deleted = false
                                                    AND t_lotnum ~* '\d{code_type}$'
                                                    ORDER BY t_prodid::INTEGER DESC)
                                                    WHERE {num1} >= range1 AND {num2} <= range2

                                                                """)
                            result = self.cursor.fetchone()

                            if result:
                                product_code_box.setText(result[2])

                                if '-' in product_code_box.text():
                                    category_box.setCurrentText('DRYCOLOR')
                                else:
                                    category_box.setCurrentText('MASTERBATCH')
                                quantity_box.setFocus()

                            else:
                                pass
                    except IndexError:
                        QMessageBox.information(self.widget, 'Data Not Found', 'Data Does Not Exist')

                    else:
                        try:
                            num1 = re.findall(r'\d{4}', lot_number_box.text().strip())[0]
                            code_type = re.findall(r'[a-zA-Z]+', lot_number_box.text().strip())[0]

                            self.cursor.execute(f"""
                                SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor
                                FROM
                                (SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor, LEFT(t_lotnum, 4):: INT as range1, 
                                CASE 
                                WHEN LENGTH(t_lotnum) = 13 THEN SUBSTRING(t_lotnum FROM 8 FOR 4)::INTEGER
                                WHEN LENGTH(t_lotnum) = 11 THEN SUBSTRING(t_lotnum FROM 7 FOR 4)::INTEGER
                                ELSE NULL
                                END AS range2
                                FROM tbl_prod01
                                WHERE t_lotnum LIKE '%-%' 
                                AND t_deleted = false
                                AND t_lotnum ~* '\d{code_type}$'
                                ORDER BY t_prodid::INTEGER DESC)
                                WHERE {num1} >= range1 AND {num1} <= range2

                                                """)
                            result = self.cursor.fetchone()

                            if result:
                                product_code_box.setText(result[2])
                            else:
                                pass

                        except:
                            QMessageBox.critical(self.widget, 'ERROR', 'LOT NUMBER not Found!')

                def save_entry():

                    if input_type_box.currentText() == 'MULTIPLE':

                        num1 = int(re.findall(r'(\d+)[A-Z]+', lot_number_box.text())[0])
                        num2 = int(re.findall(r'(\d+)[A-Z]+', lot_number_box.text())[1])
                        code = re.findall(r'[A-Z]+', lot_number_box.text())[0]
                        product_code = product_code_box.text()

                        for i in range(num1, num2 + 1):

                            while len(str(i)) != 4:
                                i = '0' + str(i)

                            # Check if data is already in the database.
                            self.cursor.execute(f"""
                                SELECT * FROM fg_incoming
                                WHERE lot_number = '{(str(i) + code).upper()}' AND quantity = {float(quantity_box.text()) / ((num2 - num1) + 1)}
                                                    AND location = '{warehouse_input.currentText() + ":" + block_input.text()}'
                                    """)

                            result = self.cursor.fetchone()

                            if result:
                                QMessageBox.information(self.widget, 'Data Exist',
                                                        f'{(str(i) + code).upper()} Already Exist.')
                                return
                            else:
                                pass

                            self.cursor.execute(f"""
                                INSERT INTO fg_incoming(product_code, production_date, lot_number, quantity, category, remarks, location)
                                VALUES('{product_code_box.text()}', TO_DATE('{production_date_box.text()}', 'MM-DD-YYYY'),
                                '{(str(i) + code).upper()}', {float(quantity_box.text()) / ((num2 - num1) + 1)}, '{category_box.currentText()}',
                                '{remarks_box.currentText()}', '{warehouse_input.currentText() + ":" + block_input.text()}')
                                                    """)

                            # add to inventory

                            self.cursor.execute(f"""
                            INSERT INTO fg_inventory
                            VALUES('{(str(i) + code).upper()}', '{product_code_box.text()}', {float(quantity_box.text()) / ((num2 - num1) + 1)}, '{date.today().strftime('%Y-%m-%d')}',
                            '{warehouse_input.currentText() + ":" + block_input.text()}', '{category_box.currentText()}')

                            """)

                        self.conn.commit()

                        QMessageBox.information(self.widget, 'Entry Success', 'Data Successfully Entered.')
                        # Clear the Output after saving
                        lot_number_val.clear()
                        product_code_box.clear()
                        quantity_box.clear()

                    else:
                        num1 = int(re.findall(r'(\d+)[A-Z]+', lot_number_box.text())[0])
                        code = re.findall(r'[A-Z]+', lot_number_box.text())[0]
                        product_code = product_code_box.text()

                        # Check if data is already in the database.
                        self.cursor.execute(f"""
                            SELECT * FROM fg_incoming
                            WHERE lot_number = '{(str(num1) + code).upper()}' AND quantity = {float(quantity_box.text())}
                                AND location = '{warehouse_input.currentText() + ":" + block_input.text()}'
                                                            """)

                        result = self.cursor.fetchone()

                        if result:
                            QMessageBox.information(self.widget, 'Data Exist',
                                                    f'{(str(num1) + code).upper()} Already Exist.')
                            return
                        else:
                            pass

                        self.cursor.execute(f"""
                            INSERT INTO fg_incoming(product_code, production_date, lot_number, quantity, category, remarks, location)
                            VALUES('{product_code_box.text()}', TO_DATE('{production_date_box.text()}', 'MM-DD-YYYY'),
                            '{lot_number_box.text()}', {quantity_box.text()}, '{category_box.currentText()}',
                            '{remarks_box.currentText()}', '{warehouse_input.currentText() + ":" + block_input.text()}')


                                            """)

                        self.cursor.execute(f"""
                            INSERT INTO fg_inventory
                            VALUES('{lot_number_box.text()}', '{product_code_box.text()}', {quantity_box.text()}, '{date.today().strftime('%Y-%m-%d')}',
                            '{warehouse_input.currentText() + ":" + block_input.text()}', '{category_box.currentText()}')


                        """)

                        self.conn.commit()

                        QMessageBox.information(self.widget, 'Entry Success', 'Data Successfully Entered.')
                        # Clear the Output after saving
                        lot_number_val.clear()
                        product_code_box.clear()
                        quantity_box.clear()

                    self.cursor.execute("""
                                        SELECT control_id, lot_number, production_date, product_code, quantity, category, remarks, location   
                                        FROM fg_incoming
                                        WHERE deleted = false
                                        ORDER BY control_id DESC

                                        """)

                    result = self.cursor.fetchall()

                    # adjust the length
                    table_widget.setRowCount(len(result))

                    table_widget.clearContents()
                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            table_widget.setItem(result.index(row), column, item)

                    # Clear after saving
                    lot_number_box.clear()
                    product_code_box.clear()
                    quantity_box.clear()
                    block_input.clear()

                self.widget = QWidget()
                self.widget.setGeometry(780, 305, 400, 500)
                self.widget.setFixedSize(400, 500)
                self.widget.setWindowModality(Qt.ApplicationModal)
                self.widget.setWindowTitle("ADD FG INCOMING")
                self.widget.setWindowIcon(QtGui.QIcon('delivery_icon.png'))
                self.widget.show()

                form_layout_widget = QWidget(self.widget)
                form_layout_widget.setGeometry(0, 0, 400, 400)
                form_layout_widget.show()

                label_font = QtGui.QFont("Arial", 12)
                input_font = QtGui.QFont("Arial", 11)

                input_type_label = QLabel()
                input_type_label.setFont(label_font)
                input_type_label.setText("Input Type")
                input_type_label.setFixedWidth(150)
                input_type_label.setFixedHeight(35)

                input_type_box = QComboBox()
                input_type_box.addItem('SINGLE')
                input_type_box.addItem('MULTIPLE')
                input_type_box.setFixedHeight(30)
                input_type_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                input_type_box.setFont(input_font)

                id_number_label = QLabel()
                id_number_label.setFont(label_font)
                id_number_label.setText("Control ID")
                id_number_label.setFixedWidth(150)
                id_number_label.setFixedHeight(35)

                id_number_box = QLineEdit()
                id_number_box.setFixedHeight(35)
                id_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                id_number_box.setEnabled(False)
                id_number_box.setFont(input_font)

                production_date_label = QLabel()
                production_date_label.setFont(label_font)
                production_date_label.setText('PROD DATE')
                production_date_label.setFixedWidth(150)

                production_date_box = QDateEdit()
                production_date_box.setFixedHeight(30)
                production_date_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                production_date_box.setDisplayFormat('MM-dd-yyyy')
                production_date_box.setDate(QtCore.QDate.currentDate())
                production_date_box.setFont(input_font)

                category_label = QLabel()
                category_label.setFont(label_font)
                category_label.setText('Category')
                category_label.setFixedWidth(150)

                category_box = QComboBox()
                category_box.addItem('MASTERBATCH')
                category_box.addItem('DRYCOLOR')
                category_box.setFixedHeight(30)
                category_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                category_box.setFont(input_font)

                lot_number_label = QLabel()
                lot_number_label.setText('LOT Number')
                lot_number_label.setFixedWidth(150)
                lot_number_label.setFont(label_font)

                lot_number_box = QLineEdit()
                lot_number_box.setFixedHeight(30)
                lot_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                lot_number_box.editingFinished.connect(autofill)
                lot_number_box.setFont(input_font)

                quantity_label = QLabel()
                quantity_label.setText('Quantity')
                quantity_label.setFont(label_font)
                quantity_label.setFixedWidth(150)

                quantity_box = QLineEdit()
                quantity_box.setFixedHeight(30)
                quantity_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                quantity_box.setFont(input_font)
                validator = QtGui.QDoubleValidator()
                validator.setNotation(QtGui.QDoubleValidator.StandardNotation)

                quantity_box.setValidator(validator)

                product_code_label = QLabel()
                product_code_label.setText('Product Code')
                product_code_label.setFixedWidth(150)
                product_code_label.setFont(label_font)

                product_code_box = QLineEdit()
                product_code_box.setFixedHeight(30)
                product_code_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                product_code_box.setFont(input_font)

                warehouse_label = QLabel()
                warehouse_label.setText('Warehouse')
                warehouse_label.setFixedWidth(150)
                warehouse_label.setFont(label_font)

                warehouse_input = QComboBox()
                warehouse_input.setFixedHeight(30)
                warehouse_input.setStyleSheet('background-color: rgb(255, 255, 17)')
                warehouse_input.addItem("WH 1")
                warehouse_input.addItem("WH 2")
                warehouse_input.addItem("WH 4")
                warehouse_input.addItem("WH 5")
                warehouse_input.setFont(input_font)

                block_label = QLabel()
                block_label.setText('Box No.')
                block_label.setFixedWidth(150)
                block_label.setFont(label_font)

                block_input = QLineEdit()
                block_input.setFixedHeight(30)
                block_input.setStyleSheet('background-color: rgb(255, 255, 17)')
                block_input.setFont(input_font)

                remarks_label = QLabel()
                remarks_label.setText('Status')
                remarks_label.setFixedWidth(150)
                remarks_label.setFont(label_font)

                comment_label = QLabel()
                comment_label.setText('Remarks')
                comment_label.setFixedWidth(150)
                comment_label.setFont(label_font)

                comment_box = QTextEdit()
                comment_box.setFixedHeight(70)
                comment_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                comment_box.setFont(input_font)

                remarks_box = QComboBox()
                remarks_box.addItem('-')
                remarks_box.addItem("NEW PASSED")
                remarks_box.addItem("NEW FAILED")
                remarks_box.addItem("RETURN PASS")
                remarks_box.addItem("RETURN FAIL")
                remarks_box.setFixedHeight(30)
                remarks_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                remarks_box.setFont(input_font)

                layout = QFormLayout(form_layout_widget)
                layout.addRow(input_type_label, input_type_box)
                layout.addRow(id_number_label, id_number_box)
                layout.addRow(lot_number_label, lot_number_box)
                layout.addRow(production_date_label, production_date_box)
                layout.addRow(product_code_label, product_code_box)
                layout.addRow(category_label, category_box)
                layout.addRow(quantity_label, quantity_box)
                layout.addRow(warehouse_label, warehouse_input)
                layout.addRow(block_label, block_input)
                layout.addRow(remarks_label, remarks_box)

                layout.setVerticalSpacing(20)

                lot_number_box.setFocus()

                save_btn = QPushButton(self.widget)
                save_btn.setGeometry(100, 450, 60, 25)
                save_btn.setText('SAVE')
                save_btn.setStyleSheet(
                    'background-color: rgb(194, 232, 255); border-radius: 5px; border: 1px solid rgb(92, 154, 255)')
                save_btn.clicked.connect(save_entry)
                save_btn.setShortcut('Return')
                save_btn.show()

                cancel_btn = QPushButton(self.widget)
                cancel_btn.setGeometry(230, 450, 60, 25)
                cancel_btn.setText('CANCEL')
                cancel_btn.setStyleSheet(
                    'background-color: rgb(194, 232, 255); border-radius: 5px; border: 1px solid rgb(92, 154, 255)')
                cancel_btn.clicked.connect(lambda: self.widget.close())
                cancel_btn.show()

            def update_entry():

                # Check if User have permission to enter QC data.
                if warehouse_access:
                    pass
                else:
                    QMessageBox.critical(self.warehouse_widget, 'Restricted Access',
                                         "You Dont Have Permission. \n Contact the Admin.")
                    return

                selected = table_widget.selectedItems()

                if selected:
                    selected = [i.text() for i in selected]

                    def update_btn_clicked():

                        try:
                            self.cursor.execute(f"""
                                UPDATE fg_incoming
                                SET  product_code = '{product_code_box.text()}', 
                                production_date = TO_DATE('{production_date_box.text()}', 'MM-DD-YYYY'), lot_number = '{lot_number_box.text().upper()}',
                                quantity = {quantity_box.text()}, category = '{category_box.currentText()}', 
                                remarks = '{remarks_box.currentText()}', location = '{warehouse_input.currentText() + ':' + block_input.text()}'
                                WHERE control_id = {selected[0]}

                                                """)

                            self.conn.commit()
                            QMessageBox.information(self.widget, 'SUCCESS', 'Update Successful!')
                            self.widget.close()
                            self.warehouse()

                        except Exception as e:
                            self.conn.rollback()
                            QMessageBox.critical(self.widget, 'ERROR', str(e))

                    self.widget = QWidget()
                    self.widget.setGeometry(780, 305, 400, 500)
                    self.widget.setFixedSize(400, 500)
                    self.widget.setWindowTitle('UPDATE ENTRY')
                    self.widget.setWindowModality(Qt.ApplicationModal)
                    self.widget.setWindowIcon(QtGui.QIcon('delivery_icon.png'))
                    self.widget.show()

                    form_layout_widget = QWidget(self.widget)
                    form_layout_widget.setGeometry(0, 0, 400, 400)
                    form_layout_widget.show()

                    label_font = QtGui.QFont("Arial", 12)
                    input_font = QtGui.QFont("Arial", 11)

                    id_number_label = QLabel()
                    id_number_label.setFont(label_font)
                    id_number_label.setText("Control ID")
                    id_number_label.setFixedWidth(150)
                    id_number_label.setFixedHeight(35)

                    id_number_box = QLineEdit()
                    id_number_box.setFixedHeight(35)
                    id_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    id_number_box.setText(selected[0])
                    id_number_box.setEnabled(False)
                    id_number_box.setFont(input_font)

                    production_date_label = QLabel()
                    production_date_label.setFont(label_font)
                    production_date_label.setText('PROD DATE')
                    production_date_label.setFixedWidth(150)

                    production_date_box = QDateEdit()
                    production_date_box.setFixedHeight(30)
                    production_date_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    production_date_box.setDisplayFormat('MM-dd-yyyy')
                    production_date_box.setFont(input_font)

                    my_date = re.findall(r'\d+', selected[2])
                    print(my_date)
                    year = int(my_date[2])
                    month = int(my_date[0])
                    day = int(my_date[1])

                    d = QtCore.QDate(year, month, day)
                    production_date_box.setDate(d)

                    warehouse_label = QLabel()
                    warehouse_label.setText('Warehouse')
                    warehouse_label.setFixedWidth(150)
                    warehouse_label.setFont(label_font)

                    warehouse_input = QComboBox()
                    warehouse_input.setFixedHeight(30)
                    warehouse_input.setStyleSheet('background-color: rgb(255, 255, 17)')
                    warehouse_input.addItem("WH 1")
                    warehouse_input.addItem("WH 2")
                    warehouse_input.addItem("WH 4")
                    warehouse_input.addItem("WH 5")
                    warehouse_input.setCurrentText(selected[7].split(":")[0])
                    warehouse_input.setFont(input_font)

                    block_label = QLabel()
                    block_label.setText('Box No.')
                    block_label.setFixedWidth(150)
                    block_label.setFont(label_font)

                    block_input = QLineEdit()
                    block_input.setFixedHeight(30)
                    block_input.setStyleSheet('background-color: rgb(255, 255, 17)')
                    block_input.setFont(input_font)

                    remarks_label = QLabel()
                    remarks_label.setText('Status')
                    remarks_label.setFixedWidth(150)
                    remarks_label.setFont(label_font)

                    remarks_box = QComboBox()
                    remarks_box.addItem('-')
                    remarks_box.addItem("NEW PASSED")
                    remarks_box.addItem("NEW FAILED")
                    remarks_box.addItem("RETURN PASS")
                    remarks_box.addItem("RETURN FAIL")
                    remarks_box.setFixedHeight(30)
                    remarks_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    remarks_box.setCurrentText(selected[6])
                    remarks_box.setFont(input_font)

                    category_label = QLabel()
                    category_label.setFont(label_font)
                    category_label.setText('Category')
                    category_label.setFixedWidth(150)

                    category_box = QComboBox()
                    category_box.addItem('MASTERBATCH')
                    category_box.addItem('DRYCOLOR')
                    category_box.setFixedHeight(30)
                    category_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    category_box.setCurrentText(selected[5])
                    category_box.setFont(input_font)

                    lot_number_label = QLabel()
                    lot_number_label.setText('LOT Number')
                    lot_number_label.setFixedWidth(150)
                    lot_number_label.setFont(label_font)

                    lot_number_box = QLineEdit()
                    lot_number_box.setFixedHeight(30)
                    lot_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    lot_number_box.setText(selected[1])
                    lot_number_box.setFont(input_font)

                    quantity_label = QLabel()
                    quantity_label.setText('Quantity')
                    quantity_label.setFont(label_font)
                    quantity_label.setFixedWidth(150)

                    quantity_box = QLineEdit()
                    quantity_box.setFixedHeight(30)
                    quantity_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    quantity_box.setText(selected[4])
                    validator = QtGui.QDoubleValidator()
                    validator.setNotation(QtGui.QDoubleValidator.StandardNotation)

                    quantity_box.setValidator(validator)

                    quantity_box.setFont(input_font)

                    product_code_label = QLabel()
                    product_code_label.setText('Product Code')
                    product_code_label.setFixedWidth(150)
                    product_code_label.setFont(label_font)

                    product_code_box = QLineEdit()
                    product_code_box.setFixedHeight(30)
                    product_code_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    product_code_box.setText(selected[3])
                    product_code_box.setFont(input_font)

                    layout = QFormLayout(form_layout_widget)
                    layout.addRow(id_number_label, id_number_box)
                    layout.addRow(lot_number_label, lot_number_box)
                    layout.addRow(production_date_label, production_date_box)
                    layout.addRow(product_code_label, product_code_box)
                    layout.addRow(category_label, category_box)
                    layout.addRow(quantity_label, quantity_box)
                    layout.addRow(warehouse_label, warehouse_input)
                    layout.addRow(block_label, block_input)
                    layout.addRow(remarks_label, remarks_box)

                    layout.setVerticalSpacing(20)

                    lot_number_box.setFocus()

                    update_btn = QPushButton(self.widget)
                    update_btn.setGeometry(100, 450, 60, 30)
                    update_btn.setText('UPDATE')
                    update_btn.clicked.connect(update_btn_clicked)
                    update_btn.show()

                    cancel_btn = QPushButton(self.widget)
                    cancel_btn.setGeometry(230, 450, 60, 30)
                    cancel_btn.setText('CANCEL')
                    cancel_btn.clicked.connect(lambda: self.widget.close())
                    cancel_btn.show()

                else:
                    QMessageBox.critical(self.warehouse_widget, 'ERROR', 'No Selected Item!')

            def show_selected():

                try:
                    selected = [i.text() for i in table_widget.selectedItems()]

                    control_num_val.setText(selected[0])
                    lot_number_val.setText(selected[1])
                    code_value.setText(selected[3])
                    category_value.setText(selected[5])
                except IndexError:
                    QMessageBox.information(self.warehouse_widget, 'ERROR', 'No Item Selected.')

            def fg_filter():

                if masterbatch_checkbox.isChecked() == True and drycolor_checkbox.isChecked() == False:

                    self.cursor.execute("""
                    SELECT control_id, lot_number, production_date, product_code,  quantity, category, remarks, location  
                    FROM fg_incoming
                    WHERE deleted = false AND category = 'MASTERBATCH'
                    ORDER BY control_id DESC

                    """)

                    result = self.cursor.fetchall()

                    table_widget.clear()
                    table_widget.setHorizontalHeaderLabels(
                        ["Control ID", "Lot No.", "Date", "Product Code", "Quantity", "Category", 'Remarks',
                         'Location'])

                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            table_widget.setItem(result.index(row), column, item)


                elif masterbatch_checkbox.isChecked() == False and drycolor_checkbox.isChecked() == True:
                    self.cursor.execute("""
                        SELECT control_id, lot_number, production_date, product_code, quantity, category   
                        FROM fg_incoming
                        WHERE deleted = false AND category = 'DRYCOLOR'
                        ORDER BY control_id DESC

                                    """)

                    result = self.cursor.fetchall()

                    table_widget.clear()
                    table_widget.setHorizontalHeaderLabels(
                        ["Control ID", "Lot No.", "Date", "Customer", "Product Code", "Color", "Quantity", "Category"])

                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            table_widget.setItem(result.index(row), column, item)

                else:
                    self.cursor.execute("""
                                                    SELECT control_id, lot_number, production_date,  product_code,  quantity, category, remarks, location  
                                                    FROM fg_incoming
                                                    WHERE deleted = false 
                                                    ORDER BY control_id DESC

                                                    """)

                    result = self.cursor.fetchall()

                    table_widget.clear()
                    table_widget.setHorizontalHeaderLabels(
                        ["Control ID", "Lot No.", "Date", "Product Code", "Quantity", "Category", "Remarks",
                         "Location"])

                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            table_widget.setItem(result.index(row), column, item)

            def search():

                lot_num = search_box.text()

                self.cursor.execute(f"""
                SELECT control_id, lot_number, production_date, product_code, quantity, category, remarks, location
                FROM fg_incoming 
                WHERE lot_number ILIKE '%{lot_num}%' AND deleted = false

                """)

                result = self.cursor.fetchall()

                table_widget.clear()
                table_widget.setHorizontalHeaderLabels(
                    ["Control ID", "Lot No.", "Date", "Product Code", "Quantity", "Category", "Remarks", "Location"])

                for row in result:
                    for column in range(len(row)):
                        item = QTableWidgetItem(str(row[column]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        table_widget.setItem(result.index(row), column, item)

            def delete_incoming():

                # Check if User have permission to enter QC data.
                if warehouse_access:
                    pass
                else:
                    QMessageBox.critical(self.warehouse_widget, 'Restricted Access',
                                         "You Dont Have Permission. \n Contact the Admin.")
                    return

                selected = table_widget.selectedItems()

                if selected:
                    self.cursor.execute(f"""
                    DELETE FROM fg_incoming
                    WHERE control_id = {selected[0].text()}

                    """)

                    self.conn.commit()
                    QMessageBox.information(self.warehouse_widget, 'Deleted', 'Data successfuly deleted!')
                    self.warehouse()  # Refreshes the Table.

                else:
                    QMessageBox.information(self.warehouse_widget, 'ERROR', 'No Selected Item!')

            def export_to_excel():

                self.cursor.execute("""
                SELECT * FROM fg_incoming
                ORDER BY control_id

                """)
                result = self.cursor.fetchall()

                df = pd.DataFrame(result)

                column_names = ['Control ID', 'Product Code', 'Date', 'Lot Number', 'Quantity', 'category', 'deleted',
                                'Color']

                try:
                    filename, _ = QtWidgets.QFileDialog.getSaveFileName(self.warehouse_widget, "Save File",
                                                                        r"C:\Users\Administrator\Documents",
                                                                        'Excel Files (*.xlsx)',
                                                                        options=QtWidgets.QFileDialog.Options())

                    if filename:
                        # Ensuring the file name ends with .xlsx
                        if not filename.lower().endswith('.xlsx'):
                            filename += '.xlsx'

                        # Print or use the file name
                        df.to_excel(excel_writer=filename, index=False,
                                    header=column_names)
                        QMessageBox.information(self.production_widget, "File Imported", "Successfully Imported Data")
                except PermissionError:
                    QMessageBox.critical(self.production_widget, "Permission Error", "Unable to Export the File. \n "
                                                                                     "Someone is using blank.xlsx")

            def auto_search():
                item = search_box.text()

                self.cursor.execute(f"""
                                SELECT control_id, lot_number, production_date, product_code, quantity, category, remarks, location
                                FROM fg_incoming 
                                WHERE lot_number ILIKE '%{item}%' OR product_code ILIKE '%{item}%'
                                ORDER BY control_id DESC
                                """)

                result = self.cursor.fetchall()

                table_widget.clear()
                table_widget.setHorizontalHeaderLabels(
                    ["Control ID", "Lot No.", "Date", "Product Code", "Quantity", "Category", "Remarks", "Location"])

                table_widget.verticalScrollBar().setValue(0)

                for row in result:
                    for column in range(len(row)):
                        item = QTableWidgetItem(str(row[column]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        table_widget.setItem(result.index(row), column, item)

            def daily_report():

                self.cursor.execute(f"""
                SELECT product_code, SUM(quantity) as total_incoming
                FROM fg_incoming
                WHERE deleted = false AND date_encoded = '{date.today().strftime('%Y-%m-%d')}'
                GROUP BY product_code
                """)
                result = self.cursor.fetchall()
                print(result)

            fg_incoming_btn = QPushButton(self.warehouse_tabs)
            fg_incoming_btn.setGeometry(30, 0, 100, 30)
            fg_incoming_btn.setText("FG INCOMING")
            fg_incoming_btn.setCursor(Qt.PointingHandCursor)
            fg_incoming_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_incoming_btn.setStyleSheet(
                "color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;")
            fg_incoming_btn.clicked.connect(self.warehouse)
            fg_incoming_btn.show()

            fg_outgoing_btn = QPushButton(self.warehouse_tabs)
            fg_outgoing_btn.setGeometry(150, 0, 100, 30)
            fg_outgoing_btn.setText("FG OUTGOING")
            fg_outgoing_btn.setCursor(Qt.PointingHandCursor)
            fg_outgoing_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_outgoing_btn.setStyleSheet("color: white; border: none")
            fg_outgoing_btn.clicked.connect(fg_outgoing)
            fg_outgoing_btn.show()

            fg_inventory_btn = QPushButton(self.warehouse_tabs)
            fg_inventory_btn.setGeometry(270, 0, 100, 30)
            fg_inventory_btn.setText("FG INVENTORY")
            fg_inventory_btn.setCursor(Qt.PointingHandCursor)
            fg_inventory_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_inventory_btn.setStyleSheet("color: white; border: none")
            fg_inventory_btn.clicked.connect(fg_inventory)
            fg_inventory_btn.show()

            self.status_border = QWidget(self.warehouse_widget)
            self.status_border.setGeometry(0, 30, 991, 35)
            self.status_border.setStyleSheet('border-bottom: 1px solid rgb(160, 160, 160)')
            self.status_border.show()

            search_box = QLineEdit(self.warehouse_widget)
            search_box.setGeometry(760, 70, 150, 25)
            search_box.setStyleSheet('border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 17);')
            search_box.setPlaceholderText('Lot Number')
            search_box.setFocus()
            search_box.textChanged.connect(auto_search)
            search_box.show()

            search_button = QPushButton(self.warehouse_widget)
            search_button.setGeometry(915, 70, 70, 25)
            search_button.setStyleSheet('border: 1px solid rgb(171, 173, 179)')
            search_button.setText('Search')
            search_button.clicked.connect(search)
            search_button.setShortcut('Return')
            search_button.show()

            masterbatch_checkbox = QCheckBox(self.warehouse_widget)
            masterbatch_checkbox.move(5, 70)
            masterbatch_checkbox.stateChanged.connect(fg_filter)
            masterbatch_checkbox.show()

            mb_checkbox_label = QLabel(self.warehouse_widget)
            mb_checkbox_label.setGeometry(22, 70, 85, 10)
            mb_checkbox_label.setText('MASTERBATCH')
            mb_checkbox_label.show()

            drycolor_checkbox = QCheckBox(self.warehouse_widget)
            drycolor_checkbox.move(110, 70)
            drycolor_checkbox.stateChanged.connect(fg_filter)
            drycolor_checkbox.show()

            dc_checkbox_label = QLabel(self.warehouse_widget)
            dc_checkbox_label.setGeometry(125, 70, 85, 10)
            dc_checkbox_label.setText("DRYCOLOR")
            dc_checkbox_label.show()

            title_label = QLabel(self.warehouse_widget)
            title_label.setGeometry(290, 80, 300, 40)
            title_label.setStyleSheet('color : rgb(41, 181, 255)')
            title_label.setFont(QtGui.QFont('Segoe UI Black', 20))
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setText("FG Incoming")
            title_label.show()

            title_icon = QLabel(self.warehouse_widget)
            title_icon.setGeometry(550, 80, 40, 40)
            title_icon.setPixmap(QtGui.QIcon('delivery_icon.png').pixmap(40, 40))
            title_icon.show()

            control_num_lbl = QLabel(self.status_border)
            control_num_lbl.setGeometry(3, 12, 100, 10)
            control_num_lbl.setText('Control Number:')
            control_num_lbl.setStyleSheet('border: none')
            control_num_lbl.show()

            control_num_val = QLabel(self.status_border)
            control_num_val.setGeometry(105, 0, 120, 35)
            control_num_val.setFont(QtGui.QFont('Arial Black', 15))
            control_num_val.setStyleSheet('color: rgb(0, 128, 192)')
            control_num_val.show()

            code_label = QLabel(self.status_border)
            code_label.setGeometry(260, 0, 40, 35)
            code_label.setFont(QtGui.QFont('Arial', 8))
            code_label.setText('CODE:')
            code_label.show()

            code_value = QLabel(self.status_border)
            code_value.setGeometry(300, 0, 150, 35)
            code_value.setFont(QtGui.QFont('Arial Black', 15))
            code_value.setStyleSheet('color: rgb(0, 128, 192)')
            code_value.show()

            lot_number_label = QLabel(self.status_border)
            lot_number_label.setGeometry(490, 0, 50, 35)
            lot_number_label.setText('LOT NO :')
            lot_number_label.setFont(QtGui.QFont('Arial', 8))
            lot_number_label.show()

            lot_number_val = QLabel(self.status_border)
            lot_number_val.setGeometry(540, 0, 180, 35)
            lot_number_val.setStyleSheet('color: rgb(0, 128, 192)')
            lot_number_val.setFont(QtGui.QFont('Arial Black', 15))
            lot_number_val.show()

            category_label = QLabel(self.status_border)
            category_label.setGeometry(735, 0, 50, 35)
            category_label.setFont(QtGui.QFont('Arial', 8))
            category_label.setText('Category:')
            category_label.show()

            category_value = QLabel(self.status_border)
            category_value.setGeometry(790, 0, 170, 35)
            category_value.setFont(QtGui.QFont('Arial Black', 15))
            category_value.setStyleSheet('color: rgb(0, 128, 192)')
            category_value.show()

            table_widget = QTableWidget(self.warehouse_widget)
            table_widget.setGeometry(0, 125, 991, 506)
            table_widget.setColumnCount(9)
            table_widget.verticalHeader().setVisible(False)
            table_widget.setHorizontalHeaderLabels(
                ["Control ID", "Lot No.", "Date", "Product Code", "Quantity", "Category", "Status", "Location",
                 "Remarks"])

            table_widget.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
            table_widget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            table_widget.itemSelectionChanged.connect(show_selected)

            self.cursor.execute("""
                    SELECT control_id, lot_number, TO_CHAR(production_date, 'MM-DD-YYYY'), product_code, quantity,
                           category, remarks, location, comments 
                    FROM fg_incoming
                    WHERE deleted = false
                    ORDER BY control_id DESC

                    """)

            table_widget.setColumnWidth(0, 90)
            table_widget.setColumnWidth(1, 120)
            table_widget.setColumnWidth(2, 100)
            table_widget.setColumnWidth(3, 100)
            table_widget.setColumnWidth(4, 100)
            table_widget.setColumnWidth(5, 100)
            table_widget.setColumnWidth(6, 150)
            table_widget.setColumnWidth(7, 220)
            table_widget.setColumnWidth(8, 200)

            table_widget.setFont(QtGui.QFont('Arial', 10))

            table_widget.horizontalHeader().setStyleSheet("""
                        QHeaderView::section{
                            font-weight: bold;
                            background-color: rgb(41, 181, 255);
                            color: black;
                            font-size: 18;
                        }
                                    """)

            result = self.cursor.fetchall()

            table_widget.setRowCount(len(result))
            print(result)

            for i in range(len(result)):
                for j in range(len(result[i])):
                    item = QTableWidgetItem(str(result[i][j]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                    if result[i][6] == 'NEW FAILED' or result[i][6] == 'RETURN FAIL':
                        item.setBackground(QtGui.QColor(255, 128, 128))

                    table_widget.setItem(i, j, item)

            table_widget.show()

            bottom_button_widget = QWidget(self.warehouse_widget)
            bottom_button_widget.setGeometry(0, 631, 991, 43)
            bottom_button_widget.show()

            date1 = QDateEdit(bottom_button_widget)
            date1.setGeometry(80, 18, 100, 25)
            date1.show()

            date2 = QDateEdit(bottom_button_widget)
            date2.setGeometry(200, 18, 100, 25)
            date2.show()

            export_logo = ClickableLabel(bottom_button_widget)
            export_logo.setGeometry(310, 18, 20, 20)
            export_logo.setPixmap(QtGui.QIcon('export.png').pixmap(20, 20))
            export_logo.setStyleSheet('border: none')
            export_logo.setCursor(Qt.PointingHandCursor)
            export_logo.clicked.connect(export_to_excel)
            export_logo.show()

            report_logo = ClickableLabel(bottom_button_widget)
            report_logo.setGeometry(335, 18, 20, 20)
            report_logo.setPixmap(QtGui.QIcon('daily_report.png').pixmap(20, 20))
            report_logo.setStyleSheet('border: none')
            report_logo.setCursor(Qt.PointingHandCursor)
            report_logo.clicked.connect(daily_report)
            report_logo.show()

            # Buttons
            add_btn = QPushButton(bottom_button_widget)
            add_btn.setGeometry(650, 18, 60, 25)
            add_btn.setText('ADD')
            add_btn.setStyleSheet("border-radius: 5px;"
                                  "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            add_btn.setCursor(Qt.PointingHandCursor)
            add_btn.clicked.connect(add_entry)
            add_btn.show()

            update_btn = QPushButton(bottom_button_widget)
            update_btn.setGeometry(715, 18, 60, 25)
            update_btn.setText('UPDATE')
            update_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                     "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            update_btn.setCursor(Qt.PointingHandCursor)
            update_btn.clicked.connect(update_entry)
            update_btn.show()

            refresh_btn = QPushButton(bottom_button_widget)
            refresh_btn.setGeometry(780, 18, 60, 25)
            refresh_btn.setText('REFRESH')
            refresh_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                      "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            refresh_btn.setCursor(Qt.PointingHandCursor)
            refresh_btn.clicked.connect(lambda: self.warehouse())
            refresh_btn.show()

            delete_btn = QPushButton(bottom_button_widget)
            delete_btn.setGeometry(845, 18, 60, 25)
            delete_btn.setText('DELETE')
            delete_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                     "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            delete_btn.setCursor(Qt.PointingHandCursor)
            delete_btn.clicked.connect(delete_incoming)
            delete_btn.show()

        def fg_outgoing():

            def add_entry():

                # Check if User have permission to enter QC data.
                if warehouse_access:
                    pass
                else:
                    QMessageBox.critical(self.warehouse_widget, 'Restricted Access',
                                         "You Dont Have Permission. \n Contact the Admin.")
                    return

                def autofill():
                    # Set to Capital Letter
                    lot_number_box.setText(lot_number_box.text().upper())

                    self.cursor.execute(f"""
                    SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor
                    FROM tbl_prod01
                    WHERE t_lotnum = '{lot_number_box.text()}' AND t_deleted = false

                    """)
                    result = self.cursor.fetchone()

                    if result:
                        product_color_box.setText(result[3])
                        product_code_box.setText(result[2])

                        if '-' in product_code_box.text():
                            category_box.setCurrentText('DRYCOLOR')
                        else:
                            category_box.setCurrentText('MASTERBATCH')
                        return

                    else:
                        pass

                    # for lot numbers that have cuts
                    if '-' in lot_number_box.text() and lot_number_box.text()[-1] != ')':
                        num1 = re.findall(r'\d{4}', lot_number_box.text().strip())[0]
                        num2 = re.findall(r'\d{4}', lot_number_box.text().strip())[1]
                        code_type = re.findall(r'[a-zA-Z]+', lot_number_box.text().strip())[0]

                        self.cursor.execute(f"""
                            SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor
                            FROM
                            (SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor, LEFT(t_lotnum, 4):: INT as range1, 
                            CASE 
                            WHEN LENGTH(t_lotnum) = 13 THEN SUBSTRING(t_lotnum FROM 8 FOR 4)::INTEGER
                            WHEN LENGTH(t_lotnum) = 11 THEN SUBSTRING(t_lotnum FROM 7 FOR 4)::INTEGER
                            ELSE NULL
                            END AS range2
                            FROM tbl_prod01
                            WHERE t_lotnum LIKE '%-%' 
                            AND t_deleted = false
                            AND t_lotnum ~* '\d{code_type}$'
                            ORDER BY t_prodid::INTEGER DESC)
                            WHERE {num1} >= range1 AND {num2} <= range2

                                        """)
                        result = self.cursor.fetchone()

                        if result:
                            product_color_box.setText(result[3])
                            product_code_box.setText(result[2])

                            if '-' in product_code_box.text():
                                category_box.setCurrentText('DRYCOLOR')
                            else:
                                category_box.setCurrentText('MASTERBATCH')
                            quantity_box.setFocus()

                        else:
                            pass

                    else:
                        try:
                            num1 = re.findall(r'\d{4}', lot_number_box.text().strip())[0]
                            code_type = re.findall(r'[a-zA-Z]+', lot_number_box.text().strip())[0]

                            self.cursor.execute(f"""
                                SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor
                                FROM
                                (SELECT t_lotnum, t_customer, t_prodcode, t_prodcolor, LEFT(t_lotnum, 4):: INT as range1, 
                                CASE 
                                WHEN LENGTH(t_lotnum) = 13 THEN SUBSTRING(t_lotnum FROM 8 FOR 4)::INTEGER
                                WHEN LENGTH(t_lotnum) = 11 THEN SUBSTRING(t_lotnum FROM 7 FOR 4)::INTEGER
                                ELSE NULL
                                END AS range2
                                FROM tbl_prod01
                                WHERE t_lotnum LIKE '%-%' 
                                AND t_deleted = false
                                AND t_lotnum ~* '\d{code_type}$'
                                ORDER BY t_prodid::INTEGER DESC)
                                WHERE {num1} >= range1 AND {num1} <= range2

                                                """)
                            result = self.cursor.fetchone()

                            if result:
                                product_color_box.setText(result[3])
                                product_code_box.setText(result[2])
                            else:
                                pass

                        except:
                            QMessageBox.critical(self.widget, 'ERROR', 'LOT NUMBER not Found!')

                def save_entry():

                    if input_type_box.currentText() == 'MULTIPLE':

                        num1 = int(re.findall(r'(\d+)[A-Z]{2}', lot_number_box.text())[0])
                        num2 = int(re.findall(r'(\d+)[A-Z]{2}', lot_number_box.text())[1])
                        code = re.findall(r'[A-Z]+', lot_number_box.text())[0]
                        product_code = product_code_box.text()

                        for i in range(num1, num2 + 1):
                            self.cursor.execute(f"""
                                                INSERT INTO fg_outgoing(product_code, date, lot_number, quantity, category)
                                                VALUES('{product_code_box.text()}',  TO_DATE('{production_date_box.text()}', 'MM-DD-YYYY'),
                                                '{str(i) + code}', {float(quantity_box.text()) / ((num2 - num1) + 1)}, '{category_box.currentText()}')
                                                    """)
                        self.conn.commit()

                        lot_number_val.clear()
                        product_code_box.clear()
                        quantity_box.clear()

                        QMessageBox.information(self.widget, 'Entry Success', 'Data Successfully Entered.')
                    else:
                        num1 = int(re.findall(r'(\d+)[A-Z]+', lot_number_box.text())[0])
                        code = re.findall(r'[A-Z]+', lot_number_box.text())[0]
                        product_code = product_code_box.text()

                        self.cursor.execute(f"""
                                            INSERT INTO fg_outgoing(product_code, date, lot_number, quantity, category)
                                            VALUES('{product_code_box.text()}', TO_DATE('{production_date_box.text()}', 'MM-DD-YYYY'),
                                            '{lot_number_box.text()}', {quantity_box.text()}, '{category_box.currentText()}')

                                            """)

                        self.cursor.execute(f"""
                        UPDATE fg_inventory
                        SET quantity = quantity - {quantity_box.text()}
                        WHERE lot_number = '{lot_number_box.text()}' AND EXISTS (SELECT 1 FROM fg_inventory
                        WHERE lot_number = '{lot_number_box.text()}')


                        """)

                        self.conn.commit()

                        lot_number_val.clear()
                        product_code_box.clear()
                        quantity_box.clear()

                        QMessageBox.information(self.widget, 'Entry Success', 'Data Successfully Entered.')

                    self.cursor.execute("""
                        SELECT control_id, lot_number, date, product_code, quantity, category 
                        FROM fg_outgoing
                        ORDER BY control_id DESC

                                                            """)

                    result = self.cursor.fetchall()

                    # adjust the length
                    outgoing_table.setRowCount(len(result))

                    outgoing_table.clearContents()
                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            outgoing_table.setItem(result.index(row), column, item)

                    # Clear after saving
                    lot_number_box.clear()
                    product_code_box.clear()
                    quantity_box.clear()

                self.widget = QWidget()
                self.widget.setGeometry(780, 305, 400, 500)
                self.widget.setFixedSize(400, 500)
                self.widget.setWindowModality(Qt.ApplicationModal)
                self.widget.setWindowTitle("ADD FG INCOMING")
                self.widget.setWindowIcon(QtGui.QIcon('delivery_icon.png'))
                self.widget.show()

                form_layout_widget = QWidget(self.widget)
                form_layout_widget.setGeometry(0, 0, 400, 400)
                form_layout_widget.show()

                label_font = QtGui.QFont("Arial", 11)

                input_type_label = QLabel()
                input_type_label.setFont(label_font)
                input_type_label.setText("Input Type")
                input_type_label.setFixedWidth(150)
                input_type_label.setFixedHeight(35)

                input_type_box = QComboBox()
                input_type_box.addItem('SINGLE')
                input_type_box.addItem('MULTIPLE')
                input_type_box.setFixedHeight(30)
                input_type_box.setStyleSheet('background-color: rgb(255, 255, 17)')

                id_number_label = QLabel()
                id_number_label.setFont(label_font)
                id_number_label.setText("Control ID")
                id_number_label.setFixedWidth(150)
                id_number_label.setFixedHeight(35)

                id_number_box = QLineEdit()
                id_number_box.setFixedHeight(35)
                id_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                id_number_box.setEnabled(False)

                production_date_label = QLabel()
                production_date_label.setFont(label_font)
                production_date_label.setText('PROD DATE')
                production_date_label.setFixedWidth(150)

                production_date_box = QDateEdit()
                production_date_box.setFixedHeight(30)
                production_date_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                production_date_box.setDisplayFormat('MM-dd-yyyy')
                production_date_box.setDate(date.today())

                category_label = QLabel()
                category_label.setFont(label_font)
                category_label.setText('Category')
                category_label.setFixedWidth(150)

                category_box = QComboBox()
                category_box.addItem('MASTERBATCH')
                category_box.addItem('DRYCOLOR')
                category_box.setFixedHeight(30)
                category_box.setStyleSheet('background-color: rgb(255, 255, 17)')

                lot_number_label = QLabel()
                lot_number_label.setText('LOT Number')
                lot_number_label.setFixedWidth(150)
                lot_number_label.setFont(label_font)

                lot_number_box = QLineEdit()
                lot_number_box.setFixedHeight(30)
                lot_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                lot_number_box.editingFinished.connect(autofill)

                quantity_label = QLabel()
                quantity_label.setText('Quantity')
                quantity_label.setFont(label_font)
                quantity_label.setFixedWidth(150)

                quantity_box = QLineEdit()
                quantity_box.setFixedHeight(30)
                quantity_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                validator = QtGui.QDoubleValidator()
                validator.setNotation(QtGui.QDoubleValidator.StandardNotation)

                quantity_box.setValidator(validator)

                product_code_label = QLabel()
                product_code_label.setText('Product Code')
                product_code_label.setFixedWidth(150)
                product_code_label.setFont(label_font)

                product_code_box = QLineEdit()
                product_code_box.setFixedHeight(30)
                product_code_box.setStyleSheet('background-color: rgb(255, 255, 17)')

                product_color_label = QLabel()
                product_color_label.setText('Color')
                product_color_label.setFont(label_font)
                product_color_label.setFixedWidth(150)

                product_color_box = QLineEdit()
                product_color_box.setFixedHeight(30)
                product_color_box.setStyleSheet('background-color: rgb(255, 255, 17)')

                layout = QFormLayout(form_layout_widget)
                layout.addRow(input_type_label, input_type_box)
                layout.addRow(id_number_label, id_number_box)
                layout.addRow(lot_number_label, lot_number_box)
                layout.addRow(production_date_label, production_date_box)
                layout.addRow(product_code_label, product_code_box)
                layout.addRow(category_label, category_box)
                layout.addRow(quantity_label, quantity_box)

                layout.setVerticalSpacing(20)

                lot_number_box.setFocus()

                save_btn = QPushButton(self.widget)
                save_btn.setGeometry(100, 450, 60, 30)
                save_btn.setText('SAVE')
                save_btn.clicked.connect(save_entry)
                save_btn.setShortcut('Return')
                save_btn.show()

                cancel_btn = QPushButton(self.widget)
                cancel_btn.setGeometry(230, 450, 60, 30)
                cancel_btn.setText('CANCEL')
                cancel_btn.clicked.connect(lambda: self.widget.close())
                cancel_btn.show()

            def update_entry():

                # Check if User have permission to enter QC data.
                if warehouse_access:
                    pass
                else:
                    QMessageBox.critical(self.warehouse_widget, 'Restricted Access',
                                         "You Dont Have Permission. \n Contact the Admin.")
                    return

                selected = outgoing_table.selectedItems()

                if selected:
                    selected = [i.text() for i in selected]

                    def update_btn_clicked():

                        try:
                            self.cursor.execute(f"""
                                UPDATE fg_outgoing
                                SET product_code = '{product_code_box.text()}', 
                                date = TO_DATE('{production_date_box.text()}', 'MM-DD-YYYY'), lot_number = '{lot_number_box.text()}',
                                quantity = {quantity_box.text()}, category = '{category_box.currentText()}'
                                WHERE control_id = {selected[0]}


                                                """)

                            self.conn.commit()
                            QMessageBox.information(self.widget, 'SUCCESS', 'Update Successful!')
                            self.widget.close()
                            fg_outgoing()

                        except Exception as e:
                            self.conn.rollback()
                            QMessageBox.critical(self.widget, 'ERROR', str(e))

                    self.widget = QWidget()
                    self.widget.setGeometry(780, 305, 400, 500)
                    self.widget.setFixedSize(400, 500)
                    self.widget.setWindowTitle('UPDATE ENTRY')
                    self.widget.setWindowModality(Qt.ApplicationModal)
                    self.widget.setWindowIcon(QtGui.QIcon('delivery_icon.png'))
                    self.widget.show()

                    form_layout_widget = QWidget(self.widget)
                    form_layout_widget.setGeometry(0, 0, 400, 400)
                    form_layout_widget.show()

                    label_font = QtGui.QFont("Arial", 11)

                    id_number_label = QLabel()
                    id_number_label.setFont(label_font)
                    id_number_label.setText("Control ID")
                    id_number_label.setFixedWidth(150)
                    id_number_label.setFixedHeight(35)

                    id_number_box = QLineEdit()
                    id_number_box.setFixedHeight(35)
                    id_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    id_number_box.setText(selected[0])
                    id_number_box.setEnabled(False)

                    production_date_label = QLabel()
                    production_date_label.setFont(label_font)
                    production_date_label.setText('PROD DATE')
                    production_date_label.setFixedWidth(150)

                    production_date_box = QDateEdit()
                    production_date_box.setFixedHeight(30)
                    production_date_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    production_date_box.setDisplayFormat('MM-dd-yyyy')

                    my_date = re.findall(r'\d+', selected[2])
                    year = int(my_date[0])
                    month = int(my_date[1])
                    day = int(my_date[2])

                    d = QtCore.QDate(year, month, day)
                    production_date_box.setDate(d)

                    category_label = QLabel()
                    category_label.setFont(label_font)
                    category_label.setText('Category')
                    category_label.setFixedWidth(150)

                    category_box = QComboBox()
                    category_box.addItem('MASTERBATCH')
                    category_box.addItem('DRYCOLOR')
                    category_box.setFixedHeight(30)
                    category_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    category_box.setCurrentText(selected[5])

                    lot_number_label = QLabel()
                    lot_number_label.setText('LOT Number')
                    lot_number_label.setFixedWidth(150)
                    lot_number_label.setFont(label_font)

                    lot_number_box = QLineEdit()
                    lot_number_box.setFixedHeight(30)
                    lot_number_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    lot_number_box.setText(selected[1])

                    quantity_label = QLabel()
                    quantity_label.setText('Quantity')
                    quantity_label.setFont(label_font)
                    quantity_label.setFixedWidth(150)

                    quantity_box = QLineEdit()
                    quantity_box.setFixedHeight(30)
                    quantity_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    validator = QtGui.QDoubleValidator()
                    validator.setNotation(QtGui.QDoubleValidator.StandardNotation)

                    quantity_box.setValidator(validator)
                    quantity_box.setText(selected[4])

                    product_code_label = QLabel()
                    product_code_label.setText('Product Code')
                    product_code_label.setFixedWidth(150)
                    product_code_label.setFont(label_font)

                    product_code_box = QLineEdit()
                    product_code_box.setFixedHeight(30)
                    product_code_box.setStyleSheet('background-color: rgb(255, 255, 17)')
                    product_code_box.setText(selected[3])

                    product_color_label = QLabel()
                    product_color_label.setText('Color')
                    product_color_label.setFont(label_font)
                    product_color_label.setFixedWidth(150)

                    layout = QFormLayout(form_layout_widget)
                    layout.addRow(id_number_label, id_number_box)
                    layout.addRow(lot_number_label, lot_number_box)
                    layout.addRow(production_date_label, production_date_box)
                    layout.addRow(product_code_label, product_code_box)
                    layout.addRow(category_label, category_box)
                    layout.addRow(quantity_label, quantity_box)

                    layout.setVerticalSpacing(20)

                    lot_number_box.setFocus()

                    update_btn = QPushButton(self.widget)
                    update_btn.setGeometry(100, 450, 60, 30)
                    update_btn.setText('UPDATE')
                    update_btn.clicked.connect(update_btn_clicked)
                    update_btn.show()

                    cancel_btn = QPushButton(self.widget)
                    cancel_btn.setGeometry(230, 450, 60, 30)
                    cancel_btn.setText('CANCEL')
                    cancel_btn.clicked.connect(lambda: self.widget.close())
                    cancel_btn.show()

                else:
                    QMessageBox.critical(self.warehouse_widget, 'ERROR', 'No Selected Item!')

            def show_selected():

                try:
                    selected = [i.text() for i in outgoing_table.selectedItems()]

                    control_num_val.setText(selected[0])
                    lot_number_val.setText(selected[1])
                    code_value.setText(selected[3])
                    category_value.setText(selected[5])
                except IndexError:
                    QMessageBox.information(self.warehouse_widget, 'ERROR', 'No Selected!')

            def fg_filter():

                if masterbatch_checkbox.isChecked() == True and drycolor_checkbox.isChecked() == False:

                    self.cursor.execute("""
                        SELECT control_id, lot_number, date, customer, product_code, color, quantity, category   
                        FROM fg_outgoing
                        WHERE deleted = false AND category = 'MASTERBATCH'
                        ORDER BY control_id DESC

                    """)

                    result = self.cursor.fetchall()

                    outgoing_table.clear()
                    outgoing_table.setHorizontalHeaderLabels(
                        ["Control ID", "Lot No.", "Date", "Customer", "Product Code", "Color", "Quantity", "Category"])

                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            outgoing_table.setItem(result.index(row), column, item)


                elif masterbatch_checkbox.isChecked() == False and drycolor_checkbox.isChecked() == True:
                    self.cursor.execute("""
                        SELECT control_id, lot_number, date, customer, product_code, color, quantity, category   
                        FROM fg_outgoing
                        WHERE deleted = false AND category = 'DRYCOLOR'
                        ORDER BY control_id DESC

                                    """)

                    result = self.cursor.fetchall()

                    outgoing_table.clear()
                    outgoing_table.setHorizontalHeaderLabels(
                        ["Control ID", "Lot No.", "Date", "Customer", "Product Code", "Color", "Quantity", "Category"])

                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            outgoing_table.setItem(result.index(row), column, item)

                else:
                    self.cursor.execute("""
                        SELECT control_id, lot_number, date, customer, product_code, color, quantity, category   
                        FROM fg_outgoing
                        WHERE deleted = false 
                        ORDER BY control_id DESC

                                        """)

                    result = self.cursor.fetchall()

                    outgoing_table.clear()
                    outgoing_table.setHorizontalHeaderLabels(
                        ["Control ID", "Lot No.", "Date", "Customer", "Product Code", "Color", "Quantity", "Category"])

                    for row in result:
                        for column in range(len(row)):
                            item = QTableWidgetItem(str(row[column]))
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignCenter)
                            outgoing_table.setItem(result.index(row), column, item)

            def search():

                lot_num = search_box.text()

                self.cursor.execute(f"""
                    SELECT control_id, lot_number, date, customer, product_code, color, quantity, category
                    FROM fg_outgoing 
                    WHERE lot_number ILIKE '%{lot_num}%' AND deleted = false

                """)

                result = self.cursor.fetchall()

                outgoing_table.clear()
                outgoing_table.setHorizontalHeaderLabels(
                    ["Control ID", "Lot No.", "Date", "Customer", "Product Code", "Color", "Quantity", "Category"])

                for row in result:
                    for column in range(len(row)):
                        item = QTableWidgetItem(str(row[column]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        outgoing_table.setItem(result.index(row), column, item)

            def delete_outgoing():

                # Check if User have permission to enter QC data.
                if warehouse_access:
                    pass
                else:
                    QMessageBox.critical(self.warehouse_widget, 'Restricted Access',
                                         "You Dont Have Permission. \n Contact the Admin.")
                    return

                selected = outgoing_table.selectedItems()

                if selected:
                    self.cursor.execute(f"""
                        UPDATE fg_outgoing
                        SET deleted = true
                        WHERE control_id = {selected[0].text()}

                    """)

                    self.conn.commit()
                    QMessageBox.information(self.warehouse_widget, 'Deleted', 'Data successfuly deleted!')
                    self.warehouse()  # Refreshes the Table.

                else:
                    QMessageBox.information(self.warehouse_widget, 'ERROR', 'No Seleceted Item!')

            def export_to_excel():

                self.cursor.execute("""
                SELECT * FROM fg_outgoing
                ORDER BY control_id

                        """)
                result = self.cursor.fetchall()

                df = pd.DataFrame(result)

                column_names = ['Control ID', 'Customer', 'Product Code', 'Date', 'Lot Number', 'Quantity', 'category',
                                'deleted', 'Color']

                try:
                    filename, _ = QtWidgets.QFileDialog.getSaveFileName(self.warehouse_widget, "Save File",
                                                                        r"C:\Users\Administrator\Documents",
                                                                        'Excel Files (*.xlsx)',
                                                                        options=QtWidgets.QFileDialog.Options())

                    if filename:
                        # Ensuring the file name ends with .xlsx
                        if not filename.lower().endswith('.xlsx'):
                            filename += '.xlsx'

                        # Print or use the file name
                        df.to_excel(excel_writer=filename, index=False,
                                    header=column_names)
                        QMessageBox.information(self.production_widget, "File Imported", "Successfully Imported Data")
                except PermissionError:
                    QMessageBox.critical(self.production_widget, "Permission Error", "Unable to Export the File. \n "
                                                                                     "Someone is using blank.xlsx")

            def auto_search():

                lot_num = search_box.text()

                self.cursor.execute(f"""
                    SELECT control_id, lot_number, date, product_code, quantity, category   
                    FROM fg_outgoing 
                    WHERE lot_number ILIKE '%{lot_num}%' OR product_code ILIKE '%{lot_num}%'
                    ORDER BY control_id DESC

                        """)

                result = self.cursor.fetchall()

                outgoing_table.clear()
                outgoing_table.setHorizontalHeaderLabels(
                    ["Control ID", "Lot No.", "Date", "Customer", "Product Code", "Color", "Quantity", "Category"])

                outgoing_table.verticalScrollBar().setValue(0)

                for row in result:
                    for column in range(len(row)):
                        item = QTableWidgetItem(str(row[column]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        outgoing_table.setItem(result.index(row), column, item)

            fg_incoming_btn = QPushButton(self.warehouse_tabs)
            fg_incoming_btn.setGeometry(30, 0, 100, 30)
            fg_incoming_btn.setText("FG INCOMING")
            fg_incoming_btn.setCursor(Qt.PointingHandCursor)
            fg_incoming_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_incoming_btn.setStyleSheet(
                "color: white; border: none; ")
            fg_incoming_btn.clicked.connect(self.warehouse)
            fg_incoming_btn.show()

            fg_outgoing_btn = QPushButton(self.warehouse_tabs)
            fg_outgoing_btn.setGeometry(150, 0, 100, 30)
            fg_outgoing_btn.setText("FG OUTGOING")
            fg_outgoing_btn.setCursor(Qt.PointingHandCursor)
            fg_outgoing_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_outgoing_btn.setStyleSheet(
                "color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;")
            fg_outgoing_btn.clicked.connect(fg_outgoing)
            fg_outgoing_btn.show()

            fg_inventory_btn = QPushButton(self.warehouse_tabs)
            fg_inventory_btn.setGeometry(270, 0, 100, 30)
            fg_inventory_btn.setText("FG INVENTORY")
            fg_inventory_btn.setCursor(Qt.PointingHandCursor)
            fg_inventory_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_inventory_btn.setStyleSheet("color: white; border: none")
            fg_inventory_btn.clicked.connect(fg_inventory)
            fg_inventory_btn.show()

            fg_incoming_widget = QWidget(self.warehouse_widget)
            fg_incoming_widget.setGeometry(0, 30, 991, 721)
            fg_incoming_widget.show()

            status_border = QWidget(fg_incoming_widget)
            status_border.setGeometry(0, 0, 991, 35)
            status_border.setStyleSheet('border-bottom: 1px solid rgb(160, 160, 160)')
            status_border.show()

            search_box = QLineEdit(fg_incoming_widget)
            search_box.setGeometry(760, 40, 150, 25)
            search_box.setStyleSheet('border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 17);')
            search_box.setPlaceholderText('Lot Number')
            search_box.setFocus()
            search_box.textChanged.connect(auto_search)
            search_box.show()

            search_button = QPushButton(fg_incoming_widget)
            search_button.setGeometry(915, 40, 70, 25)
            search_button.setStyleSheet('border: 1px solid rgb(171, 173, 179)')
            search_button.setText('Search')
            search_button.clicked.connect(search)
            search_button.setShortcut('Return')
            search_button.show()

            masterbatch_checkbox = QCheckBox(fg_incoming_widget)
            masterbatch_checkbox.move(5, 40)
            masterbatch_checkbox.stateChanged.connect(fg_filter)
            masterbatch_checkbox.show()

            mb_checkbox_label = QLabel(fg_incoming_widget)
            mb_checkbox_label.setGeometry(22, 40, 85, 10)
            mb_checkbox_label.setText('MASTERBATCH')
            mb_checkbox_label.show()

            drycolor_checkbox = QCheckBox(fg_incoming_widget)
            drycolor_checkbox.move(110, 40)
            drycolor_checkbox.stateChanged.connect(fg_filter)
            drycolor_checkbox.show()

            dc_checkbox_label = QLabel(fg_incoming_widget)
            dc_checkbox_label.setGeometry(125, 40, 85, 10)
            dc_checkbox_label.setText("DRYCOLOR")
            dc_checkbox_label.show()

            title_label = QLabel(fg_incoming_widget)
            title_label.setGeometry(290, 50, 300, 40)
            title_label.setStyleSheet('color : red')
            title_label.setFont(QtGui.QFont('Segoe UI Black', 20))
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setText("FG OUTGOING")
            title_label.show()

            title_icon = QLabel(fg_incoming_widget)
            title_icon.setGeometry(550, 50, 40, 40)
            title_icon.setPixmap(QtGui.QIcon('delivery_icon_outgoing.png').pixmap(40, 40))
            title_icon.show()

            control_num_lbl = QLabel(fg_incoming_widget)
            control_num_lbl.setGeometry(3, 12, 100, 10)
            control_num_lbl.setText('Control Number:')
            control_num_lbl.show()

            control_num_val = QLabel(fg_incoming_widget)
            control_num_val.setGeometry(105, 0, 120, 35)
            control_num_val.setFont(QtGui.QFont('Arial Black', 15))
            control_num_val.setStyleSheet('color: rgb(0, 128, 192); border-bottom: 1px solid rgb(160, 160, 160)')
            control_num_val.show()

            code_label = QLabel(fg_incoming_widget)
            code_label.setGeometry(260, 0, 40, 35)
            code_label.setFont(QtGui.QFont('Arial', 8))
            code_label.setStyleSheet('border-bottom: 1px solid rgb(160, 160, 160)')
            code_label.setText('CODE:')
            code_label.show()

            code_value = QLabel(fg_incoming_widget)
            code_value.setGeometry(300, 0, 150, 35)
            code_value.setFont(QtGui.QFont('Arial Black', 15))
            code_value.setStyleSheet('color: rgb(0, 128, 192); border-bottom: 1px solid rgb(160, 160, 160)')
            code_value.show()

            lot_number_label = QLabel(fg_incoming_widget)
            lot_number_label.setGeometry(490, 0, 50, 35)
            lot_number_label.setText('LOT NO :')
            lot_number_label.setFont(QtGui.QFont('Arial', 8))
            lot_number_label.setStyleSheet('border-bottom: 1px solid rgb(160, 160, 160)')
            lot_number_label.show()

            lot_number_val = QLabel(fg_incoming_widget)
            lot_number_val.setGeometry(540, 0, 180, 35)
            lot_number_val.setStyleSheet('color: rgb(0, 128, 192); border-bottom: 1px solid rgb(160, 160, 160)')
            lot_number_val.setFont(QtGui.QFont('Arial Black', 15))
            lot_number_val.show()

            category_label = QLabel(fg_incoming_widget)
            category_label.setGeometry(735, 0, 50, 35)
            category_label.setFont(QtGui.QFont('Arial', 8))
            category_label.setText('Category:')
            category_label.setStyleSheet('border-bottom: 1px solid rgb(160, 160, 160)')
            category_label.show()

            category_value = QLabel(fg_incoming_widget)
            category_value.setGeometry(790, 0, 170, 35)
            category_value.setFont(QtGui.QFont('Arial Black', 15))
            category_value.setStyleSheet('color: rgb(0, 128, 192); border-bottom: 1px solid rgb(160, 160, 160)')
            category_value.show()

            outgoing_table = QTableWidget(fg_incoming_widget)
            outgoing_table.setGeometry(0, 95, 991, 506)
            outgoing_table.setColumnCount(6)
            outgoing_table.verticalHeader().setVisible(False)
            outgoing_table.setHorizontalHeaderLabels(
                ["Control ID", "Lot No.", "Date", "Product Code", "Quantity", "Category"])

            outgoing_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
            outgoing_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            outgoing_table.itemSelectionChanged.connect(show_selected)

            outgoing_table.setColumnWidth(0, 90)
            outgoing_table.setColumnWidth(1, 120)
            outgoing_table.setColumnWidth(2, 100)
            outgoing_table.setColumnWidth(3, 270)
            outgoing_table.setColumnWidth(4, 100)
            outgoing_table.setColumnWidth(5, 100)
            outgoing_table.setColumnWidth(6, 90)

            outgoing_table.horizontalHeader().setStyleSheet("""
                        QHeaderView::section{
                            font-weight: bold;
                            background-color: rgb(255, 0, 0);
                            color: black;
                            font-size: 18;
                        }
                                    """)

            self.cursor.execute("""
                                SELECT control_id, lot_number, date, product_code, quantity, category   
                                FROM fg_outgoing
                                WHERE deleted = false
                                ORDER BY control_id DESC

                                """)

            result = self.cursor.fetchall()
            outgoing_table.setRowCount(len(result))

            for row in result:
                for column in range(len(row)):
                    item = QTableWidgetItem(str(row[column]))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    item.setTextAlignment(Qt.AlignCenter)
                    outgoing_table.setItem(result.index(row), column, item)

            outgoing_table.show()

            bottom_button_widget = QWidget(self.warehouse_widget)
            bottom_button_widget.setGeometry(0, 631, 991, 43)
            bottom_button_widget.show()

            date1 = QDateEdit(bottom_button_widget)
            date1.setGeometry(80, 18, 100, 25)
            date1.show()

            date2 = QDateEdit(bottom_button_widget)
            date2.setGeometry(200, 18, 100, 25)
            date2.show()

            export_logo = ClickableLabel(bottom_button_widget)
            export_logo.setGeometry(310, 18, 20, 20)
            export_logo.setPixmap(QtGui.QIcon('export.png').pixmap(20, 20))
            export_logo.setStyleSheet('border: none')
            export_logo.setCursor(Qt.PointingHandCursor)
            export_logo.clicked.connect(export_to_excel)
            export_logo.show()

            # Buttons
            add_btn = QPushButton(bottom_button_widget)
            add_btn.setGeometry(650, 18, 60, 25)
            add_btn.setText('ADD')
            add_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                  "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            add_btn.setCursor(Qt.PointingHandCursor)
            add_btn.clicked.connect(add_entry)
            add_btn.show()

            update_btn = QPushButton(bottom_button_widget)
            update_btn.setGeometry(715, 18, 60, 25)
            update_btn.setText('UPDATE')
            update_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                     "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            update_btn.setCursor(Qt.PointingHandCursor)
            update_btn.clicked.connect(update_entry)
            update_btn.show()

            refresh_btn = QPushButton(bottom_button_widget)
            refresh_btn.setGeometry(780, 18, 60, 25)
            refresh_btn.setText('REFRESH')
            refresh_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                      "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            refresh_btn.setCursor(Qt.PointingHandCursor)
            refresh_btn.clicked.connect(lambda: fg_outgoing())
            refresh_btn.show()

            delete_btn = QPushButton(bottom_button_widget)
            delete_btn.setGeometry(845, 18, 60, 25)
            delete_btn.setText('DELETE')
            delete_btn.setStyleSheet("background-color : rgb(240,240,240); border-radius: 5px;"
                                     "border: 1px solid rgb(92, 154, 255); background-color: rgb(194, 232, 255)")
            delete_btn.setCursor(Qt.PointingHandCursor)
            delete_btn.clicked.connect(delete_outgoing)
            delete_btn.setShortcut('Delete')
            delete_btn.show()

        def fg_inventory():

            def get_daily_report():

                # Query for getting the fg inventory
                self.cursor.execute("""
                    WITH incoming_total AS (
                        SELECT lot_number, product_code, SUM(quantity) as total_qty FROM fg_incoming
                        GROUP BY lot_number, product_code

                    ),
                    outgoing_total AS (
                        SELECT lot_number, product_code, SUM(quantity) as total_qty FROM fg_outgoing
                        GROUP BY lot_number, product_code)

                    , lot_location AS (
                        SELECT lot_number, array_agg(location) as location FROM fg_incoming
                        GROUP BY lot_number
                    )
                    , inventory AS	(SELECT a.lot_number, a.product_code,
                            CASE WHEN a.total_qty - b.total_qty IS NULL THEN a.total_qty
                                 ELSE a.total_qty - b.total_qty
                                 END AS total_qty, 
                            c.location, split_part(location[1], ':', 1) as wh
                        FROM incoming_total a
                        LEFT JOIN outgoing_total b ON a.lot_number = b.lot_number
                        JOIN lot_location c ON a.lot_number = c.lot_number)

                    SELECT lot_number, product_code, ROUND(total_qty::numeric, 2), location FROM inventory
                """)

                pass

            def auto_search():  # Auto search in fg inventory whenever the text changes
                self.cursor.execute(f"""
                    WITH inventory AS (
                        WITH incoming_total AS (
                                    SELECT lot_number, product_code, SUM(quantity) as total_qty FROM fg_incoming
                                    GROUP BY lot_number, product_code

                                ),
                                outgoing_total AS (
                                    SELECT lot_number, product_code, SUM(quantity) as total_qty FROM fg_outgoing
                                    GROUP BY lot_number, product_code)

                                , lot_location AS (
                                    SELECT lot_number, array_agg(location) as location FROM fg_incoming
                                    GROUP BY lot_number

                                )

                                , inventory AS	(SELECT a.lot_number, a.product_code,
                                        CASE WHEN a.total_qty - b.total_qty IS NULL THEN a.total_qty
                                             ELSE a.total_qty - b.total_qty
                                             END AS total_qty, 
                                        c.location, split_part(location[1], ':', 1) as wh
                                    FROM incoming_total a
                                    LEFT JOIN outgoing_total b ON a.lot_number = b.lot_number
                                    JOIN lot_location c ON a.lot_number = c.lot_number)

                                SELECT lot_number, product_code, ROUND(total_qty::numeric, 2), location FROM inventory

                    )

                    SELECT * FROM inventory
                    WHERE lot_number ILIKE '%{search_box.text()}%' OR product_code ILIKE '%{search_box.text()}%'


                                                """)
                result = self.cursor.fetchall()

                table_widget.clearContents()
                table_widget.verticalScrollBar().setValue(0)

                for row in result:
                    for column in range(len(row)):
                        item = QTableWidgetItem(str(row[column]))
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                        table_widget.setItem(result.index(row), column, item)

            fg_incoming_btn = QPushButton(self.warehouse_tabs)
            fg_incoming_btn.setGeometry(30, 0, 100, 30)
            fg_incoming_btn.setText("FG INCOMING")
            fg_incoming_btn.setCursor(Qt.PointingHandCursor)
            fg_incoming_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_incoming_btn.setStyleSheet(
                "color: white; border: none; ")
            fg_incoming_btn.clicked.connect(self.warehouse)
            fg_incoming_btn.show()

            fg_outgoing_btn = QPushButton(self.warehouse_tabs)
            fg_outgoing_btn.setGeometry(150, 0, 100, 30)
            fg_outgoing_btn.setText("FG OUTGOING")
            fg_outgoing_btn.setCursor(Qt.PointingHandCursor)
            fg_outgoing_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_outgoing_btn.setStyleSheet(
                "color: white; border: none; ")
            fg_outgoing_btn.clicked.connect(fg_outgoing)
            fg_outgoing_btn.show()

            fg_inventory_btn = QPushButton(self.warehouse_tabs)
            fg_inventory_btn.setGeometry(270, 0, 100, 30)
            fg_inventory_btn.setText("FG INVENTORY")
            fg_inventory_btn.setCursor(Qt.PointingHandCursor)
            fg_inventory_btn.setFont(QtGui.QFont('Berlin Sans FB Demi', 10))
            fg_inventory_btn.setStyleSheet(
                "color: white; border: none; padding-bottom: 5px; border-bottom: 2px solid white;")
            fg_inventory_btn.clicked.connect(fg_inventory)
            fg_inventory_btn.show()

            self.status_border = QWidget(self.warehouse_widget)
            self.status_border.setGeometry(0, 30, 991, 35)
            self.status_border.setStyleSheet('border-bottom: 1px solid rgb(160, 160, 160)')
            self.status_border.show()

            search_box = QLineEdit(self.warehouse_widget)
            search_box.setGeometry(760, 70, 150, 25)
            search_box.setStyleSheet('border: 1px solid rgb(171, 173, 179); background-color: rgb(255, 255, 17);')
            search_box.setPlaceholderText('Lot Number')
            search_box.setFocus()
            search_box.textChanged.connect(auto_search)
            search_box.show()

            search_button = QPushButton(self.warehouse_widget)
            search_button.setGeometry(915, 70, 70, 25)
            search_button.setStyleSheet('border: 1px solid rgb(171, 173, 179)')
            search_button.setText('Search')
            search_button.setShortcut('Return')
            search_button.show()

            masterbatch_checkbox = QCheckBox(self.warehouse_widget)
            masterbatch_checkbox.move(5, 70)

            masterbatch_checkbox.show()

            mb_checkbox_label = QLabel(self.warehouse_widget)
            mb_checkbox_label.setGeometry(22, 70, 85, 10)
            mb_checkbox_label.setText('MASTERBATCH')
            mb_checkbox_label.show()

            drycolor_checkbox = QCheckBox(self.warehouse_widget)
            drycolor_checkbox.move(110, 70)

            drycolor_checkbox.show()

            dc_checkbox_label = QLabel(self.warehouse_widget)
            dc_checkbox_label.setGeometry(125, 70, 85, 10)
            dc_checkbox_label.setText("DRYCOLOR")
            dc_checkbox_label.show()

            title_label = QLabel(self.warehouse_widget)
            title_label.setGeometry(290, 80, 300, 40)
            title_label.setStyleSheet('color : rgb(41, 181, 255)')
            title_label.setFont(QtGui.QFont('Segoe UI Black', 20))
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setText("FG INVENTORY")
            title_label.show()

            title_icon = QLabel(self.warehouse_widget)
            title_icon.setGeometry(550, 80, 40, 40)
            title_icon.setPixmap(QtGui.QIcon('inventory.png').pixmap(40, 40))
            title_icon.show()

            control_num_lbl = QLabel(self.status_border)
            control_num_lbl.setGeometry(3, 12, 100, 10)
            control_num_lbl.setText('Control Number:')
            control_num_lbl.setStyleSheet('border: none')
            control_num_lbl.show()

            control_num_val = QLabel(self.status_border)
            control_num_val.setGeometry(105, 0, 120, 35)
            control_num_val.setFont(QtGui.QFont('Arial Black', 15))
            control_num_val.setStyleSheet('color: rgb(0, 128, 192)')
            control_num_val.show()

            code_label = QLabel(self.status_border)
            code_label.setGeometry(260, 0, 40, 35)
            code_label.setFont(QtGui.QFont('Arial', 8))
            code_label.setText('CODE:')
            code_label.show()

            code_value = QLabel(self.status_border)
            code_value.setGeometry(300, 0, 150, 35)
            code_value.setFont(QtGui.QFont('Arial Black', 15))
            code_value.setStyleSheet('color: rgb(0, 128, 192)')
            code_value.show()

            lot_number_label = QLabel(self.status_border)
            lot_number_label.setGeometry(490, 0, 50, 35)
            lot_number_label.setText('LOT NO :')
            lot_number_label.setFont(QtGui.QFont('Arial', 8))
            lot_number_label.show()

            lot_number_val = QLabel(self.status_border)
            lot_number_val.setGeometry(540, 0, 180, 35)
            lot_number_val.setStyleSheet('color: rgb(0, 128, 192)')
            lot_number_val.setFont(QtGui.QFont('Arial Black', 15))
            lot_number_val.show()

            category_label = QLabel(self.status_border)
            category_label.setGeometry(735, 0, 50, 35)
            category_label.setFont(QtGui.QFont('Arial', 8))
            category_label.setText('Category:')
            category_label.show()

            category_value = QLabel(self.status_border)
            category_value.setGeometry(790, 0, 170, 35)
            category_value.setFont(QtGui.QFont('Arial Black', 15))
            category_value.setStyleSheet('color: rgb(0, 128, 192)')
            category_value.show()

            table_widget = QTableWidget(self.warehouse_widget)
            table_widget.setGeometry(0, 125, 991, 506)
            table_widget.setColumnCount(4)
            table_widget.verticalHeader().setVisible(False)
            table_widget.setHorizontalHeaderLabels(
                ["Lot Number", "Product Code", "Quantity", "Location"])

            table_widget.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
            table_widget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)

            table_widget.setColumnWidth(1, 200)
            table_widget.setColumnWidth(2, 120)
            table_widget.setColumnWidth(3, 400)

            # Query for Getting the Inventory
            self.cursor.execute("""
                WITH incoming_total AS (
                    SELECT lot_number, product_code, SUM(quantity) as total_qty FROM fg_incoming
                    GROUP BY lot_number, product_code

                ),
                outgoing_total AS (
                    SELECT lot_number, product_code, SUM(quantity) as total_qty FROM fg_outgoing
                    GROUP BY lot_number, product_code)

                , lot_location AS (
                    SELECT lot_number, array_agg(location) as location FROM fg_incoming
                    GROUP BY lot_number

                )

                , inventory AS	(SELECT a.lot_number, a.product_code,
                        CASE WHEN a.total_qty - b.total_qty IS NULL THEN a.total_qty
                             ELSE a.total_qty - b.total_qty
                             END AS total_qty, 
                        c.location, split_part(location[1], ':', 1) as wh
                    FROM incoming_total a
                    LEFT JOIN outgoing_total b ON a.lot_number = b.lot_number
                    JOIN lot_location c ON a.lot_number = c.lot_number)

                SELECT lot_number, product_code, ROUND(total_qty::numeric, 2), location FROM inventory

                                """)

            table_widget.setFont(QtGui.QFont('Arial', 10))

            table_widget.horizontalHeader().setStyleSheet("""
                                    QHeaderView::section{
                                        font-weight: bold;
                                        background-color: rgb(41, 181, 255);
                                        color: black;
                                        font-size: 18;
                                    }
                                                """)

            result = self.cursor.fetchall()

            table_widget.setRowCount(len(result))
            # Populating the table
            for row in result:
                for column in range(len(row)):
                    item = QTableWidgetItem(str(row[column]))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    item.setTextAlignment(Qt.AlignCenter)
                    table_widget.setItem(result.index(row), column, item)

            table_widget.show()

            bottom_button_widget = QWidget(self.warehouse_widget)
            bottom_button_widget.setGeometry(0, 631, 991, 43)
            bottom_button_widget.show()

            date1 = QDateEdit(bottom_button_widget)
            date1.setGeometry(80, 18, 100, 25)
            date1.show()

            date2 = QDateEdit(bottom_button_widget)
            date2.setGeometry(200, 18, 100, 25)
            date2.show()

            export_logo = ClickableLabel(bottom_button_widget)
            export_logo.setGeometry(310, 18, 20, 20)
            export_logo.setPixmap(QtGui.QIcon('export.png').pixmap(20, 20))
            export_logo.setStyleSheet('border: none')
            export_logo.setCursor(Qt.PointingHandCursor)
            export_logo.show()

            report_logo = ClickableLabel(bottom_button_widget)
            report_logo.setGeometry(335, 18, 20, 20)
            report_logo.setPixmap(QtGui.QIcon('daily_report.png').pixmap(20, 20))
            report_logo.setStyleSheet('border: none')
            report_logo.clicked.connect(get_daily_report)
            report_logo.setCursor(Qt.PointingHandCursor)

            report_logo.show()

        self.warehouse_widget = QWidget(self.main_widget)
        self.warehouse_widget.setGeometry(0, 0, 991, 751)
        self.warehouse_widget.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.warehouse_widget.show()

        self.warehouse_tabs = QtWidgets.QWidget(self.warehouse_widget)
        self.warehouse_tabs.setGeometry(0, 0, 991, 30)
        self.warehouse_tabs.setStyleSheet(
            'border-bottom: 1px solid rgb(160, 160, 160); background-color: rgb(92, 154, 255)')
        self.warehouse_tabs.show()

        try:
            if self.fgOutgoing_btn_clicked_status == True:
                fg_outgoing()
                self.fgOutgoing_btn_clicked_status = False
                return
            else:
                pass
        except Exception as e:
            print(e)

        fg_incoming()


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    LoginWindow = QtWidgets.QMainWindow()
    ui = Ui_LoginWindow()
    ui.setupUi(LoginWindow)
    LoginWindow.show()
    sys.exit(app.exec_())

